from flask import Flask, render_template, jsonify, request, send_file, session, redirect, url_for, make_response, send_from_directory
import pandas as pd
import data_store
from datetime import datetime, date, timedelta
import os
import json
import threading
import time
import uuid
import traceback
import numpy as np
from collections import defaultdict
import glob
import re
from urllib.parse import unquote, quote
import revision_metrics
import issue_prediction
import due_date_engine
import scoring
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.image import MIMEImage
import base64
from werkzeug.utils import secure_filename
import secrets
from datetime import datetime, timedelta
from datetime import timedelta
from functools import wraps
from datetime import datetime
from apscheduler.schedulers.background import BackgroundScheduler
from apscheduler.triggers.cron import CronTrigger
from functools import wraps
from flask import g, flash, redirect, url_for
from flask_limiter import Limiter
from flask_limiter.util import get_remote_address
import redis
from datetime import datetime, timedelta, date
import math

# ==================== تنظیمات اولیه Flask ====================
app = Flask(__name__)


app.secret_key = 'my-super-secret-key-12345'  # یک رشته تصادفی و مخفی
# تولید یک کلید تصادفی ۳۲ بایتی (فقط یک بار اجرا شود)
app.permanent_session_lifetime = timedelta(days=120)  # 120 روز

# ==================== تنظیمات لاگ و کدینگ ====================
import logging
import sys

# Optional but helpful
os.environ["PYTHONIOENCODING"] = "utf-8"
os.environ["PYTHONUTF8"] = "1"

# Logging setup
logger = logging.getLogger("app_safe_logger")
logger.setLevel(logging.DEBUG)

if not logger.handlers:
    h = logging.StreamHandler()
    h.setFormatter(logging.Formatter("%(asctime)s [%(levelname)s] %(message)s"))
    logger.addHandler(h)

def safe_log(message, level="info"):
    """
    Encoding-safe logger for Windows consoles (cp1256/cp1252/...)
    Never raises UnicodeEncodeError.
    """
    try:
        msg = str(message)
    except Exception:
        msg = "[unprintable message]"

    # Force safe output for problematic terminals
    try:
        safe_msg = msg.encode("ascii", "replace").decode("ascii")
    except Exception:
        safe_msg = "[encoding failed]"

    if level == "error":
        logger.error(safe_msg)
    elif level == "warning":
        logger.warning(safe_msg)
    elif level == "debug":
        logger.debug(safe_msg)
    else:
        logger.info(safe_msg)


dashboard_data_cache = {}
dashboard_cache_lock = threading.Lock()

# اتصال به Redis — protocol=2 یعنی HELLO نمی‌فرستد
redis_client = redis.Redis(
    host='localhost',
    port=6379,
    decode_responses=True,
    protocol=2,  # ✅ مهم
    socket_connect_timeout=3,
)

# تست اتصال (اختیاری ولی مفید)
try:
    redis_client.ping()
    safe_log("Redis connected OK")
except Exception as e:
    safe_log(f"Redis not available: {e}", level="warning")

limiter = Limiter(
    key_func=get_remote_address,
    app=app,
    default_limits=["800 per day", "800 per hour"],
    storage_uri="redis://localhost:6379",
    storage_options={"protocol": 2},  # ✅ مهم
)

def get_redis_client():
    try:
        client = redis.Redis(
            host='localhost', port=6379,
            decode_responses=True, protocol=2,
            socket_connect_timeout=2,
        )
        client.ping()
        return client
    except Exception as e:
        safe_log(f"Redis unavailable, using memory fallback: {e}", level="warning")
        return None

redis_client = get_redis_client()

# ==================== لایه‌ی کش مشترک (داشبورد + اینباکس) ====================
# اگر Redis در دسترس باشد از آن استفاده می‌شود (بین همه‌ی workerها/پردازش‌ها
# مشترک است)؛ در غیر این صورت به دیکشنری حافظه‌ای (dashboard_data_cache) برمی‌گردد.
# کلیدها شامل data_store.get_data_version() هستند، پس با آپلود فایل جدید
# به‌صورت خودکار باطل می‌شوند و نیازی به پاک‌سازی زمان‌محور (TTL) نیست.
CACHE_NAMESPACE = "dash_cache_v3"


def cache_get(key: str):
    full_key = f"{CACHE_NAMESPACE}:{key}"
    if redis_client:
        try:
            raw = redis_client.get(full_key)
            if raw:
                return json.loads(raw)
            return None
        except Exception as e:
            safe_log(f"[cache_get] خطای Redis: {e}", level="warning")
    with dashboard_cache_lock:
        item = dashboard_data_cache.get(full_key)
        return item['data'] if item else None


def cache_set(key: str, value):
    full_key = f"{CACHE_NAMESPACE}:{key}"
    if redis_client:
        try:
            redis_client.set(full_key, json.dumps(value, default=str))
            return
        except Exception as e:
            safe_log(f"[cache_set] خطای Redis: {e}", level="warning")
    with dashboard_cache_lock:
        dashboard_data_cache[full_key] = {'data': value, 'timestamp': time.time()}


def cache_clear_all():
    if redis_client:
        try:
            for k in redis_client.scan_iter(f"{CACHE_NAMESPACE}:*"):
                redis_client.delete(k)
        except Exception as e:
            safe_log(f"[cache_clear_all] خطای Redis: {e}", level="warning")
    with dashboard_cache_lock:
        dashboard_data_cache.clear()


def cache_clear_old(current_version: str):
    """کلیدهای نسخه قدیمی را پاک می‌کند؛ نسخه فعلی دست نخورده می‌ماند."""
    if not current_version:
        return
    if redis_client:
        try:
            for k in redis_client.scan_iter(f"{CACHE_NAMESPACE}:*"):
                if current_version not in k:
                    redis_client.delete(k)
        except Exception as e:
            safe_log(f"[cache_clear_old] خطای Redis: {e}", level="warning")
    with dashboard_cache_lock:
        for k in list(dashboard_data_cache.keys()):
            if current_version not in k:
                dashboard_data_cache.pop(k, None)


if redis_client:
    limiter = Limiter(
        key_func=get_remote_address,
        app=app,
        default_limits=["300 per day", "50 per hour"],
        storage_uri="redis://localhost:6379",
        storage_options={"protocol": 2},
    )
else:
    limiter = Limiter(
        key_func=get_remote_address,
        app=app,
        default_limits=["300 per day", "300 per hour"],
        storage_uri="memory://",  # موقت تا Redis درست شود
    )

from flask import g, session, flash, redirect, url_for
from functools import wraps

ACTIVITY_KEY_PREFIX = "user:last_activity:"
ACTIVITY_TTL_SECONDS = 30 * 24 * 60 * 60  # ۳۰ روز — طبق درخواست کاربر، تاریخچه فقط تا ۱ ماه نگه داشته می‌شود

def set_user_activity(username):
    if not username or not redis_client:
        return
    redis_client.set(
        f"{ACTIVITY_KEY_PREFIX}{username}",
        datetime.now().isoformat(),
        ex=ACTIVITY_TTL_SECONDS
    )

def get_user_activity(username):
    if not redis_client:
        return user_last_activity.get(username)  # fallback
    val = redis_client.get(f"{ACTIVITY_KEY_PREFIX}{username}")
    if not val:
        return None
    try:
        return datetime.fromisoformat(val)
    except Exception:
        return None
    
def check_user_active(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        username = session.get('username')
        if not username:
            flash('لطفاً وارد شوید.', 'error')
            return redirect(url_for('login'))

        # اگر فعلاً مکانیزم active/inactive نداری، همین کافی است.
        # اگر بعداً status کاربر را از دیتابیس خواندی، اینجا چک کن.
        return f(*args, **kwargs)

    return decorated_function

# ==================== وضعیت آنلاین کاربران ====================
# دیکشنری برای ذخیره آخرین زمان فعالیت هر کاربر
user_last_activity = {}
ONLINE_TIMEOUT_SECONDS = 120  # 4 دقیقه

# ==================== تنظیمات احراز هویت ====================
# دیکشنری برای ذخیره کدهای موقت
temp_codes = {}

# لیست کاربران با ایمیل (همان لیست ایمیل‌ها)
# می‌توانید از email_config.json هم استفاده کنید
USERS_EMAILS = {
    # ===== کاربران داخلی (psmco.co) =====
    'احسان برازنده راد': 'e.barazandeh@psmco.co',
    'احسان برازنده راد': 'ehsan.barazande@gmail.com',
    'آرش عیسی زاده': 'eesazadeh@psmco.co',                  
    'سارا صابر': 's.saber@psmco.co',
    'مرجانه اسماعیل زاده': 'm.esmaeilzadeh@psmco.co',          
    'حنانه محمودی': 'h.mahmoudi@psmco.co',                    
    'کیمیا کربلایی محمد': 'k.kmohammad@psmco.co',             
    'نیما معززی': 'n.moazezi@psmco.co',                       
    'قاسم جواد پور': 'gh.javadpour@psmco.co',                 
    'علی خسروی': 'khosravi@psmco.co',
    'احسان راشدی': 'e.rashedi@psmco.co',
    'سید علیرضا حسینی': 'a.hosseini@psmco.co',
    'مهدی اکبری': 'm.akbari@psmco.co',
    'مجید حسینی': 'm.hoseini@psmco.co',
    'عباس امیرآبادی': 'a.amirabadi@psmco.co',
    'علی ملکی': 'a.maleki@psmco.co',
    'علی اردوخانی': 'a.ordoukhani@psmco.co',
    'علیرضا کارگر': 'karegar@psmco.co',
    'محمود موفق': 'm.movafagh@psmco.co',
    'علیرضا عابر': 'a.aber@psmco.co',
    'مریم تاجیک': 'm.tajik@psmco.co',
    'بهاره داداش‌پور': 'b.dadashpour@psmco.co',
    'محسن مبارکی': 'm.mobaraki@psmco.co',
    'محمدرضا وحیدی': 'vahidi@petrotechco.com',

    # ===== کاربران خارجی =====
    'بهنام تاکی': 'behnamtaki56@gmail.com',
    'ناصر لطفی': 'naserlotfi2009@gmail.com',
    'محمد امینی': 'amini.mohammad.eng@mail.ir',

    # ===== کاربران CC =====
    'امیر جهانی': 'a.jahani@psmco.co',
    'نیما مرادخانی': 'n.moradkhani@petrotechco.com',
    'مریم وحیدی': 'm.vahidi@psmco.co'                      
}

# دیکشنری معکوس: ایمیل -> نام کاربری
EMAIL_TO_USERNAME = {v: k for k, v in USERS_EMAILS.items()}

# ==================== نگاشت نام‌های انگلیسی به فارسی ====================
NAME_MAPPING = {
    # نام‌هایی که در گزارشات ظاهر می‌شوند (از To_Name در history)
    'ehsan barazande rad (CTR)': 'احسان برازنده راد',
    'Arash esazadeh (MDS)': 'آرش عیسی زاده',
    'sara saberhosseini (MDJ)': 'سارا صابر',
    'MARJANEH esmaeilzadeh (MDJ)': 'مرجانه اسماعیل زاده',
    'Hannaneh Mahmoudi (MDJ)': 'حنانه محمودی',
    'Kimia Karbalaiimohamad (MDS)': 'کیمیا کربلایی محمد',
    'Nima moazezi (MDJ)': 'نیما معززی',
    'Ghasem Javadpour (Procurement)': 'قاسم جواد پور',
    'Ali Khosravi (MDS)': 'علی خسروی',
    'Ehsan rashedi (MDJ)': 'احسان راشدی',
    'seyed alireza hosseini (MDS)': 'سید علیرضا حسینی',
    'Mahdi Akbari (Engineer)': 'مهدی اکبری',
    'Majid Hoseini (MDJ)': 'مجید حسینی',
    'Abbas Amirabadi (MDS)': 'عباس امیرآبادی',
    'Ali maleki (MDJ)': 'علی ملکی',
    'Ali Ordoukhani (MDS)': 'علی اردوخانی',
    'Alireza karegar (MDS)': 'علیرضا کارگر',
    'Mahmoud salimmovafagh (EM)': 'محمود سلیم موفق',
    'Alireza Aber (EM)': 'علیرضا عابر',
    'maryam tajik (MDJ)': 'مریم تاجیک',
    'bahareh dadashpour (MDJ)': 'بهاره داداش‌پور',
    'Amir Jahani (CEO)': 'امیر جهانی',
    'Nima Moradkhani (Pmo)': 'نیما مرادخانی',
    'Maryam Vahidi (DCC)': 'مریم وحیدی',
    'Naser Lotfi (MDS)': 'ناصر لطفی',
    'Nader Abdi (MDJ)': 'نادر عبدی',
    'Ali Akbarzadeh (MDJ)': 'علی اکبرزاده',
    'MOHAMAD amini (MDS)': 'محمد امینی',
    'Abbass Raiss Shaghaghi (MDS)': 'عباس رییس شقاقی',
    'behnam taki (MDS)': 'بهنام تاکی',
    'Saeed Yari (MDJ)': 'سعید یاری',
    'Mohsen Mobaraki (MDS)': 'محسن مبارکی',

    # اگر نام بدون پسوند هم ظاهر می‌شود
    'Ehsan Barazandeh': 'احسان برازنده راد',
    'Arash esazadeh': 'آرش عیسی زاده',
    'Sara Saber': 'سارا صابر',
    'Marjaneh Esmaeilzadeh': 'مرجانه اسماعیل زاده',
    'Hananeh Mahmoudi': 'حنانه محمودی',
    'Kimia Karbalaei Mohammad': 'کیمیا کربلایی محمد',
    'Nima Moazezi': 'نیما معززی',
    'Ghasem Javadpour': 'قاسم جواد پور',
    'Ali Khosravi': 'علی خسروی',
    'Ehsan Rashedi': 'احسان راشدی',
    'Seyed Alireza Hosseini': 'سید علیرضا حسینی',
    'Mahdi Akbari': 'مهدی اکبری',
    'Majid Hosseini': 'مجید حسینی',
    'Abbas Amirabadi': 'عباس امیرآبادی',
    'Ali Maleki': 'علی ملکی',
    'Ali Ordoukhani': 'علی اردوخانی',
    'Alireza Karegar': 'علیرضا کارگر',
    'Mahmoud Movafagh': 'محمود سلیم موفق',
    'Mahmoud Salim Movafagh': 'محمود سلیم موفق',
    'Alireza Aber': 'علیرضا عابر',
    'Maryam Tajik': 'مریم تاجیک',
    'Behareh Dadashpour': 'بهاره داداش‌پور',
    'Amir Jahani': 'امیر جهانی',
    'Nima Moradkhani': 'نیما مرادخانی',
    'Maryam Vahidi': 'مریم وحیدی',
    'Naser Lotfi': 'ناصر لطفی',
    'Nader Abdi': 'نادر عبدی',
    'Ali Akbarzadeh': 'علی اکبرزاده',
    'MOHAMAD amini': 'محمد امینی',
    'Abbass Raiss Shaghaghi': 'عباس رییس شقاقی',
    'behnam taki': 'بهنام تاکی',
    'Saeed Yari': 'سعید یاری',
    'Mohsen Mobaraki': 'محسن مبارکی'
}


# ==================== تنظیمات ====================
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
UPLOAD_FOLDER = os.path.join(BASE_DIR, 'uploads')
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

from datetime import datetime

def parse_date(value):
    if value is None or value == "":
        raise ValueError("empty date")

    if isinstance(value, datetime):
        return value

    if hasattr(value, "to_pydatetime"):
        return value.to_pydatetime()

    text = str(value).strip()
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%Y/%m/%d", "%d/%m/%Y"):
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            pass

    parsed = pd.to_datetime(text, errors="coerce")
    if pd.notna(parsed):
        return parsed.to_pydatetime()

    raise ValueError(f"Unrecognized date format: {value}")


import os

# ==================== Maintenance Mode ====================
MAINTENANCE_FILE = os.path.join(BASE_DIR, 'maintenance_mode.txt')
MAINTENANCE_IP_WHITELIST = ['127.0.0.1', '10.0.0.171']  # آی‌پی‌های مجاز (اختیاری)
# ==================== حالت تعمیر ====================
MAINTENANCE_MODE = os.path.exists(MAINTENANCE_FILE)
MAINTENANCE_START_TIME = datetime(2026, 8, 9, 7, 0, 0)  # مثال: ۸ آگوست ۲۰۲۶ ساعت ۱۶:۰۰
MAINTENANCE_END_TIME = datetime(2026, 8, 9, 14, 30, 0)  # سال، ماه، روز، ساعت، دقیقه، ثانیه

@app.before_request
def check_maintenance():
    if not MAINTENANCE_MODE:
        return None
    
    # دسترسی ادمین
    if session.get('username') in ['احسان برازنده راد', 'admin']:
        return None
    
    if request.remote_addr in MAINTENANCE_IP_WHITELIST:
        return None
    
    if request.path.startswith('/static/') or request.path.startswith('/chat_uploads/'):
        return None
    
    # ارسال زمان پایان به قالب (با فرمت ISO)
    end_time_str = MAINTENANCE_END_TIME.isoformat()
    return render_template('maintenance.html', end_time=end_time_str), 503

@app.before_request
def check_maintenance():
    # اگر حالت تعمیر غیرفعال است، کاری نکن
    if not MAINTENANCE_MODE:
        return None
    
    # اگر کاربر ادمین است (از Session)، دسترسی بده
    if session.get('username') in ['احسان برازنده راد', 'admin']:
        return None
    
    # اگر آی‌پی در لیست سفید است، دسترسی بده
    if request.remote_addr in MAINTENANCE_IP_WHITELIST:
        return None
    
    # اگر درخواست برای فایل‌های استاتیک یا آپلود چت است
    if request.path.startswith('/static/') or request.path.startswith('/chat_uploads/'):
        return None
    
    # ارسال زمان پایان به قالب
    end_time_str = MAINTENANCE_END_TIME.strftime('%Y-%m-%d %H:%M:%S')
    return render_template('maintenance.html', end_time=end_time_str), 503

# ==================== فایل‌های چت ====================
CHAT_FILE = os.path.join(BASE_DIR, 'chat_messages.json')
IP_MAPPING_FILE = os.path.join(BASE_DIR, 'ip_mapping.json')

UPLOAD_FOLDER_CHAT = os.path.join(BASE_DIR, 'chat_uploads')
os.makedirs(UPLOAD_FOLDER_CHAT, exist_ok=True)
# ==================== تنظیمات ایمیل ====================

EMAIL_CONFIG_FILE = os.path.join(BASE_DIR, 'email_config.json')
EMAIL_TEMPLATE_FILE = os.path.join(BASE_DIR, 'templates', 'email_template.html')
DASHBOARD_INTERNAL_URL = 'http://10.0.0.171:5000'
DASHBOARD_EXTERNAL_URL = 'http://5.160.148.115:5000'
_EMAIL_WEEKDAYS_FA = ['دوشنبه', 'سه‌شنبه', 'چهارشنبه', 'پنجشنبه', 'جمعه', 'شنبه', 'یکشنبه']


def _html_esc(text):
    return (
        str(text or '')
        .replace('&', '&amp;')
        .replace('<', '&lt;')
        .replace('>', '&gt;')
        .replace('"', '&quot;')
    )


def _email_display_date(raw=None):
    dt = datetime.now()
    if raw:
        try:
            dt = datetime.strptime(str(raw)[:10], '%Y-%m-%d')
        except ValueError:
            pass
    return f"{_EMAIL_WEEKDAYS_FA[dt.weekday()]} {dt.strftime('%Y-%m-%d')}"


def _email_hotspot_block(client_docs, contractor_docs):
    pool = list(client_docs or []) + list(contractor_docs or [])
    pool.sort(key=lambda d: int(d.get('days') or 0), reverse=True)
    if not pool:
        return (
            '<table width="100%" cellpadding="0" cellspacing="0" border="0" bgcolor="#e8f5e9" '
            'style="background:#e8f5e9; border-radius:12px; border-right:4px solid #2e7d32;">'
            '<tr><td style="padding:14px 16px; font-size:13px; color:#1b5e20; line-height:1.9;">'
            'خبر خوب: در صف تأخیر کارفرما و پیمانکار مورد داغی نیست. برای دیدن پیشرفت پروژه‌ها و تخته امتیازات وارد داشبورد شوید.'
            '</td></tr></table>'
        )

    top = pool[:3]
    rows = []
    for doc in top:
        kind = 'کارفرما' if doc.get('type') == 'client' else 'پیمانکار'
        kind_color = '#e65100' if doc.get('type') == 'client' else '#c2185b'
        doc_no = _html_esc(doc.get('document_no') or '—')
        project = _html_esc(doc.get('project') or 'نامشخص')
        days = int(doc.get('days') or 0)
        rows.append(
            '<tr>'
            f'<td style="padding:10px 12px; border-bottom:1px solid #eef0f6; font-size:13px; color:#1a237e; font-weight:700; direction:ltr; text-align:left;">{doc_no}</td>'
            f'<td style="padding:10px 12px; border-bottom:1px solid #eef0f6; font-size:12px; color:#546e7a;">{project}</td>'
            f'<td style="padding:10px 12px; border-bottom:1px solid #eef0f6; font-size:12px; color:{kind_color}; font-weight:700; white-space:nowrap;">{kind}</td>'
            f'<td style="padding:10px 12px; border-bottom:1px solid #eef0f6; font-size:13px; color:#c62828; font-weight:800; white-space:nowrap;">{days} روز</td>'
            '</tr>'
        )
    more = len(pool) - len(top)
    footer = ''
    if more > 0:
        footer = (
            f'<tr><td colspan="4" style="padding:12px; text-align:center; font-size:13px; color:#1a237e; font-weight:700;">'
            f'و {more} مدرک تأخیری دیگر فقط روی داشبورد است — نام مسئول و فیلتر دیسیپلین را آنجا ببینید.'
            '</td></tr>'
        )
    else:
        footer = (
            '<tr><td colspan="4" style="padding:12px; text-align:center; font-size:13px; color:#546e7a;">'
            'برای دیدن مسئول هر مدرک و اقدام بعدی، وارد داشبورد شوید.'
            '</td></tr>'
        )
    return (
        '<table width="100%" cellpadding="0" cellspacing="0" border="0" style="border:1px solid #e3e8f5; border-radius:12px; overflow:hidden;">'
        '<tr bgcolor="#f4f7ff">'
        '<th style="padding:8px 12px; font-size:11px; color:#5c6bc0; text-align:right;">شماره مدرک</th>'
        '<th style="padding:8px 12px; font-size:11px; color:#5c6bc0; text-align:right;">پروژه</th>'
        '<th style="padding:8px 12px; font-size:11px; color:#5c6bc0; text-align:right;">طرف</th>'
        '<th style="padding:8px 12px; font-size:11px; color:#5c6bc0; text-align:right;">تأخیر</th>'
        '</tr>'
        + ''.join(rows) + footer +
        '</table>'
    )


def _email_top_project_alert_block(client_docs, contractor_docs):
    """
    یک بنر کوتاه می‌سازد: پروژه‌ای که الان بیشترین تعداد مدرک تأخیری را دارد.
    کاملاً بر اساس داده‌ی واقعی (نه ساختگی) — اگر هیچ تأخیری نباشد، رشته‌ی
    خالی برمی‌گردد و بنر اصلاً رندر نمی‌شود.
    """
    from collections import Counter
    pool = list(client_docs or []) + list(contractor_docs or [])
    if not pool:
        return ''
    counter = Counter()
    for d in pool:
        proj = str(d.get('project') or 'نامشخص').strip() or 'نامشخص'
        counter[proj] += 1
    if not counter:
        return ''
    top_proj, top_count = counter.most_common(1)[0]
    proj_esc = _html_esc(top_proj)
    return (
        '<table width="100%" cellpadding="0" cellspacing="0" border="0" '
        'style="background:#fff3e0; border-radius:12px; border-right:4px solid #ef6c00; margin-top:10px;">'
        '<tr><td style="padding:12px 16px; font-size:13px; color:#e65100; line-height:1.9;">'
        f'📌 بیشترین تأخیر الان مال پروژه «<strong>{proj_esc}</strong>» است — {top_count} مدرک منتظرِ اقدام. '
        'قبل از این‌که بیشتر بشه، روی داشبورد ببینش.'
        '</td></tr></table>'
    )


def _email_headline(overdue_client_n, overdue_contractor_n, progress):
    if overdue_client_n:
        return f"الان {overdue_client_n} مدرک کارفرما تأخیر دارد — جزئیات و مسئول هر کدام روی داشبورد است."
    if overdue_contractor_n:
        return f"{overdue_contractor_n} مدرک پیمانکار در صف پیگیری است. لیست کامل را روی سایت ببینید."
    return f"داشبورد به‌روز شد؛ پیشرفت کلی {progress}٪. رتبه این هفته و پیش‌بینی صدور را از دست ندهید."


def collect_email_digest(integrated=None, update_date=None, extra_content=''):
    integrated = integrated or {}
    stats = integrated.get('stats') or {}
    client = integrated.get('overdue_client') or []
    contractor = integrated.get('overdue_contractor') or []
    hold = integrated.get('hold_docs') or []
    progress = round(float(stats.get('avg_progress') or stats.get('overall_progress') or 0), 1)
    oc = len(client)
    ok = len(contractor)
    total = int(stats.get('total') or 0)
    not_issued = int(stats.get('not_issued') or 0)
    hold_n = len(hold) if hold else int(stats.get('hold') or 0)
    date_str = str(update_date or datetime.now().strftime('%Y-%m-%d'))[:10]
    return {
        'update_date': date_str,
        'display_date': _email_display_date(date_str),
        'headline': _email_headline(oc, ok, progress),
        'total_docs': total,
        'not_issued': not_issued,
        'overall_progress': progress,
        'overdue_client': oc,
        'overdue_contractor': ok,
        'hold_count': hold_n,
        'hotspot_block': _email_hotspot_block(client, contractor),
        'top_project_alert_block': _email_top_project_alert_block(client, contractor),
        'internal_url': DASHBOARD_INTERNAL_URL,
        'external_url': DASHBOARD_EXTERNAL_URL,
        'extra_content': extra_content or '',
        'overdue_client_n': oc,
    }


def fill_email_template(values=None):
    values = values or {}
    if not os.path.exists(EMAIL_TEMPLATE_FILE):
        return None
    with open(EMAIL_TEMPLATE_FILE, 'r', encoding='utf-8') as f:
        html = f.read()
    for key, val in values.items():
        text = '' if val is None else str(val)
        html = html.replace('{{ ' + key + ' }}', text)
        html = html.replace('{{' + key + '}}', text)
    html = re.sub(r'\{\{\s*[\w]+\s*\}\}', '', html)
    return html


def render_update_email(integrated=None, update_date=None, extra_content=''):
    return fill_email_template(collect_email_digest(integrated, update_date, extra_content))

def load_email_config():
    """بارگذاری تنظیمات ایمیل"""
    if os.path.exists(EMAIL_CONFIG_FILE):
        try:
            with open(EMAIL_CONFIG_FILE, 'r', encoding='utf-8') as f:
                return json.load(f)
        except:
            return {}
    return {}

def save_email_config(config):
    """ذخیره تنظیمات ایمیل"""
    with open(EMAIL_CONFIG_FILE, 'w', encoding='utf-8') as f:
        json.dump(config, f, ensure_ascii=False, indent=2)

def get_email_list():
    """دریافت لیست ایمیل‌ها"""
    config = load_email_config()
    return config.get('emails', [])

def add_email(email):
    """افزودن ایمیل به لیست"""
    config = load_email_config()
    if 'emails' not in config:
        config['emails'] = []
    if email not in config['emails']:
        config['emails'].append(email)
        save_email_config(config)
        return True
    return False

def remove_email(email):
    """حذف ایمیل از لیست"""
    config = load_email_config()
    if 'emails' in config and email in config['emails']:
        config['emails'].remove(email)
        save_email_config(config)
        return True
    return False

def send_email(to_emails, subject, html_content, logo_path=None, cc_emails=None):
    """
    ارسال ایمیل به لیست مخاطبین با پشتیبانی از CC
    
    Args:
        to_emails: لیست ایمیل‌های گیرندگان اصلی
        subject: موضوع ایمیل
        html_content: محتوای HTML ایمیل
        logo_path: مسیر لوگو (اختیاری)
        cc_emails: لیست ایمیل‌های CC (اختیاری)
    """
   
    config = load_email_config()
    
    smtp_server = config.get('smtp_server')
    smtp_port = config.get('smtp_port',578)
    sender_email = config.get('sender_email')
    sender_password = config.get('sender_password')
    
    if not all([smtp_server, smtp_port, sender_email]):
        raise Exception("تنظیمات SMTP کامل نیست. لطفاً ابتدا تنظیمات را ذخیره کنید.")
    
    if isinstance(to_emails, str):
        to_emails = [to_emails]
    
    if cc_emails and isinstance(cc_emails, str):
        cc_emails = [cc_emails]
    
    msg = MIMEMultipart('related')
    msg['From'] = sender_email
    msg['To'] = ', '.join(to_emails)
    if cc_emails:
        msg['Cc'] = ', '.join(cc_emails)
    msg['Subject'] = subject
    
    # ایجاد بخش HTML
    msg_alternative = MIMEMultipart('alternative')
    msg.attach(msg_alternative)
    
    # اضافه کردن لوگو
    if logo_path and os.path.exists(logo_path):
        with open(logo_path, 'rb') as f:
            logo_data = f.read()
            logo_base64 = base64.b64encode(logo_data).decode()
            html_content = html_content.replace('cid:logo', f'data:image/png;base64,{logo_base64}')
    
    msg_alternative.attach(MIMEText(html_content, 'html', 'utf-8'))
    
    # ===== ارسال ایمیل =====
    # ترکیب گیرندگان اصلی و CC برای ارسال
    all_recipients = to_emails + (cc_emails if cc_emails else [])
    safe_log(f"📧 تلاش برای اتصال SMTP به {smtp_server}:{smtp_port} ...")
    
    try:
        server = smtplib.SMTP(smtp_server, smtp_port)
        server.starttls()
        if sender_password:
            server.login(sender_email, sender_password)
        server.sendmail(sender_email, all_recipients, msg.as_string())
        server.quit()
        return True
        
    except Exception as e1:
        safe_log(f"⚠️ اتصال اولیه SMTP ({smtp_server}:{smtp_port}) ناموفق: {e1}", level="warning")

        # ===== امتحان پورت‌ها به ترتیب (بدون تکرار پورتی که همین الان تست شد) =====
        ports_to_try = list(dict.fromkeys([smtp_port, 587, 465, 25]))
        last_error = e1

        for port in ports_to_try:
            try:
                if port == 465:
                    # SSL
                    import ssl
                    context = ssl.create_default_context()
                    server = smtplib.SMTP_SSL(smtp_server, port, context=context)
                else:
                    # STARTTLS یا بدون امنیت
                    server = smtplib.SMTP(smtp_server, port)
                    if port != 25:
                        server.starttls()  # برای پورت‌های ۵۸۷ و غیره
                
                if sender_password:
                    server.login(sender_email, sender_password)
                
                server.sendmail(sender_email, all_recipients, msg.as_string())
                server.quit()
                return True
                
            except Exception as e:
                last_error = e
                safe_log(f"⚠️ اتصال با پورت {port} ناموفق: {e}", level="warning")
                continue

        # اگر خطا از نوع «resolve نشدن نام سرور» باشد (DNS)، پیام واضح‌تری بده
        error_text = str(last_error)
        if 'getaddrinfo' in error_text or '11002' in error_text or 'Name or service not known' in error_text:
            raise Exception(
                f"سرور SMTP با آدرس «{smtp_server}» پیدا نشد (خطای DNS). "
                f"این یعنی این سرور الان به «{smtp_server}» دسترسی شبکه ندارد — "
                f"نه پورت‌ها. با nslookup/ping آن را از روی همین سرور تست کنید. "
                f"جزئیات فنی: {error_text}"
            )
        raise Exception(f"همه پورت‌های SMTP به {smtp_server} ناموفق بودند. آخرین خطا: {error_text}")

# ==================== توابع گزارش DISTRIBUTE ====================
def get_distribute_docs():
    """استخراج تمام مدارک با آخرین وضعیت 'Distribute' برای هر شخص"""
    data = load_all_data()
    data = process_master(data)
    data = process_history(data)
    data = process_vendor_master(data)
    data = process_vendor_history(data)
    
    history_df = data.get('history')
    history_cols = data.get('history_cols', {})
    vendor_history_df = data.get('vendor_history')
    vendor_history_cols = data.get('vendor_history_cols', {})
    
    result = {}  # key: persian_name, value: list of docs

    def process_distribute(df, cols, doc_type, master_df, master_cols):
        if df is None or cols is None:
            return
        
        doc_no_col = cols.get('doc_no')
        action_date_col = cols.get('action_date')
        to_name_col = cols.get('to_name')
        from_name_col = cols.get('from_name')
        log_status_col = cols.get('log_status')
        project_col = cols.get('project')
        close_date_col = cols.get('close_date')
        
        if not all([doc_no_col, action_date_col, to_name_col, log_status_col]):
            return
        
        # فیلتر رکوردهای Distribute
        distribute_df = df[df[log_status_col].astype(str).str.strip() == 'Distribute']
        if distribute_df.empty:
            return

        # ===== رفع باگ ۲: فقط رکوردهایی که واقعاً هنوز باز هستن =====
        # خودِ سطر Distribute اگه Ongoing='No' باشه یا Close Date پر شده
        # باشه، یعنی رویژنی که این Distribute توش اتفاق افتاده از قبل بسته
        # شده — پس این مدرک صددرصد از اینباکس خارج شده، فارغ از اینکه
        # پاسخ صریحی (Comment) پیدا کردیم یا نه.
        ongoing_col = cols.get('ongoing')
        if ongoing_col and ongoing_col in distribute_df.columns:
            distribute_df = distribute_df[distribute_df[ongoing_col].astype(str).str.strip() == 'Yes']
        if close_date_col and close_date_col in distribute_df.columns:
            distribute_df = distribute_df[distribute_df[close_date_col].isna()]
        if distribute_df.empty:
            return
        
        # ===== اصلاح: گروه‌بندی بر اساس doc_no + to_name =====
        # برای هر مدرک و هر شخص، آخرین رکورد Distribute را نگه دار
        latest = distribute_df.sort_values(action_date_col).groupby([doc_no_col, to_name_col]).last().reset_index()

        # ===== رفع باگ ۱: حذف مدارکی که از اینباکس خارج شده‌اند =====
        # پاسخ به یک Distribute با یک ردیف Log Status='Comment' ثبت می‌شود که
        # طرفینش برعکسِ ردیف Distribute است (From/To جابه‌جا می‌شوند). اگر چنین
        # ردیفی بعد از تاریخ Distribute برای همان مدرک پیدا شود، یعنی دیگر پاسخ
        # داده شده و نباید در ایمیل «هنوز پاسخ نداده‌اید» ظاهر شود.
        if from_name_col and from_name_col in df.columns and not latest.empty:
            comment_df = df[df[log_status_col].astype(str).str.strip() == 'Comment']
            if not comment_df.empty:
                comment_df = comment_df[[doc_no_col, from_name_col, to_name_col, action_date_col]].copy()
                comment_df.columns = ['_c_doc_no', '_c_from', '_c_to', '_c_date']

                merged = latest.merge(
                    comment_df,
                    left_on=[doc_no_col, to_name_col],
                    right_on=['_c_doc_no', '_c_from'],
                    how='left'
                )
                # پاسخِ معتبر: طرفِ مقابلِ Comment باید همون فرستنده‌ی اصلیِ Distribute باشه
                # و تاریخ Comment باید بعد از تاریخ Distribute باشه
                if from_name_col in merged.columns:
                    answered_mask = (
                        (merged['_c_to'] == merged[from_name_col]) &
                        (merged['_c_date'] > merged[action_date_col])
                    )
                    answered_keys = set(
                        zip(merged.loc[answered_mask, doc_no_col], merged.loc[answered_mask, to_name_col])
                    )
                    if answered_keys:
                        keep_mask = ~latest.apply(
                            lambda r: (r[doc_no_col], r[to_name_col]) in answered_keys, axis=1
                        )
                        latest = latest[keep_mask]
        
        master_doc_no_col = master_cols.get('doc_no')
        master_title_col = master_cols.get('doc_title')
        master_discipline_col = master_cols.get('discipline')
        master_project_col = master_cols.get('project')
        
        for _, row in latest.iterrows():
            doc_no = row[doc_no_col]
            action_date = row[action_date_col]
            to_name = row[to_name_col]
            from_name_raw = row.get(from_name_col) if from_name_col else None
            project = row.get(project_col) if project_col else None
            
            if pd.isna(to_name):
                continue
            
            raw_name = str(to_name).strip()

            def to_persian(raw):
                if not raw:
                    return raw
                raw_str = str(raw).strip()
                # اول: تطبیق دقیق (سریع‌ترین حالت)
                persian = NAME_MAPPING.get(raw_str)
                if persian:
                    return persian
                # دوم: تطبیق بدون حساسیت به بزرگی/کوچکی حروف - چون تاریخچه
                # گاهی همون اسم رو با کیس متفاوت ثبت می‌کنه (مثلاً
                # "mohsen mobaraki (MDS)" به‌جای "Mohsen Mobaraki (MDS)")
                # و تطبیق دقیقِ بالا در این حالت شکست می‌خورد.
                normalized = normalize_name(raw_str)
                for key, val in NAME_MAPPING.items():
                    if normalize_name(key) == normalized:
                        return val
                # سوم: تطبیق مستقیم با جدول کاربران (fallback نهایی)
                for name in USERS_EMAILS.keys():
                    if normalize_name(name) == normalized:
                        return name
                return raw_str

            persian_name = to_persian(raw_name)

            # ===== نام فرستنده (دیسیپلین مبدا) برای CC کردن روی ایمیل =====
            sender_persian_name = None
            if from_name_raw is not None and not pd.isna(from_name_raw):
                sender_persian_name = to_persian(str(from_name_raw).strip())
            
            # دریافت اطلاعات مدرک از master
            doc_title = None
            discipline = None
            if master_df is not None and master_doc_no_col:
                master_row = master_df[master_df[master_doc_no_col].astype(str).str.strip() == doc_no]
                if not master_row.empty:
                    doc_title = master_row.iloc[0].get(master_title_col) if master_title_col else None
                    discipline = master_row.iloc[0].get(master_discipline_col) if master_discipline_col else None
                    if master_project_col and project is None:
                        project = master_row.iloc[0].get(master_project_col)
            
            # محاسبه دیرکرد
            today = datetime.now().date()
            if isinstance(action_date, pd.Timestamp):
                action_date = action_date.date()
            elif isinstance(action_date, datetime):
                action_date = action_date.date()
            days = (today - action_date).days if action_date else 0
            
            if persian_name not in result:
                result[persian_name] = []
            
            result[persian_name].append({
                'document_no': doc_no,
                'document_title': str(doc_title) if doc_title else 'نامشخص',
                'discipline': str(discipline) if discipline else 'نامشخص',
                'distribute_date': action_date.strftime('%Y-%m-%d') if action_date else None,
                'days': days,
                'project': str(project) if project else 'نامشخص',
                'doc_type': doc_type,
                'to_name': persian_name,
                'from_name': sender_persian_name,
                'overdue': days > 3
            })
    
    master_df = data.get('master')
    master_cols = data.get('master_cols', {})
    process_distribute(history_df, history_cols, 'MASTER', master_df, master_cols)
    
    vendor_master_df = data.get('vendor_master')
    vendor_master_cols = data.get('vendor_cols', {})
    process_distribute(vendor_history_df, vendor_history_cols, 'VENDOR', vendor_master_df, vendor_master_cols)
    
    return result

def get_email_for_person(person_name):
    """دریافت ایمیل با استفاده از NAME_MAPPING و نرمال‌سازی"""
    if not person_name:
        return None
    
    person_name = str(person_name).strip()
    
    # 1. تطابق با KEYهای NAME_MAPPING (برای پیدا کردن نام فارسی)
    for eng_name, persian_name in NAME_MAPPING.items():
        if persian_name == person_name:
            if persian_name in USERS_EMAILS:
                return USERS_EMAILS[persian_name]
    
    # 2. تطابق مستقیم با USERS_EMAILS
    if person_name in USERS_EMAILS:
        return USERS_EMAILS[person_name]
    
    # 3. تطابق با نرمال‌سازی
    person_norm = normalize_name(person_name)
    for name, email in USERS_EMAILS.items():
        if normalize_name(name) == person_norm:
            return email
    
    return None

def generate_distribute_report_html(person_name, docs, show_warning=False):
    """تولید HTML جدول برای یک شخص خاص با تفکیک مهندسی/وندور"""
    if not docs:
        return None
    
    # تفکیک مدارک بر اساس نوع
    engineering_docs = [d for d in docs if d['doc_type'] == 'MASTER']
    vendor_docs = [d for d in docs if d['doc_type'] == 'VENDOR']
    
    html = f"""
    <h3 style="color:#1a237e;text-align: right;">📋{person_name} شده برای DISTRIBUTE گزارش مدارک </h3>
    <p style="color:#C62828;">Total Document: {len(docs)}</p>
    """
    
    # ===== بخش مهندسی =====
    if engineering_docs:
        html += build_table_html(engineering_docs, "📄 Engineering Document")
    
    # ===== بخش وندور =====
    if vendor_docs:
        html += build_table_html(vendor_docs, "📦 Vendor Document")
    
    # اگر هیچکدام نبود
    if not engineering_docs and not vendor_docs:

        html += "<p style='color:#999;'>هیچ مدرکی یافت نشد.</p>"

        # ===== اضافه کردن متن هشدار (فقط در صورت درخواست) =====
    if show_warning:
        html += """
        <div style="margin-top:20px; padding:16px 20px; background:#fff3e0; border-radius:8px; border-right:4px solid #e65100; direction:rtl; text-align:right;">
            <p style="margin:0; font-size:14px; color:#4e342e; line-height:2;">
                با سلام و احترام
            </p>
            <p style="margin:8px 0 0 0; font-size:14px; color:#4e342e; line-height:2;">
                براساس جدول پیوست، مدارک زیر برای شما ارسال شده و هنوز پاسخی به دیسیپلین مربوطه داده نشده است. پیرو جلسه مورخ <strong>31/03/1405</strong> مبنی بر لزوم پاسخگویی به مدارک ارسالی از دیسیپلین‌های دیگر حداکثر ظرف <strong>3 روز کاری</strong>، هرچه سریع‌تر نسبت به بررسی و ارسال آن اقدام لازم به عمل آید.
            </p>
            <p style="margin:12px 0 0 0; font-size:13px; color:#bf360c; font-weight:600;">
                <i class="fas fa-exclamation-triangle" style="margin-left:6px;"></i>
                لطفاً در اسرع وقت نسبت به بررسی و پاسخگویی اقدام فرمایید.
            </p>
        </div>
        """
    
    return html


def build_table_html(docs, title):
    """ساخت جدول HTML برای یک گروه از مدارک (با تفکیک پروژه)"""
    if not docs:
        return ""
    
    # گروه‌بندی بر اساس پروژه
    projects = {}
    for doc in docs:
        proj = doc['project'] or 'نامشخص'
        if proj not in projects:
            projects[proj] = []
        projects[proj].append(doc)
    
    html = f"""
    <h4 style='color:#0d47a1; margin-top:16px;'><i class="fas fa-folder-open"></i> {title}</h4>
    """
    
    for proj, proj_docs in projects.items():
        html += f"""
        <h5 style='color:#1a237e; margin:8px 0 4px 0;'><i class="fas fa-folder"></i> Project: {proj}</h5>
        <table style='width:100%; border-collapse:collapse; font-size:13px; direction:rtl; margin-bottom:16px; table-layout:fixed;'>
            <thead style='background:#1a237e; color:white;'>
                <tr>
                    <th style='padding:8px 12px; text-align:center; width:20%;'>Document No.</th>
                    <th style='padding:8px 12px; text-align:center; width:55%;'>Document Title</th>
                    <th style='padding:8px 12px; text-align:center; width:15%;'>DISTRIBUTE Date</th>
                    <th style='padding:8px 12px; text-align:center; width:10%;'>Delay</th>
                </tr>
            </thead>
            <tbody>
        """
        for doc in proj_docs:
            days = doc['days']
            row_style = ""
            if days > 3:
                row_style = "font-weight:bold; color:#c62828; background:#ffebee;"
            elif days > 0:
                row_style = "color:#e65100;"
            
            html += f"""
                <tr style='border-bottom:1px solid #eee; {row_style}'>
                    <td style='padding:6px 12px; direction:ltr; text-align:left; font-family:monospace; white-space:nowrap;'>{doc['document_no']}</td>
                    <td style='padding:6px 12px; white-space:nowrap; overflow:hidden; text-overflow:ellipsis; max-width:400px;' 
                        title='{doc['document_title']}'>{doc['document_title']}</td>
                    <td style='padding:6px 12px; white-space:nowrap;'>{doc['distribute_date']}</td>
                    <td style='padding:6px 12px; text-align:center; white-space:nowrap;'>{days} روز</td>
                </tr>
            """
        html += "</tbody></table>"
    return html

@app.route('/api/distribute-report')
def distribute_report():
    """API برای دریافت گزارش DISTRIBUTE به‌صورت JSON"""
    try:
        result = get_distribute_docs()
        output = {}
        for person, docs in result.items():
            output[person] = {
                'email': get_email_for_person(person),
                'docs': docs,
                'count': len(docs)
            }
        return jsonify({'success': True, 'data': output})
    except Exception as e:
        safe_log(f"❌ خطا در دریافت گزارش DISTRIBUTE: {e}")
        return jsonify({'error': str(e)}), 500

def get_cc_emails():
    """دریافت لیست ایمیل‌های CC از email_config.json"""
    config = load_email_config()
    return config.get('cc_emails', [])

@app.route('/api/send-distribute-report', methods=['POST'])
def send_distribute_report():
    """ارسال گزارش DISTRIBUTE به همه افراد با CC"""
    try:
        current_user = session.get('username')
        admin_names = ['admin', 'مدیر', 'ادمین', 'ehsan', 'barazande']
        is_admin = current_user and current_user.lower() in admin_names
        if not is_admin:
            return jsonify({'error': 'تنها ادمین می‌تواند گزارش را ارسال کند'}), 403
        
        result = get_distribute_docs()
        cc_emails = get_cc_emails()  # دریافت ایمیل‌های CC
        
        sent_count = 0
        failed_count = 0
        report_data = []
        
        for person, docs in result.items():
            email = get_email_for_person(person)
            if not email:
                failed_count += 1
                continue
            
            html_table = generate_distribute_report_html(person, docs, show_warning=False)
            if not html_table:
                continue
            
            subject = f"📋 {person} - شده  DISTRIBUTE گزارش مدارک "
            logo_path = os.path.join(BASE_DIR, 'static', 'logo.png')
            
            template_path = os.path.join(BASE_DIR, 'templates', 'email_template.html')
            if os.path.exists(template_path):
                with open(template_path, 'r', encoding='utf-8') as f:
                    email_template = f.read()
                email_body = email_template.replace('{{ update_date }}', datetime.now().strftime('%Y-%m-%d'))
                email_body = email_body.replace('{{ total_docs }}', str(len(docs)))
                email_body = email_body.replace('{{ not_issued }}', '0')
                email_body = email_body.replace('{{ overall_progress }}', '0')
                email_body = email_body.replace('📊 خلاصه وضعیت:', html_table)
            else:
                email_body = f"""
                <html>
                <body style="font-family: Arial, sans-serif; direction: rtl;">
                    <h2 style="color:#1a237e;text-align: right;">📋 گزارش مدارک DISTRIBUTE شده</h2>
                    <p>{datetime.now().strftime('%Y-%m-%d')}</p>
                    {html_table}
                </body>
                </html>
                """
            
            try:
                # ===== CC کردن فرستنده‌ها (دیسیپلین مبدا) روی همین ایمیل =====
                sender_emails = set()
                for d in docs:
                    sender_name = d.get('from_name')
                    if sender_name:
                        sender_email = get_email_for_person(sender_name)
                        if sender_email and sender_email != email:
                            sender_emails.add(sender_email)
                person_cc = list(set(cc_emails) | sender_emails)

                # ارسال ایمیل با CC
                send_email([email], subject, email_body, logo_path, person_cc)
                sent_count += 1
                report_data.append({
                    'person': person,
                    'email': email,
                    'cc': person_cc,
                    'count': len(docs)
                })
            except Exception as e:
                safe_log(f"❌ خطا در ارسال ایمیل به {person}: {e}")
                failed_count += 1
        
        return jsonify({
            'success': True,
            'message': f'گزارش برای {sent_count} نفر ارسال شد (CC: {len(cc_emails)} نفر)',
            'sent': sent_count,
            'failed': failed_count,
            'cc_count': len(cc_emails),
            'details': report_data
        })
    except Exception as e:
        safe_log(f"❌ خطا در ارسال گزارش DISTRIBUTE: {e}")
        return jsonify({'error': str(e)}), 500

# ==================== تابع ارسال خودکار (استفاده‌شده توسط scheduler) ====================
def send_distribute_report_auto():
    """ارسال خودکار گزارش (بدون نیاز به ادمین)"""

    # ===== قفل توزیع‌شده روی Redis: تضمین می‌کنه صرف‌نظر از اینکه چند
    # پردازش/نسخه از این تابع صدا زده می‌شه (چه به‌خاطر reloader، چه چند
    # پردازش موازی سرور، چه هر دلیل دیگه)، ایمیل واقعاً فقط یک‌بار ارسال بشه.
    # قفل به مدت ۶۰ دقیقه معتبره (خیلی بیشتر از زمان لازم برای اجرای این
    # تابع)، پس اجرای بعدیِ واقعی (چند روز دیگه طبق زمان‌بندی) مشکلی نداره.
    lock_key = "distribute_report_send_lock"
    if redis_client:
        try:
            acquired = redis_client.set(lock_key, "1", nx=True, ex=3600)
            if not acquired:
                safe_log("⏭️ اجرای گزارش DISTRIBUTE رد شد — قفل توسط اجرای دیگری گرفته شده (جلوگیری از ارسال تکراری)")
                return
        except Exception as e:
            safe_log(f"⚠️ خطا در گرفتن قفل Redis برای گزارش DISTRIBUTE: {e} — ادامه بدون قفل", level="warning")

    safe_log("🚀 شروع اجرای خودکار گزارش DISTRIBUTE...")
    with app.app_context():
        try:
            result = get_distribute_docs()
            safe_log(f"📊 تعداد افراد در نتیجه: {len(result)}")
            
            if not result:
                safe_log("⚠️ هیچ مدرک DISTRIBUTE یافت نشد، ایمیل ارسال نمی‌شود")
                return
            
            sent_count = 0
            failed_count = 0
            cc_emails = get_cc_emails()
            
            for person, docs in result.items():
                email = get_email_for_person(person)
                safe_log(f"👤 {person} -> ایمیل: {email} (تعداد مدارک: {len(docs)})")
                
                if not email:
                    safe_log(f"⚠️ ایمیل برای {person} پیدا نشد، رد شد")
                    failed_count += 1
                    continue
                
                html_table = generate_distribute_report_html(person, docs)
                if not html_table:
                    safe_log(f"⚠️ جدول HTML برای {person} ساخته نشد")
                    failed_count += 1
                    continue
                
                subject = f"📋 گزارش هفتگی مدارک DISTRIBUTE شده - {person}"
                logo_path = os.path.join(BASE_DIR, 'static', 'logo.png')

                # ===== CC کردن فرستنده‌ها (دیسیپلین مبدا) روی همین ایمیل =====
                # تا فرستنده هم بدونه که پیگیری شده و منتظر پاسخ گیرنده‌ست
                sender_emails = set()
                for d in docs:
                    sender_name = d.get('from_name')
                    if sender_name:
                        sender_email = get_email_for_person(sender_name)
                        if sender_email and sender_email != email:
                            sender_emails.add(sender_email)
                person_cc = list(set(cc_emails) | sender_emails)
                
                try:
                    send_email([email], subject, html_table, logo_path, person_cc)
                    safe_log(f"✅ ایمیل برای {person} ارسال شد (CC فرستنده‌ها: {len(sender_emails)})")
                    sent_count += 1
                except Exception as e:
                    safe_log(f"❌ خطا در ارسال ایمیل به {person}: {e}")
                    failed_count += 1
            
            safe_log(f"📊 گزارش خودکار DISTRIBUTE: {sent_count} موفق، {failed_count} ناموفق")
            
        except Exception as e:
            safe_log(f"❌ خطا در ارسال خودکار گزارش: {e}")
            import traceback
            safe_log(traceback.format_exc())

@app.route('/api/distribute-report-preview')
def distribute_report_preview():
    """پیش‌نمایش HTML گزارش DISTRIBUTE"""
    try:
        result = get_distribute_docs()
        
        html = """
        <!DOCTYPE html>
        <html lang="fa" dir="rtl">
        <head>
            <meta charset="UTF-8">
            <meta name="viewport" content="width=device-width, initial-scale=1.0">
            <title>پیش‌نمایش گزارش DISTRIBUTE</title>
            <link rel="stylesheet" href="https://cdnjs.cloudflare.com/ajax/libs/font-awesome/6.4.0/css/all.min.css">
            <style>
                body { font-family: 'Segoe UI', Tahoma, sans-serif; background: #f0f2f5; padding: 20px; direction: rtl; }
                .container { max-width: 1200px; margin: 0 auto; background: white; padding: 24px; border-radius: 12px; box-shadow: 0 2px 12px rgba(0,0,0,0.08); }
                h1 { color: #1a237e; border-bottom: 2px solid #eee; padding-bottom: 12px; }
                h2 { color: #0d47a1; margin-top: 24px; }
                .summary { background: #e3f2fd; padding: 12px 20px; border-radius: 8px; margin-bottom: 20px; }
                table { width: 100%; border-collapse: collapse; font-size: 13px; margin-bottom: 20px; table-layout:fixed; }
                thead { background: #1a237e; color: white; }
                th { padding: 10px 14px; text-align: right; }
                td { padding: 8px 14px; border-bottom: 1px solid #eee; }
                .overdue { font-weight: bold; color: #c62828; background: #ffebee; }
                .warning { color: #e65100; }
                .doc-type-header { background: #e8eaf6; font-weight: bold; }
                .project-header { background: #f5f5f5; font-weight: 600; }
                .badge { display: inline-block; padding: 2px 10px; border-radius: 12px; font-size: 11px; font-weight: 600; }
                .badge-master { background: #bbdefb; color: #0d47a1; }
                .badge-vendor { background: #c8e6c9; color: #1b5e20; }
            </style>
        </head>
        <body>
            <div class="container">
                <h1><i class="fas fa-file-alt"></i> گزارش مدارک DISTRIBUTE شده</h1>
                <div class="summary">
                    <p><strong>تاریخ گزارش:</strong> """ + datetime.now().strftime('%Y-%m-%d %H:%M') + """</p>
                    <p><strong>تعداد افراد:</strong> """ + str(len(result)) + """</p>
                </div>
        """
        
        for person, docs in result.items():
            email = get_email_for_person(person)
            email_display = f"<span style='color:#666; font-size:12px;'>({email})</span>" if email else "<span style='color:#999; font-size:12px;'>(ایمیل یافت نشد)</span>"
            
            html += f"""
                <h2><i class="fas fa-user"></i> {person} {email_display}</h2>
                <p style='color:#666;'>تعداد مدارک: {len(docs)}</p>
            """
            
            # تفکیک مدارک مهندسی و وندور
            engineering = [d for d in docs if d['doc_type'] == 'MASTER']
            vendor = [d for d in docs if d['doc_type'] == 'VENDOR']
            
            if engineering:
                html += build_table_html(engineering, "📄 Engineering Document")
            
            if vendor:
                html += build_table_html(vendor, "📦 Vendor Document")
            
            if not engineering and not vendor:
                html += "<p style='color:#999;'>هیچ مدرکی یافت نشد.</p>"
            
            html += "<hr style='border: 1px dashed #ddd; margin: 24px 0;'>"
        
        html += """
            </div>
        </body>
        </html>
        """
        return html
        
    except Exception as e:
        safe_log(f"❌ خطا در پیش‌نمایش گزارش: {e}")
        return f"<h3 style='color:#c62828;'>خطا: {e}</h3>"


def build_table_html(docs, title):
    """ساخت جدول HTML برای یک گروه از مدارک"""
    if not docs:
        return ""
    
    # گروه‌بندی بر اساس پروژه
    projects = {}
    for doc in docs:
        proj = doc['project'] or 'نامشخص'
        if proj not in projects:
            projects[proj] = []
        projects[proj].append(doc)
    
    html = f"""
    <h4 style='color:#0d47a1; margin-top:16px;'><i class="fas fa-folder-open"></i> {title}</h4>
    """
    
    for proj, proj_docs in projects.items():
        html += f"""
        <h5 style='color:#1a237e; margin:8px 0 4px 0;'><i class="fas fa-folder"></i> Project: {proj}</h5>
        <table>
            <thead>
                <tr>
                    <th style="width:20%;">Document No.</th>
                    <th style='width:50%;text-align:right;'>Document Title</th>
                    <th style="width:10%;">DISTRIBUTE Date</th>
                    <th style="width:10%;">Delay</th>
                    <th style="width:10%;">Discipline</th>
                </tr>
            </thead>
            <tbody>
        """
        for doc in proj_docs:
            days = doc['days']
            row_class = "overdue" if days > 3 else ("warning" if days > 0 else "")
            date_display = doc['distribute_date'] or 'نامشخص'
            
            html += f"""
                <tr class="{row_class}">
                    <td style="direction:ltr; text-align:left; font-family:monospace;">{doc['document_no']}</td>
                    <td>{doc['document_title']}</td>
                    <td>{date_display}</td>
                    <td style="text-align:center;">{days} days</td>
                    <td>{doc['discipline']}</td>
                </tr>
            """
        html += "</tbody></table>"
    
    return html
# ==================== فایل‌های چت خصوصی ====================
PRIVATE_CHAT_DIR = os.path.join(BASE_DIR, 'private_chats')
os.makedirs(PRIVATE_CHAT_DIR, exist_ok=True)

def get_private_chat_file(user1, user2):
    """دریافت مسیر فایل چت خصوصی بین دو کاربر"""
    # نام فایل بر اساس حروف الفبا مرتب شود تا همیشه یکسان باشد
    users = sorted([user1.lower(), user2.lower()])
    filename = f"chat_{users[0]}_{users[1]}.json"
    return os.path.join(PRIVATE_CHAT_DIR, filename)

def load_private_messages(user1, user2):
    """بارگذاری پیام‌های خصوصی بین دو کاربر"""
    filepath = get_private_chat_file(user1, user2)
    if os.path.exists(filepath):
        try:
            with open(filepath, 'r', encoding='utf-8') as f:
                return json.load(f)
        except:
            return []
    return []

def save_private_message(user1, user2, message):
    """ذخیره پیام خصوصی بین دو کاربر"""
    # اگر پیام سیستمی است، تاریخ امروز را ذخیره کن
    if message.get('is_system'):
        message['system_date'] = datetime.now().strftime('%Y-%m-%d')
    
    messages = load_private_messages(user1, user2)
    messages.append(message)
    if len(messages) > 500:
        messages = messages[-500:]
    filepath = get_private_chat_file(user1, user2)
    with open(filepath, 'w', encoding='utf-8') as f:
        json.dump(messages, f, ensure_ascii=False, indent=2)

def clean_old_system_messages():
    """حذف پیام‌های سیستمی که مربوط به روزهای قبل هستند"""
    today = datetime.now().strftime('%Y-%m-%d')
    cleaned_count = 0
    
    # بررسی همه فایل‌های چت
    for filename in os.listdir(PRIVATE_CHAT_DIR):
        if not filename.endswith('.json'):
            continue
        
        filepath = os.path.join(PRIVATE_CHAT_DIR, filename)
        try:
            with open(filepath, 'r', encoding='utf-8') as f:
                messages = json.load(f)
            
            # فیلتر کردن پیام‌های سیستمی قدیمی
            new_messages = []
            for msg in messages:
                if msg.get('is_system'):
                    # اگر تاریخ پیام با امروز برابر نیست، حذف کن
                    if msg.get('system_date') != today:
                        cleaned_count += 1
                        continue
                new_messages.append(msg)
            
            # اگر تغییری کرده بود، ذخیره کن
            if len(new_messages) != len(messages):
                with open(filepath, 'w', encoding='utf-8') as f:
                    json.dump(new_messages, f, ensure_ascii=False, indent=2)
                    
        except Exception as e:
            safe_log(f"⚠️ خطا در پاک‌سازی فایل {filename}: {e}")
    
    if cleaned_count > 0:
        safe_log(f"🧹 {cleaned_count} پیام سیستمی قدیمی حذف شد")
    
    return cleaned_count

def update_activity(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        username = session.get('username')
        if username:
            set_user_activity(username)  # ✅ Redis
        return f(*args, **kwargs)
    return decorated_function

# ==================== توابع کمکی ====================
def normalize_doc_type_filter(value):
    value = (value or 'همه').strip().lower()

    if value in {'مهندسی', 'engineering', 'master'}:
        return {'MASTER', 'مهندسی', 'ENGINEERING'}
    if value in {'وندور', 'vendor'}:
        return {'VENDOR', 'وندور'}
    return None


def resolve_baseline_date(all_dates, latest_date, period_days, ref_date=None):
    """
    Pick baseline date near (ref_date - period_days), but never equal to latest_date.
    """
    if not all_dates or latest_date is None:
        return None

    if ref_date is None:
        ref_date = datetime.now().date()

    target_date = ref_date - timedelta(days=period_days)
    c = find_closest_date(all_dates, target_date)
    if c is None:
        return None

    if c == latest_date:
        earlier = [d for d in all_dates if d < latest_date]
        return max(earlier) if earlier else None
    return c


def safe_change(current_value, old_value):
    try:
        c = float(current_value or 0)
        o = float(old_value or 0)
        return round(c - o, 2)
    except Exception:
        return 0.0


def calculate_project_average_progress(docs, projects, target_date=None):
    """
    محاسبه میانگین ساده پیشرفت پروژه‌ها
    
    Args:
        docs: لیست مدارک
        projects: لیست پروژه‌ها
        target_date: تاریخ هدف (اختیاری)
    
    Returns:
        float: میانگین ساده پیشرفت
    """
    project_progress_list = []
    
    for proj in projects:
        # فیلتر مدارک پروژه
        if target_date:
            # اگر تاریخ هدف مشخص شده، فقط مدارک قبل از آن تاریخ را در نظر بگیر
            proj_docs = []
            for doc in docs:
                if doc.get('project') == proj:
                    doc_date = doc.get('date')
                    if doc_date is not None and not pd.isna(doc_date):
                        if isinstance(doc_date, pd.Timestamp):
                            doc_date = doc_date.date()
                        elif isinstance(doc_date, datetime):
                            doc_date = doc_date.date()
                        if doc_date <= target_date:
                            proj_docs.append(doc)
            
            # گرفتن آخرین نسخه هر مدرک در تاریخ هدف
            latest_old = {}
            for d in proj_docs:
                doc_no = d.get('document_no')
                if not doc_no:
                    continue
                doc_date = d.get('date')
                if doc_date is None or pd.isna(doc_date):
                    continue
                if isinstance(doc_date, pd.Timestamp):
                    doc_date = doc_date.date()
                elif isinstance(doc_date, datetime):
                    doc_date = doc_date.date()
                if doc_no not in latest_old or doc_date > latest_old[doc_no]['date']:
                    latest_old[doc_no] = d
                    latest_old[doc_no]['date'] = doc_date
            
            proj_docs = list(latest_old.values())
        else:
            # بدون تاریخ هدف، از همه مدارک استفاده کن
            proj_docs = [d for d in docs if d.get('project') == proj]
        
        if proj_docs:
            proj_progress = calculate_adjusted_progress(proj_docs)
            project_progress_list.append(proj_progress)
    
    if project_progress_list:
        return round(sum(project_progress_list) / len(project_progress_list), 2)
    return 0

def get_progress_at_date_fixed_weight(docs, target_date, current_weights):
    """
    Fast version:
    محاسبه پیشرفت وزنی تا تاریخ target_date با وزن‌های ثابت current (typed-key).
    key format: TYPE::DOC_NO
    """
    if not docs or not current_weights:
        return 0.0

    # --- local bindings for speed ---
    _to_datetime = pd.to_datetime
    _isna = pd.isna
    _float = float
    _get_adj = get_adjusted_progress
    _weights_get = current_weights.get

    def to_ts(v):
        t = _to_datetime(v, errors='coerce')
        return t if not _isna(t) else pd.NaT

    target_ts = to_ts(target_date)
    if _isna(target_ts):
        return 0.0

    # latest_by_key[typed_key] = (timestamp, doc)
    latest_by_key = {}

    for d in docs:
        if not isinstance(d, dict):
            continue

        doc_no = str(d.get('document_no', '')).strip()
        if not doc_no:
            continue

        doc_type = str(d.get('doc_type', '')).strip().upper() or 'MASTER'
        tk = f"{doc_type}::{doc_no}"

        # اگر اصلاً وزن current برای این سند نداریم، از همین اول رد کن (صرفه‌جویی مهم)
        w0 = _weights_get(tk, 0)
        try:
            w0 = _float(w0 or 0)
        except Exception:
            w0 = 0.0
        if _isna(w0) or w0 <= 0:
            continue

        dts = to_ts(d.get('date'))
        if _isna(dts):
            dts = to_ts(d.get('comment_date'))
        if _isna(dts):
            dts = to_ts(d.get('issued_date'))
        if _isna(dts) or dts > target_ts:
            continue

        prev = latest_by_key.get(tk)
        if (prev is None) or (dts > prev[0]):
            latest_by_key[tk] = (dts, d)

    if not latest_by_key:
        return 0.0

    total_weight = 0.0
    total_weighted_progress = 0.0

    for tk, (_, doc) in latest_by_key.items():
        w = _weights_get(tk, 0)
        try:
            w = _float(w or 0)
        except Exception:
            w = 0.0
        if _isna(w) or w <= 0:
            continue

        p = _get_adj(doc)
        try:
            p = _float(0 if p is None else p)
        except Exception:
            p = 0.0
        if _isna(p):
            p = 0.0

        # clamp
        if p < 0.0:
            p = 0.0
        elif p > 100.0:
            p = 100.0

        total_weight += w
        total_weighted_progress += (p * w)  # هنوز درصدی

    if total_weight <= 0:
        return 0.0

    # چون p را 0..100 گرفتیم، تقسیم مستقیم کافیست
    return total_weighted_progress / total_weight


def calculate_project_average_progress(docs, projects, target_date=None):
    """
    محاسبه میانگین ساده پیشرفت پروژه‌ها
    """
    project_progress_list = []
    
    for proj in projects:
        if target_date:
            proj_docs = []
            for doc in docs:
                if doc.get('project') == proj:
                    doc_date = doc.get('date')
                    if doc_date is not None and not pd.isna(doc_date):
                        if isinstance(doc_date, pd.Timestamp):
                            doc_date = doc_date.date()
                        elif isinstance(doc_date, datetime):
                            doc_date = doc_date.date()
                        if doc_date <= target_date:
                            proj_docs.append(doc)
            
            latest_old = {}
            for d in proj_docs:
                doc_no = d.get('document_no')
                if not doc_no:
                    continue
                doc_date = d.get('date')
                if doc_date is None or pd.isna(doc_date):
                    continue
                if isinstance(doc_date, pd.Timestamp):
                    doc_date = doc_date.date()
                elif isinstance(doc_date, datetime):
                    doc_date = doc_date.date()
                if doc_no not in latest_old or doc_date > latest_old[doc_no]['date']:
                    latest_old[doc_no] = d
                    latest_old[doc_no]['date'] = doc_date
            
            proj_docs = list(latest_old.values())
        else:
            proj_docs = [d for d in docs if d.get('project') == proj]
        
        if proj_docs:
            proj_progress = calculate_adjusted_progress(proj_docs)
            project_progress_list.append(proj_progress)
    
    if project_progress_list:
        return round(sum(project_progress_list) / len(project_progress_list), 2)
    return 0


def calculate_discipline_average_progress(docs, discipline_name, target_date=None):
    """
    محاسبه میانگین ساده پیشرفت یک دیسیپلین
    """
    if target_date:
        disc_docs = []
        for doc in docs:
            if doc.get('discipline') == discipline_name:
                doc_date = doc.get('date')
                if doc_date is not None and not pd.isna(doc_date):
                    if isinstance(doc_date, pd.Timestamp):
                        doc_date = doc_date.date()
                    elif isinstance(doc_date, datetime):
                        doc_date = doc_date.date()
                    if doc_date <= target_date:
                        disc_docs.append(doc)
        
        latest_old = {}
        for d in disc_docs:
            doc_no = d.get('document_no')
            if not doc_no:
                continue
            doc_date = d.get('date')
            if doc_date is None or pd.isna(doc_date):
                continue
            if isinstance(doc_date, pd.Timestamp):
                doc_date = doc_date.date()
            elif isinstance(doc_date, datetime):
                doc_date = doc_date.date()
            if doc_no not in latest_old or doc_date > latest_old[doc_no]['date']:
                latest_old[doc_no] = d
                latest_old[doc_no]['date'] = doc_date
        
        disc_docs = list(latest_old.values())
    else:
        disc_docs = [d for d in docs if d.get('discipline') == discipline_name]
    
    if disc_docs:
        return calculate_adjusted_progress(disc_docs)
    return 0

def normalize_progress(progress):
    if pd.isna(progress):
        return 0
    if isinstance(progress, (int, float)):
        if progress > 1:
            return progress / 100
        else:
            return progress
    return 0

def get_responsible_category(responsible):
    """
    تشخیص دسته‌بندی Responsible
    """
    if responsible is None or pd.isna(responsible):
        return 'unknown'
    
    resp = str(responsible).lower().strip()
    
    # صادر نشده
    if 'waiting for first issue' in resp:
        return 'not_issued'
    
    # دست کارفرما
    if 'waiting client approve' in resp or 'waiting client comments' in resp:
        return 'with_customer'
    
    # تایید شده
    if 'finished' in resp:
        return 'approved'
    
    return 'other'

def get_adjusted_progress(doc):
    """
    تنظیم پیشرفت بر اساس وضعیت مدارک
    
    منطق جدید:
    - برای مدارک مهندسی: اعمال تعدیل ۱۰۰% → ۹۵% (با شرایط)
    - برای مدارک وندور: بدون تعدیل (پیشرفت واقعی)
    """
    progress = doc.get('doc_progress', 0)
    if pd.isna(progress):
        progress = 0
    
    doc_type = str(doc.get('doc_type', '')).strip().upper()
    
    # ===== مدارک وندور: بدون تعدیل =====
    if doc_type == 'VENDOR':
        return progress
    
    # ===== مدارک مهندسی: اعمال تعدیل =====
    doc_no = doc.get('document_no', 'N/A')
    
    # شرط 1: Deleted = No
    deleted = str(doc.get('deleted', '')).lower().strip()
    if deleted == 'yes':
        return progress
    
    # شرط 2: Hold = No
    hold = str(doc.get('hold', '')).lower().strip()
    if hold == 'yes':
        return progress
    
    # شرط 3: Responsible = Contractor
    responsible = doc.get('responsible', '')
    if responsible is None or pd.isna(responsible):
        return progress
    
    resp_str = str(responsible).lower().strip()
    if 'contractor' not in resp_str:
        return progress
    
    # شرط 4: Progress = 100
    if progress != 100:
        return progress
    
    # شرط 5: Comment = Approved
    comment = doc.get('comment', '')
    if comment is None or pd.isna(comment):
        return progress
    
    comment_str = str(comment).lower().strip()
    if comment_str != 'approved':
        return progress
    
    # همه شرط‌ها برقرار است → تعدیل به 95
    return 95

def calculate_progress(docs):
    """
    محاسبه پیشرفت وزنی مدارک با تفکیک نوع
    """
    eng_docs = [d for d in docs if str(d.get('doc_type', '')).strip().upper() == 'MASTER']
    ven_docs = [d for d in docs if str(d.get('doc_type', '')).strip().upper() == 'VENDOR']
    
    eng_progress = calculate_adjusted_progress(eng_docs) if eng_docs else 0.0
    ven_progress = calculate_vendor_progress(ven_docs) if ven_docs else 0.0
    
    total_docs = len(eng_docs) + len(ven_docs)
    if total_docs > 0:
        return round(
            (eng_progress * len(eng_docs) + ven_progress * len(ven_docs)) / total_docs,
            2
        )
    return 0.0

def calculate_adjusted_progress(docs):
    """
    محاسبه پیشرفت وزنی با اعمال تعدیل (۱۰۰% → ۹۵% برای مدارک مهندسی واجد شرایط)
    """
    if not docs:
        return 0.0

    adjusted_docs = []
    for doc in docs:
        adj_doc = doc.copy()
        adj_doc['doc_progress'] = get_adjusted_progress(doc)   # تعدیل تکی
        adjusted_docs.append(adj_doc)

    # محاسبه وزنی (با fallback به وزن ۱)
    total_weighted = 0.0
    total_weight = 0.0

    for doc in adjusted_docs:
        doc_type = str(doc.get('doc_type', '')).strip().upper()

        # انتخاب وزن مناسب
        if doc_type == 'MASTER':
            weight = doc.get('eng_weight', 0)
        elif doc_type == 'VENDOR':
            weight = doc.get('weight', 0)
        else:
            weight = 0

        # اگر وزن صفر یا نامعتبر بود، وزن ۱ در نظر گرفته شود
        try:
            weight = float(weight)
            if pd.isna(weight) or weight <= 0:
                weight = 1.0
        except:
            weight = 1.0

        progress = doc.get('doc_progress', 0)
        try:
            progress = float(progress)
            if pd.isna(progress):
                progress = 0.0
        except:
            progress = 0.0

        total_weighted += weight * progress
        total_weight += weight

    if total_weight == 0:
        return 0.0

    return round((total_weighted / total_weight), 2)

def fix_document_number(doc_no):
    """
    اصلاح شماره مدرک
    تبدیل 85205-000-IN-YCC-TP-002 به 000-85205-IN-YCC-TP-002
    """
    if doc_no is None or pd.isna(doc_no):
        return None
    
    doc_no = str(doc_no).strip()
    
    # الگوی شماره مدرک: 000-85205-IN-YCC-TP-002
    # اگر با 3 رقم شروع شد و سپس خط تیره و 5 رقم، اصلاح کن
    pattern = r'^(\d{3})-(\d{5})-([A-Z]{2,3}-[A-Z]{3,4}-[A-Z]{2,3}-\d{3})$'
    match = re.match(pattern, doc_no)
    if match:
        first = match.group(1)   # 85205
        second = match.group(2)  # 000
        rest = match.group(3)    # IN-YCC-TP-002
        return f"{second}-{first}-{rest}"
    
    return doc_no

def extract_person_docs(df, cols, person_key_norm, project_filter=None, source_type=''):
    """
    اسناد مربوط به یک فرد خاص را از دیتافریم تاریخچه استخراج می‌کند.
    
    Args:
        df: دیتافریم تاریخچه (مهندسی یا وندور)
        cols: دیکشنری نگاشت نام ستون‌ها
        person_key_norm: نام نرمال‌شده فرد
        project_filter: فیلتر پروژه (اختیاری)
        source_type: نوع منبع ('مهندسی' یا 'وندور')
        
    Returns:
        دیکشنری با کلید doc_no و مقدار اطلاعات سند
    """
    # استخراج نام ستون‌ها
    to_name_col = cols.get('to_name')
    log_status_col = cols.get('log_status')
    ongoing_col = cols.get('ongoing')
    close_date_col = cols.get('close_date')
    doc_no_col = cols.get('doc_no')
    discipline_col = cols.get('discipline')
    action_date_col = cols.get('action_date')
    project_col = cols.get('project')
    doc_title_col = cols.get('doc_title')
    
    if not all([to_name_col, log_status_col, ongoing_col, close_date_col, 
                doc_no_col, discipline_col, action_date_col, project_col]):
        return {}
    
    entries = {}
    
    for _, row in df.iterrows():
        # بررسی doc_no معتبر
        doc_no = row[doc_no_col]
        if pd.isna(doc_no):
            continue
        
        # فیلتر پروژه
        if project_filter:
            project = row.get(project_col)
            if pd.isna(project) or str(project).strip() != project_filter:
                continue
        
        # بررسی log_status
        log_status = row.get(log_status_col)
        if pd.isna(log_status) or str(log_status).strip() not in ['Assign', 'Issue', 'Distribute']:
            continue
        
        # بررسی ongoing (شرط یکسان)
        ongoing = row.get(ongoing_col)
        ongoing_str = '' if pd.isna(ongoing) else str(ongoing).strip().lower()
        if ongoing_str not in ['yes', 'بله', 'true', '']:  # خالی هم مجاز است
            continue
        
        # بررسی close_date خالی
        close_date = row.get(close_date_col)
        if not pd.isna(close_date) and str(close_date).strip():
            continue
        
        # بررسی تطابق نام فرد
        to_name = row.get(to_name_col)
        if pd.isna(to_name):
            continue
        
        norm_to_name = normalize_name(to_name)
        if norm_to_name != person_key_norm:
            continue
        
        # بررسی action_date معتبر
        action_date_val = row.get(action_date_col)
        if pd.isna(action_date_val):
            continue
        
        try:
            action_date = pd.to_datetime(action_date_val).date()
        except:
            continue
        
        # ذخیره اطلاعات
        doc_info = {
            'doc_no': str(doc_no).strip(),
            'doc_title': str(row.get(doc_title_col, '')).strip() if not pd.isna(row.get(doc_title_col)) else '',
            'discipline': str(row.get(discipline_col, '')).strip() if not pd.isna(row.get(discipline_col)) else '',
            'log_status': str(log_status).strip(),
            'action_date': action_date,
            'project': str(row.get(project_col, '')).strip() if not pd.isna(row.get(project_col)) else '',
            'source': source_type
        }
        
        # نگه داشتن جدیدترین action_date برای هر doc_no
        existing = entries.get(doc_info['doc_no'])
        if not existing or action_date > existing['action_date']:
            entries[doc_info['doc_no']] = doc_info
    
    return entries

def convert_excel_date(val):
    if pd.isna(val):
        return pd.NaT
    if isinstance(val, (int, float)):
        try:
            return pd.to_datetime('1899-12-30') + pd.to_timedelta(val, unit='D')
        except:
            return pd.NaT
    try:
        return pd.to_datetime(val, errors='coerce')
    except:
        return pd.NaT


def find_column(df, possible_names):
    if df is None:
        return None
    for name in possible_names:
        if name in df.columns:
            return name
    for col in df.columns:
        for name in possible_names:
            if col.lower() == name.lower():
                return col
    return None

def find_closest_date(all_dates, target_date):
    """
    نزدیک‌ترین تاریخ منطقی برای مقایسه را پیدا می‌کند:
    1) ترجیح با آخرین تاریخ <= target_date
    2) اگر وجود نداشت، نزدیک‌ترین تاریخ موجود
    """
    if not all_dates:
        return None

    valid_dates = sorted(set(all_dates))

    before_or_equal = [dt for dt in valid_dates if dt <= target_date]
    if before_or_equal:
        return max(before_or_equal)

    return min(valid_dates, key=lambda dt: abs((dt - target_date).days))


def get_progress_at_date(docs, target_date):
    """
    محاسبه پیشرفت در یک تاریخ مشخص با استفاده از پیشرفت تعدیل شده
    ✅ همه مدارک (حتی حذف شده) محاسبه می‌شوند
    """
    target_date = pd.to_datetime(target_date).date()
    
    # دیکشنری برای نگهداری آخرین نسخه هر مدرک در تاریخ هدف
    latest_at_target = {}
    
    for doc in docs:
        doc_no = doc.get('document_no')
        if not doc_no:
            continue
            
        doc_date = doc.get('date')
        if doc_date is None or pd.isna(doc_date):
            continue
            
        # تبدیل به تاریخ
        if isinstance(doc_date, pd.Timestamp):
            doc_date = doc_date.date()
        elif isinstance(doc_date, datetime):
            doc_date = doc_date.date()
        
        # ✅ فقط مدارکی که تاریخ آنها <= تاریخ هدف است
        if doc_date <= target_date:
            # نگهداری جدیدترین نسخه برای هر مدرک در تاریخ هدف
            if doc_no not in latest_at_target:
                latest_at_target[doc_no] = doc
                latest_at_target[doc_no]['_date'] = doc_date
            else:
                existing_date = latest_at_target[doc_no].get('_date')
                if existing_date is None or doc_date > existing_date:
                    latest_at_target[doc_no] = doc
                    latest_at_target[doc_no]['_date'] = doc_date
    
    # ✅ اگر هیچ مدرکی در تاریخ هدف نبود، 0 برگردان
    if not latest_at_target:
        return 0
    
    # اعمال get_adjusted_progress روی هر مدرک قبل از محاسبه
    adjusted_docs = []
    for doc in latest_at_target.values():
        adjusted_doc = doc.copy()
        # حذف فیلد موقت _date
        if '_date' in adjusted_doc:
            del adjusted_doc['_date']
        adjusted_doc['doc_progress'] = get_adjusted_progress(doc)
        adjusted_docs.append(adjusted_doc)
    
    # محاسبه پیشرفت با مقادیر تعدیل شده
    return calculate_progress(adjusted_docs)

def normalize_package(name):
    if pd.isna(name):
        return ''
    name = str(name).strip()
    # حذف فاصله‌های اضافی
    name = re.sub(r'\s+', ' ', name)
    # حذف کاراکترهای اضافی (فقط حروف، اعداد، فاصله و خط تیره)
    name = re.sub(r'[^a-zA-Z0-9\u0600-\u06FF\s\-]', '', name)
    # تبدیل به حروف کوچک
    return name.lower().strip()

def normalize_vendor(name):
    if pd.isna(name):
        return ''
    name = str(name).strip()
    # حذف فاصله‌های اضافی
    name = re.sub(r'\s+', ' ', name)
    # حذف کاراکترهای اضافی
    name = re.sub(r'[^a-zA-Z0-9\u0600-\u06FF\s\-\.]', '', name)
    # تبدیل به حروف کوچک
    return name.lower().strip()

def normalize_name(name):
    if pd.isna(name):
        return ''
    name = str(name).strip()
    name = re.sub(r'\s*\([^)]*\)', '', name)
    return name.lower().strip()

def normalize_project(name):
    if pd.isna(name):
        return ''
    name = str(name).strip()
    # حذف فاصله‌های اضافی
    name = re.sub(r'\s+', ' ', name)
    # جایگزینی کاراکترهای خاص
    name = name.replace('&', 'and')
    name = name.replace('/', ' and ')
    # حذف کاراکترهای غیرمجاز
    name = re.sub(r'[^a-zA-Z0-9\u0600-\u06FF\s]', '', name)
    return name.lower().strip()

def clean_for_json(obj):
    if isinstance(obj, dict):
        # حذف کلیدهای None و پردازش مقادیر
        new_dict = {}
        for k, v in obj.items():
            if k is None:  # کلید None را نادیده بگیر
                continue
            new_dict[k] = clean_for_json(v)
        return new_dict
    elif isinstance(obj, list):
        return [clean_for_json(item) for item in obj]
    elif isinstance(obj, set):
        return [clean_for_json(item) for item in obj]
    elif isinstance(obj, float) and np.isnan(obj):
        return None
    elif isinstance(obj, (pd.Timestamp, datetime, date)):
        return obj.strftime('%Y-%m-%d') if not pd.isna(obj) else None
    elif pd.isna(obj):
        return None
    else:
        return obj

# ==================== بارگذاری و پردازش فایل‌ها ====================
def find_files():
    files = {'master': None, 'history': None, 'vendor_history': None, 'vendor_master': None, 'person': None}
    search_dirs = [BASE_DIR, UPLOAD_FOLDER]
    name_patterns = {
        'master': ['master.xlsx', 'Master.xlsx', 'Master*.xlsx'],
        'history': ['history.xlsx', 'History.xlsx', 'History*.xlsx', 'historylog.xlsx', 'History Log.xlsx'],
        'vendor_history': ['vendor_history.xlsx', 'vendor-historylog.xlsx', 'Vendor History.xlsx', 'vendor*.xlsx'],
        'vendor_master': ['vendor_master.xlsx', 'Vendor Master.xlsx', 'vendor_master*.xlsx'],
        'person': ['person.xlsx', 'Person.xlsx', 'person name.xlsx', 'Person Name.xlsx']
    }
    for key, patterns in name_patterns.items():
        for search_dir in search_dirs:
            if not os.path.exists(search_dir):
                continue
            for pattern in patterns:
                matches = glob.glob(os.path.join(search_dir, pattern))
                if matches:
                    files[key] = matches[0]
                    break
            if files[key]:
                break
    return files

def load_all_data():
    # ===== پاک‌سازی پیام‌های سیستمی قدیمی =====
    clean_old_system_messages()

    excel_files_map = find_files()
    data_store.initialize(excel_files_map)

    data = {}
    data_keys = ["person", "master", "history", "vendor_master", "vendor_history"]

    for key in data_keys:
        df = data_store.get_dataframe(key)
        if df is not None:
            data_dict_key = "persons" if key == "person" else key
            data[data_dict_key] = df

    return data

def clear_dashboard_cache():
    cache_clear_all()
    safe_log(" Dashboard cache cleared.")

def process_master(data):
    if 'master' not in data or data['master'] is None:
        return data
    df = data['master'].copy()

    col = {
        'project': find_column(df, ['Project']),
        'date': find_column(df, ['Date']),
        'discipline': find_column(df, ['Discipline']),
        'doc_no': find_column(df, ['Document No.', 'Document No']),
        'doc_title': find_column(df, ['Document Title', 'Document Title']),
        'category': find_column(df, ['Category']),
        'responsible': find_column(df, ['Responsible']),
        'issued_date': find_column(df, ['Issued Date', 'IssuedDate', 'Issue Date']),
        'comment_date': find_column(df, ['Comment Date', 'CommentDate']),
        'comment': find_column(df, ['Comment']),
        'eng_weight': find_column(df, ['Eng.Weight', 'Eng Weight']),
        'doc_progress': find_column(df, ['Document Progress']),
        'deleted': find_column(df, ['Deleted']),
        'hold': find_column(df, ['Hold']),
        'hold_cause': find_column(df, ['Hold.1.Hold Cause', 'Hold Cause']),
        'progress': find_column(df, ['Progress'])
    }
    
    for c in [col['date'], col['issued_date'], col['comment_date']]:
        if c:
            df[c] = df[c].apply(convert_excel_date)
    
    if col['responsible']:
        def get_status(resp):
            if pd.isna(resp):
                return 'نامشخص'
            resp = str(resp).lower()
            if 'finished' in resp:
                return 'تایید شده'
            elif 'client' in resp:
                return 'دست کارفرما'
            elif 'contractor' in resp:
                if 'first issue' in resp:
                    return 'صادر نشده'
                else:
                    return 'در انتظار پاسخ'
            return 'سایر'
        df['وضعیت'] = df[col['responsible']].apply(get_status)
    if col['eng_weight']:
        df[col['eng_weight']] = pd.to_numeric(df[col['eng_weight']], errors='coerce')
    if col['doc_progress']:
        df[col['doc_progress']] = pd.to_numeric(df[col['doc_progress']], errors='coerce')
    data['master'] = df
    data['master_cols'] = col
    return data

def process_vendor_master(data):
    if 'vendor_master' not in data or data['vendor_master'] is None:
        return data
    df = data['vendor_master'].copy()
    
    col = {
        'package': find_column(df, ['Package Name']),
        'vendor': find_column(df, ['Vendor Name']),
        'discipline': find_column(df, ['Discipline']),
        'doc_no': find_column(df, ['Document No.', 'Document No']),
        'doc_title': find_column(df, ['Document Title']),
        'category': find_column(df, ['Category']),
        'responsible': find_column(df, ['Responsible']),
        'issued_date': find_column(df, ['Issued Date', 'Issue Date']),
        'comment_date': find_column(df, ['Comment Date']),
        'comment': find_column(df, ['Comment']),
        'doc_progress': find_column(df, ['Document Progress']),
        'deleted': find_column(df, ['Deleted']),
        'date': find_column(df, ['Date'])
    }
    
    
    for c in [col['date'], col['issued_date'], col['comment_date']]:
        if c:
            df[c] = df[c].apply(convert_excel_date)
    
    # ✅ محاسبه وزن بر اساس مدارک فعال (حذف نشده)
    if col['package'] and col['deleted']:
        # فقط مدارک فعال را در نظر بگیر
        active_mask = df[col['deleted']].astype(str).str.lower() != 'yes'
        active_df = df[active_mask]
        
        # تعداد مدارک فعال در هر پکیج
        package_counts = active_df[col['package']].value_counts()
        
        def calculate_weight(row):
            pkg = row.get(col['package'])
            if pd.isna(pkg):
                return 0
            # وزن بر اساس تعداد مدارک فعال در پکیج
            count = package_counts.get(pkg, 0)
            if count == 0:
                return 0
            return 1.0 / count
        
        df['Weight'] = df.apply(calculate_weight, axis=1)
    elif col['package']:
        # اگر ستون deleted وجود نداشت، همه مدارک را در نظر بگیر
        package_counts = df[col['package']].value_counts()
        df['Weight'] = df[col['package']].apply(lambda x: 1.0 / package_counts.get(x, 1) if package_counts.get(x, 0) > 0 else 0)
    
    if col['responsible']:
        def get_status(resp):
            if pd.isna(resp):
                return 'نامشخص'
            resp = str(resp).lower()
            if 'finished' in resp:
                return 'تایید شده'
            elif 'client' in resp:
                return 'دست کارفرما'
            elif 'contractor' in resp:
                if 'first issue' in resp:
                    return 'صادر نشده'
                else:
                    return 'در انتظار پاسخ'
            return 'سایر'
        df['وضعیت'] = df[col['responsible']].apply(get_status)
    
    if col['doc_progress']:
        df[col['doc_progress']] = pd.to_numeric(df[col['doc_progress']], errors='coerce')
    
    data['vendor_master'] = df
    data['vendor_cols'] = col
    return data

def process_history(data):
    if 'history' not in data or data['history'] is None:
        return data
    df = data['history'].copy()
    
    col = {
        'project': find_column(df, ['Project']),
        'date': find_column(df, ['Date']),
        'discipline': find_column(df, ['Discipline']),
        'doc_no': find_column(df, ['Document No.', 'Document No']),
        'doc_title': find_column(df, ['Document Title']),
        'action_date': find_column(df, ['Action Date']),
        'from_name': find_column(df, ['From Name', 'From_Name', 'From name', 'From']),
        'to_name': find_column(df, ['To Name']),
        'log_status': find_column(df, ['Log Status']),
        'category': find_column(df, ['Category']),
        'ongoing': find_column(df, ['Ongoing']),
        'close_date': find_column(df, ['Close Date']),
        'comment_date': find_column(df, ['Comment Date'])
    }
    
    for c in [col['action_date'], col['close_date'], col['comment_date'], col['date']]:
        if c:
            df[c] = df[c].apply(convert_excel_date)
    data['history'] = df
    data['history_cols'] = col
    return data

def calculate_overdue_docs(docs, data, person_df=None):
    today = datetime.now().date()
    overdue_list = []
    
    client_docs_checked = 0
    client_with_issued_date = 0
    client_overdue_calculated = 0
    
    person_names = set()
    if person_df is not None and not person_df.empty:
        person_col = find_column(person_df, ['Person Name', 'Name', 'Full Name'])
        if person_col:
            for name in person_df[person_col].dropna():
                name_str = str(name).strip()
                norm = normalize_name(name_str)
                if norm:
                    person_names.add(norm)
    
    for doc in docs:
        # ===== برای مدارک وندور، فقط کارفرما (client) بررسی شود =====
        if doc.get('doc_type') == 'VENDOR':
            responsible = doc.get('responsible', '')
            if pd.isna(responsible):
                continue
            resp_lower = str(responsible).lower()
            if 'client' not in resp_lower:
                continue  # پیمانکار وندور را رد کن (جداگانه محاسبه می‌شود)
        
        if str(doc.get('hold', '')).lower() == 'yes':
            continue
        if str(doc.get('deleted', '')).lower() == 'yes':
            continue
        
        responsible = doc.get('responsible', '')
        if pd.isna(responsible):
            continue
        
        resp_lower = str(responsible).lower()
        
        # بررسی بسته شدن
        is_closed = False
        if 'contractor' in resp_lower:
            to_name = doc.get('to_name', '')
            if to_name and not pd.isna(to_name):
                to_name_norm = normalize_name(to_name)
                if to_name_norm not in person_names:
                    continue
            
            if 'history' in data and data['history'] is not None:
                hist = data['history']
                cols = data.get('history_cols', {})
                if cols.get('doc_no') and cols.get('close_date'):
                    doc_history = hist[hist[cols['doc_no']] == doc.get('document_no')]
                    if not doc_history.empty:
                        last_close = doc_history[cols['close_date']].iloc[-1]
                        if pd.notna(last_close):
                            is_closed = True
        if is_closed:
            continue
        
        overdue_days = 0
        doc_type = None
        
        if 'contractor' in resp_lower:
            to_name = doc.get('to_name', '')
            if to_name and not pd.isna(to_name):
                to_name_norm = normalize_name(to_name)
                if to_name_norm not in person_names:
                    continue
            
            comment_date = doc.get('comment_date')
            if pd.notna(comment_date):
                if isinstance(comment_date, pd.Timestamp):
                    comment_date = comment_date.date()
                days_since_comment = (today - comment_date).days
                overdue_days = days_since_comment - 7
                if overdue_days > 0:
                    doc_type = 'contractor'
                else:
                    overdue_days = 0
        
        elif 'client' in resp_lower:
            client_docs_checked += 1
            issued_date = doc.get('issued_date')
            if pd.notna(issued_date):
                client_with_issued_date += 1
                if isinstance(issued_date, pd.Timestamp):
                    issued_date = issued_date.date()
                days_since_issued = (today - issued_date).days
                overdue_days = days_since_issued - 14
                if overdue_days > 0:
                    client_overdue_calculated += 1
                    doc_type = 'client'
                else:
                    overdue_days = 0
        
        if overdue_days > 0 and doc_type is not None:
            overdue_list.append({
                'document_no': doc.get('document_no'),
                'document_title': doc.get('document_title'),
                'discipline': doc.get('discipline'),
                'responsible': responsible,
                'days': overdue_days,
                'project': doc.get('project'),
                'type': doc_type,
                'over_150': overdue_days > 150,
                'doc_type': doc.get('doc_type', 'UNKNOWN')  # ✅ اضافه شد
            })
    
   
    return overdue_list

def calculate_vendor_overdue(vendor_history_df, vendor_history_cols, person_df):
    """
    محاسبه مدارک دیرکرد وندور از vendor_history
    
    منطق صحیح:
        - From Name: نباید در Person باشد (deselect)
        - To Name: باید در Person باشد (select)
        - Close Date: خالی باشد
        - Action Date: وجود داشته باشد و از امروز گذشته باشد
        - دیرکرد = Today - Action Date
    """
    if vendor_history_df is None or vendor_history_df.empty:
        safe_log("⚠️ vendor_history خالی است")
        return []
    
    today = datetime.now().date()
    overdue_list = []
    
    # ساخت مجموعه نام‌های Person
    person_names = set()
    if person_df is not None and not person_df.empty:
        person_col = find_column(person_df, ['Person Name', 'Name', 'Full Name'])
        if person_col:
            for name in person_df[person_col].dropna():
                norm = normalize_name(name)
                if norm:
                    person_names.add(norm)
    
    if not person_names:
        safe_log("⚠️ لیست Person خالی است، دیرکرد وندور محاسبه نمی‌شود")
        return []
    
    # ستون‌های مورد نیاز
    from_name_col = vendor_history_cols.get('from_name')
    to_name_col = vendor_history_cols.get('to_name')
    close_date_col = vendor_history_cols.get('close_date')
    action_date_col = vendor_history_cols.get('action_date')
    doc_no_col = vendor_history_cols.get('doc_no')
    doc_title_col = vendor_history_cols.get('doc_title')
    discipline_col = vendor_history_cols.get('discipline')
    project_col = vendor_history_cols.get('project')
    
    if not all([from_name_col, to_name_col, close_date_col, action_date_col, doc_no_col]):
        safe_log("⚠️ ستون‌های مورد نیاز در vendor_history پیدا نشد")
        return []
    
    matched_count = 0
    for idx, row in vendor_history_df.iterrows():
        # ===== شرط 1: Close Date خالی =====
        close_date = row.get(close_date_col)
        if pd.notna(close_date) and str(close_date).strip() != '' and str(close_date).strip() != 'NaT':
            continue
        
        # ===== شرط 2: From Name (deselect) - نباید در Person باشد =====
        from_name = row.get(from_name_col)
        if pd.isna(from_name):
            continue
        from_norm = normalize_name(from_name)
        if from_norm in person_names:
            # در Person است → حذف (deselect)
            continue
        
        # ===== شرط 3: To Name (select) - باید در Person باشد =====
        to_name = row.get(to_name_col)
        if pd.isna(to_name):
            continue
        to_norm = normalize_name(to_name)
        if to_norm not in person_names:
            # در Person نیست → حذف
            continue
        
        # ===== شرط 4: Action Date =====
        action_date = row.get(action_date_col)
        if pd.isna(action_date):
            continue
        if isinstance(action_date, pd.Timestamp):
            action_date = action_date.date()
        elif isinstance(action_date, datetime):
            action_date = action_date.date()
        
        days = (today - action_date).days
        if days <= 0:
            continue
        
        matched_count += 1
        
        doc_no = str(row.get(doc_no_col)).strip() if not pd.isna(row.get(doc_no_col)) else 'N/A'
        doc_title = str(row.get(doc_title_col)).strip() if doc_title_col and not pd.isna(row.get(doc_title_col)) else '-'
        import org_structure
        discipline = org_structure.canonicalize_discipline(row.get(discipline_col) if discipline_col else '-')
        project = str(row.get(project_col)).strip() if project_col and not pd.isna(row.get(project_col)) else 'نامشخص'
        
        overdue_list.append({
            'document_no': doc_no,
            'document_title': doc_title,
            'discipline': discipline,
            'responsible': 'وندور',
            'days': days,
            'project': project,
            'type': 'vendor',
            'over_150': days > 150,
            'action_date': action_date.strftime('%Y-%m-%d'),
            'doc_type': 'VENDOR'  # ✅ اضافه شد
        })
    
    
    return overdue_list

def to_ts(v):
    t = pd.to_datetime(v, errors='coerce')
    return t if pd.notna(t) else pd.NaT

def calculate_inbox_stats(
    history_df,
    history_cols,
    vendor_history_df,
    vendor_history_cols,
    person_df,
    project_filter=None
):
    """
    نسخه سریع‌تر calculate_inbox_stats
    - بدون iterrows در هسته پردازش
    - dedup با sort + drop_duplicates
    - خروجی سازگار با نسخه قبلی
    """
    import org_structure

    # -----------------------------
    # 1) person map
    # -----------------------------
    person_display_map = {}
    if person_df is not None and not person_df.empty:
        person_col = find_column(person_df, ['Person Name', 'Name', 'Full Name'])
        if person_col:
            s = person_df[person_col].dropna().astype(str).str.strip()
            for name_str in s:
                norm = normalize_name(name_str)
                if norm:
                    person_display_map[norm] = name_str

    if not person_display_map:
        return {}

    valid_statuses = {'Assign', 'Issue', 'Distribute'}
    false_like = {'no', 'خیر', 'false', '0', 'n'}

    def process_source(df, cols, source_type):
        if df is None or cols is None or df.empty:
            return pd.DataFrame(columns=['person_key', 'doc_no', 'discipline', 'action_date', 'log_status', 'source'])

        to_name_col = cols.get('to_name')
        log_status_col = cols.get('log_status')
        ongoing_col = cols.get('ongoing')
        close_date_col = cols.get('close_date')
        doc_no_col = cols.get('doc_no')
        discipline_col = cols.get('discipline')
        action_date_col = cols.get('action_date')
        project_col = cols.get('project')

        if not to_name_col or not log_status_col or not doc_no_col:
            return pd.DataFrame(columns=['person_key', 'doc_no', 'discipline', 'action_date', 'log_status', 'source'])

        # فقط ستون‌های لازم
        needed_cols = [c for c in [
            to_name_col, log_status_col, ongoing_col, close_date_col,
            doc_no_col, discipline_col, action_date_col, project_col
        ] if c is not None and c in df.columns]

        x = df[needed_cols].copy()

        # rename به اسامی استاندارد
        rename_map = {
            to_name_col: 'to_name',
            log_status_col: 'log_status',
            doc_no_col: 'doc_no',
        }
        if ongoing_col: rename_map[ongoing_col] = 'ongoing'
        if close_date_col: rename_map[close_date_col] = 'close_date'
        if discipline_col: rename_map[discipline_col] = 'discipline'
        if action_date_col: rename_map[action_date_col] = 'action_date'
        if project_col: rename_map[project_col] = 'project'
        x = x.rename(columns=rename_map)

        # doc_no
        x = x[x['doc_no'].notna()].copy()
        x['doc_no'] = x['doc_no'].astype(str).str.strip()
        x = x[x['doc_no'] != '']

        # project filter
        if project_filter and 'project' in x.columns:
            norm_filter = normalize_project(project_filter)
            x = x[x['project'].notna()].copy()
            x['project_norm'] = x['project'].map(normalize_project)
            x = x[x['project_norm'] == norm_filter]

        # status filter
        x = x[x['log_status'].notna()].copy()
        x['log_status'] = x['log_status'].astype(str).str.strip()
        x = x[x['log_status'].isin(valid_statuses)]

        # ongoing filter
        if 'ongoing' in x.columns:
            og = x['ongoing'].astype(str).str.strip().str.lower()
            x = x[~((x['ongoing'].notna()) & (og.isin(false_like)))]

        # close_date filter
        if 'close_date' in x.columns:
            x = x[x['close_date'].isna()]

        # to_name + person membership
        x = x[x['to_name'].notna()].copy()
        x['to_name'] = x['to_name'].astype(str).str.strip()
        x['person_key'] = x['to_name'].map(normalize_name)
        x = x[x['person_key'].isin(set(person_display_map.keys()))]

        # action_date
        if 'action_date' in x.columns:
            x['action_date'] = x['action_date'].apply(to_ts)
        else:
            x['action_date'] = pd.NaT

        if 'discipline' not in x.columns:
            x['discipline'] = None

        # dedup: آخرین رکورد برای هر (person_key, doc_no)
        x = x.sort_values(['person_key', 'doc_no', 'action_date'], ascending=[True, True, True])
        x = x.drop_duplicates(subset=['person_key', 'doc_no'], keep='last')

        x['source'] = source_type

        return x[['person_key', 'doc_no', 'discipline', 'action_date', 'log_status', 'source']]

    eng = process_source(history_df, history_cols, 'مهندسی')
    ven = process_source(vendor_history_df, vendor_history_cols, 'وندور')

    all_entries = pd.concat([eng, ven], ignore_index=True)
    if all_entries.empty:
        return {}

    # -----------------------------
    # 2) آمار کلی
    # -----------------------------
    # شمارش status
    status_counts = (
        all_entries
        .pivot_table(
            index='person_key',
            columns='log_status',
            values='doc_no',
            aggfunc='count',
            fill_value=0
        )
        .reset_index()
    )

    for col in ['Assign', 'Issue', 'Distribute']:
        if col not in status_counts.columns:
            status_counts[col] = 0

    # شمارش by_source + status
    src_status = (
        all_entries
        .groupby(['person_key', 'source', 'log_status'])
        .size()
        .reset_index(name='cnt')
    )

    # total by source
    src_total = (
        all_entries
        .groupby(['person_key', 'source'])
        .size()
        .reset_index(name='total')
    )

    # disciplines
    disc_df = all_entries.copy()
    disc_df['discipline'] = disc_df['discipline'].map(lambda v: org_structure.canonicalize_discipline(v))
    disc_df.loc[disc_df['discipline'].isin(['', 'nan', 'None']), 'discipline'] = 'نامشخص'
    disciplines_map = (
        disc_df.groupby('person_key')['discipline']
        .apply(lambda s: sorted(set(s.tolist())))
        .to_dict()
    )

    # sources list
    sources_map = (
        all_entries.groupby('person_key')['source']
        .apply(lambda s: sorted(set(s.tolist())))
        .to_dict()
    )

    # -----------------------------
    # 3) مونتاژ خروجی
    # -----------------------------
    stats = {}
    person_keys = sorted(all_entries['person_key'].unique().tolist())

    # lookup table برای سرعت
    status_lookup = status_counts.set_index('person_key')[['Assign', 'Issue', 'Distribute']].to_dict('index')

    src_total_lookup = {}
    for _, r in src_total.iterrows():
        pk = r['person_key']
        src = r['source']
        src_total_lookup.setdefault(pk, {'مهندسی': 0, 'وندور': 0})
        src_total_lookup[pk][src] = int(r['total'])

    src_status_lookup = {}
    for _, r in src_status.iterrows():
        pk, src, st, cnt = r['person_key'], r['source'], r['log_status'], int(r['cnt'])
        src_status_lookup.setdefault(pk, {}).setdefault(src, {'Assign': 0, 'Issue': 0, 'Distribute': 0})
        src_status_lookup[pk][src][st] = cnt

    for pk in person_keys:
        by_source = {
            'مهندسی': {'total': 0, 'Assign': 0, 'Issue': 0, 'Distribute': 0},
            'وندور': {'total': 0, 'Assign': 0, 'Issue': 0, 'Distribute': 0},
        }

        # total per source
        if pk in src_total_lookup:
            by_source['مهندسی']['total'] = src_total_lookup[pk].get('مهندسی', 0)
            by_source['وندور']['total'] = src_total_lookup[pk].get('وندور', 0)

        # status per source
        if pk in src_status_lookup:
            for src in ['مهندسی', 'وندور']:
                st_map = src_status_lookup[pk].get(src, {})
                for st in ['Assign', 'Issue', 'Distribute']:
                    by_source[src][st] = st_map.get(st, 0)

        assign = int(status_lookup.get(pk, {}).get('Assign', 0))
        issue = int(status_lookup.get(pk, {}).get('Issue', 0))
        distribute = int(status_lookup.get(pk, {}).get('Distribute', 0))
        total = by_source['مهندسی']['total'] + by_source['وندور']['total']

        stats[pk] = {
            'display_name': person_display_map.get(pk, pk),
            'total': int(total),
            'Assign': assign,
            'Issue': issue,
            'Distribute': distribute,
            'disciplines': disciplines_map.get(pk, ['نامشخص']),
            'sources': sources_map.get(pk, []),
            'by_source': by_source
        }

    return stats


# ==================== ادغام داده‌ها ====================
def _row_get(row, col, default=None):
    if not col:
        return default
    val = row.get(col, default)
    try:
        if val is None or pd.isna(val):
            return default
    except (ValueError, TypeError):
        pass
    return val


def integrate_data(data):
    import org_structure
    integrated = {
        'projects': [],
        'disciplines': [],
        'master_docs': [],
        'vendor_docs': [],
        'stats': {},
        'inbox_stats': {},
        'package_stats': {},
        'hold_docs': [],
        'overdue_docs': [],
        'overdue_client': [],
        'overdue_contractor': [],
        'overdue_contractor_all': [],
        'contractor_over_150_count': 0,
        'total_contractor_overdue': 0,
        'deleted_count': 0
    }

    # ---- مدارک اصلی (مهندسی) ----
    if 'master' in data and data['master'] is not None:
        df = data['master']
        cols = data.get('master_cols', {})
        for row in df.to_dict(orient='records'):
            doc_no_raw = _row_get(row, cols.get('doc_no'))
            doc_no = fix_document_number(doc_no_raw)
            eng_weight = _row_get(row, cols.get('eng_weight'), 0)
            doc_progress = _row_get(row, cols.get('doc_progress'), 0)
            if pd.isna(eng_weight):
                eng_weight = 0
            if pd.isna(doc_progress):
                doc_progress = 0

            doc = {
                'project': _row_get(row, cols.get('project')),
                'document_no': doc_no,
                'document_title': _row_get(row, cols.get('doc_title')),
                'discipline': org_structure.canonicalize_discipline(_row_get(row, cols.get('discipline'))),
                'category': _row_get(row, cols.get('category')),
                'status': row.get('وضعیت', 'نامشخص'),
                'responsible': _row_get(row, cols.get('responsible')),
                'issued_date': _row_get(row, cols.get('issued_date')),
                'comment_date': _row_get(row, cols.get('comment_date')),
                'comment': _row_get(row, cols.get('comment')),
                'eng_weight': eng_weight,
                'doc_progress': doc_progress,
                'hold': _row_get(row, cols.get('hold'), 'No') if cols.get('hold') else 'No',
                'hold_cause': _row_get(row, cols.get('hold_cause')),
                'deleted': _row_get(row, cols.get('deleted'), 'No') if cols.get('deleted') else 'No',
                'doc_type': 'MASTER',
                'date': _row_get(row, cols.get('date'))
            }
            # ✅ فقط دیکشنری‌های معتبر اضافه می‌شوند
            if isinstance(doc, dict) and doc.get('document_no') is not None:
                integrated['master_docs'].append(doc)
            #else:
                #safe_log(f"⚠️ آیتم نامعتبر در master_docs: {doc}")

        if cols.get('project'):
            integrated['projects'].extend(df[cols['project']].dropna().unique().tolist())

    # ---- مدارک وندور ----
    if 'vendor_master' in data and data['vendor_master'] is not None:
        df = data['vendor_master']
        cols = data.get('vendor_cols', {})
        for row in df.to_dict(orient='records'):
            doc_progress = _row_get(row, cols.get('doc_progress'), 0)
            doc_no_raw = _row_get(row, cols.get('doc_no'))
            doc_no = fix_document_number(doc_no_raw)
            if pd.isna(doc_progress):
                doc_progress = 0

            doc = {
                'project': None,
                'package_name': _row_get(row, cols.get('package')),
                'vendor_name': _row_get(row, cols.get('vendor')),
                'document_no': doc_no,
                'document_title': _row_get(row, cols.get('doc_title')),
                'discipline': org_structure.canonicalize_discipline(_row_get(row, cols.get('discipline'))),
                'category': _row_get(row, cols.get('category')),
                'status': row.get('وضعیت', 'نامشخص'),
                'responsible': _row_get(row, cols.get('responsible')),
                'issued_date': _row_get(row, cols.get('issued_date')),
                'comment_date': _row_get(row, cols.get('comment_date')),
                'comment': _row_get(row, cols.get('comment')),
                'doc_progress': doc_progress,
                'weight': row.get('Weight', 0),
                'deleted': _row_get(row, cols.get('deleted'), 'No') if cols.get('deleted') else 'No',
                'doc_type': 'VENDOR',
                'date': _row_get(row, cols.get('date')),
                'to_name': None
            }
            # ✅ فقط دیکشنری‌های معتبر به vendor_docs اضافه می‌شوند
            if isinstance(doc, dict) and doc.get('document_no') is not None:
                integrated['vendor_docs'].append(doc)
            else:
                safe_log(f"⚠️ آیتم نامعتبر در vendor_docs: {doc}")

    # ---- استخراج پروژه برای وندور ----
    project_from_master = {}
    if 'vendor_master' in data and data['vendor_master'] is not None:
        vm = data['vendor_master']
        vm_doc_no = find_column(vm, ['Document No.', 'Document No'])
        vm_project = find_column(vm, ['Project'])
        if vm_doc_no and vm_project:
            for row in vm.to_dict(orient='records'):
                doc_no = row.get(vm_doc_no)
                if pd.isna(doc_no):
                    continue
                doc_no = str(doc_no).strip()
                project_val = row.get(vm_project)
                if pd.notna(project_val) and project_val:
                    project_from_master[doc_no] = str(project_val).strip()

    if 'vendor_history' in data and data['vendor_history'] is not None:
        vh = data['vendor_history']
        vh_doc_no = find_column(vh, ['Document No.', 'Document No'])
        vh_project = find_column(vh, ['Project'])
        vh_action_date = find_column(vh, ['Action Date'])
        vh_to_name = find_column(vh, ['To Name'])

        if vh_doc_no and vh_project:
            latest_vh = {}
            for row in vh.to_dict(orient='records'):
                doc_no = row.get(vh_doc_no)
                if pd.isna(doc_no):
                    continue
                doc_no = str(doc_no).strip()
                action_date = row.get(vh_action_date) if vh_action_date else None
                project_val = row.get(vh_project)
                to_name_val = row.get(vh_to_name) if vh_to_name else None

                if pd.isna(project_val) or not project_val:
                    continue
                if doc_no not in latest_vh or (
                    action_date is not None and
                    (latest_vh[doc_no].get('action_date') is None or
                     action_date > latest_vh[doc_no].get('action_date'))
                ):
                    latest_vh[doc_no] = {
                        'project': str(project_val).strip(),
                        'action_date': action_date,
                        'to_name': str(to_name_val).strip() if to_name_val and not pd.isna(to_name_val) else None
                    }

            for doc in integrated['vendor_docs']:
                doc_no = doc.get('document_no')
                if doc_no:
                    if doc_no in latest_vh:
                        doc['project'] = latest_vh[doc_no]['project']
                        doc['to_name'] = latest_vh[doc_no].get('to_name')
                    elif doc_no in project_from_master:
                        doc['project'] = project_from_master[doc_no]

    # ---- ترکیب همه مدارک و فیلتر نهایی ----
    all_docs = integrated['master_docs'] + integrated['vendor_docs']
    # ✅ فیلتر نهایی: فقط دیکشنری‌های با document_no معتبر
    all_docs = [d for d in all_docs if isinstance(d, dict) and d.get('document_no') is not None]

    integrated['disciplines'] = sorted(set(d.get('discipline') for d in all_docs if d.get('discipline')))
    integrated['projects'] = sorted(set(d.get('project') for d in all_docs if d.get('project')))

    # ---- آخرین رکورد هر مدرک ----
    latest_records = {}
    for doc in all_docs:
        doc_no = doc.get('document_no')
        if not doc_no:
            continue
        doc_date = doc.get('date')
        if doc_date is None or pd.isna(doc_date):
            continue
        if doc.get('project') is None or pd.isna(doc.get('project')):
            continue
        if doc_no not in latest_records:
            latest_records[doc_no] = doc
            latest_records[doc_no]['date'] = doc_date
        else:
            existing_date = latest_records[doc_no].get('date')
            if existing_date is None or pd.isna(existing_date):
                latest_records[doc_no] = doc
                latest_records[doc_no]['date'] = doc_date
            elif doc_date > existing_date:
                latest_records[doc_no] = doc
                latest_records[doc_no]['date'] = doc_date

    # یکتا‌سازی نهایی
    unique_records = {}
    for doc_no, doc in latest_records.items():
        if doc_no not in unique_records:
            unique_records[doc_no] = doc
        else:
            existing = unique_records[doc_no]
            existing_date = existing.get('date')
            new_date = doc.get('date')
            if new_date and (existing_date is None or new_date > existing_date):
                unique_records[doc_no] = doc

    latest_records = unique_records
    all_latest_docs = list(latest_records.values())

    # ---- تفکیک مهندسی و وندور ----
    engineering_latest = [d for d in all_latest_docs if d.get('doc_type') == 'MASTER']
    vendor_latest = [d for d in all_latest_docs if d.get('doc_type') == 'VENDOR']

    engineering_progress = calculate_adjusted_progress(engineering_latest) if engineering_latest else 0
    vendor_progress = calculate_vendor_progress(vendor_latest) if vendor_latest else 0
    overall_progress = round((engineering_progress + vendor_progress) / 2, 2) if (engineering_latest and vendor_latest) else (engineering_progress or vendor_progress)

    # ---- آمار ----
    active_latest = [d for d in all_latest_docs if str(d.get('deleted', '')).lower() != 'yes']
    total = len(active_latest)

    with_customer = sum(1 for d in active_latest
                        if get_responsible_category(d.get('responsible')) == 'with_customer'
                        and (d.get('comment_date') is None or pd.isna(d.get('comment_date'))))
    not_issued = sum(1 for d in active_latest
                     if get_responsible_category(d.get('responsible')) == 'not_issued'
                     and (d.get('issued_date') is None or pd.isna(d.get('issued_date'))))
    approved = sum(1 for d in active_latest
                   if get_responsible_category(d.get('responsible')) == 'approved'
                   and d.get('doc_progress', 0) == 100)
    hold = sum(1 for d in active_latest if str(d.get('hold', '')).lower() == 'yes')
    deleted = sum(1 for d in active_latest if str(d.get('deleted', '')).lower() == 'yes')

    integrated['stats'] = {
        'total': total,
        'master_count': len([d for d in active_latest if d.get('doc_type') == 'MASTER']),
        'vendor_count': len([d for d in active_latest if d.get('doc_type') == 'VENDOR']),
        'with_customer': with_customer,
        'not_issued': not_issued,
        'approved': approved,
        'hold': hold,
        'deleted': deleted,
        'avg_progress': overall_progress,
        'engineering_progress': engineering_progress,
        'vendor_progress': vendor_progress
    }
    integrated['overall_progress'] = overall_progress
    integrated['engineering_progress'] = engineering_progress
    integrated['vendor_progress'] = vendor_progress

    # ---- اینباکس ----
    integrated['inbox_stats'] = calculate_inbox_stats(
        data.get('history'),
        data.get('history_cols', {}),
        data.get('vendor_history'),
        data.get('vendor_history_cols', {}),
        data.get('persons'),
        project_filter=None
    )

    # ---- پکیج‌های وندور ----
    package_stats = {}
    latest_docs_by_no = {}
    for doc in all_docs:
        doc_no = doc.get('document_no')
        if not doc_no:
            continue
        doc_date = doc.get('date')
        if doc_date is None or pd.isna(doc_date):
            continue
        if doc_no not in latest_docs_by_no:
            latest_docs_by_no[doc_no] = doc
        else:
            existing_date = latest_docs_by_no[doc_no].get('date')
            if existing_date is None or pd.isna(existing_date) or doc_date > existing_date:
                latest_docs_by_no[doc_no] = doc

    for doc in latest_docs_by_no.values():
        pkg = doc.get('package_name')
        if not pkg:
            continue
        proj = doc.get('project', 'نامشخص')
        if not proj or pd.isna(proj):
            proj = 'نامشخص'
        pkg_norm = normalize_package(pkg)
        vendor_name = doc.get('vendor_name', 'نامشخص')
        vendor_norm = normalize_vendor(vendor_name)

        if proj not in package_stats:
            package_stats[proj] = {}
        if pkg_norm not in package_stats[proj]:
            package_stats[proj][pkg_norm] = {
                'count': 0,
                'progress': 0,
                'vendor': vendor_norm,
                'vendors': set(),
                'weight_sum': 0,
                'weighted_progress': 0,
                'original_name': pkg,
                'original_vendor': vendor_name
            }

        package_stats[proj][pkg_norm]['vendors'].add(vendor_norm)
        package_stats[proj][pkg_norm]['count'] += 1

        weight = doc.get('weight', 0)
        progress = get_adjusted_progress(doc)
        if pd.isna(weight):
            weight = 0
        if pd.isna(progress):
            progress = 0

        if weight > 0:
            normalized_progress = normalize_progress(progress)
            package_stats[proj][pkg_norm]['weight_sum'] += weight
            package_stats[proj][pkg_norm]['weighted_progress'] += weight * normalized_progress

    for proj in package_stats:
        for pkg in package_stats[proj]:
            vendors_list = sorted(list(package_stats[proj][pkg]['vendors']))
            package_stats[proj][pkg]['vendor'] = ', '.join(vendors_list)
            del package_stats[proj][pkg]['vendors']
            if package_stats[proj][pkg]['weight_sum'] > 0:
                package_stats[proj][pkg]['progress'] = round(
                    (package_stats[proj][pkg]['weighted_progress'] / package_stats[proj][pkg]['weight_sum']) * 100, 2
                )
            else:
                package_stats[proj][pkg]['progress'] = 0

    integrated['package_stats'] = package_stats

    # ---- دیرکردها ----
    overdue_docs_eng = calculate_overdue_docs(all_latest_docs, data, person_df=data.get('persons'))
    overdue_docs_vendor = calculate_vendor_overdue(
        data.get('vendor_history'),
        data.get('vendor_history_cols', {}),
        data.get('persons')
    )
    all_overdue = overdue_docs_eng + overdue_docs_vendor

    seen_docs = set()
    unique_overdue = []
    for doc in all_overdue:
        doc_no = doc.get('document_no')
        if doc_no and doc_no not in seen_docs:
            seen_docs.add(doc_no)
            unique_overdue.append(doc)

    overdue_docs = unique_overdue

    integrated['overdue_docs'] = overdue_docs
    integrated['overdue_client'] = [d for d in overdue_docs if d.get('type') == 'client']
    integrated['overdue_contractor'] = [d for d in overdue_docs if d.get('type') in ['contractor', 'vendor'] and not d.get('over_150', False)]
    integrated['overdue_contractor_all'] = [d for d in overdue_docs if d.get('type') in ['contractor', 'vendor']]
    integrated['contractor_over_150_count'] = sum(1 for d in overdue_docs if d.get('type') in ['contractor', 'vendor'] and d.get('over_150', False))
    integrated['total_contractor_overdue'] = len(integrated['overdue_contractor_all'])
    integrated['hold_docs'] = [d for d in active_latest if str(d.get('hold', '')).lower() == 'yes']
    integrated['deleted_count'] = deleted

    # ---- فیلتر نهایی برای اطمینان از صحت داده‌ها ----
    integrated['master_docs'] = [d for d in integrated['master_docs'] if isinstance(d, dict) and d.get('document_no') is not None]
    integrated['vendor_docs'] = [d for d in integrated['vendor_docs'] if isinstance(d, dict) and d.get('document_no') is not None]
    return integrated

def process_vendor_history(data):
    if 'vendor_history' not in data or data['vendor_history'] is None:
        return data
    df = data['vendor_history'].copy()
    
    col = {
        'project': find_column(df, ['Project']),
        'date': find_column(df, ['Date']),
        'discipline': find_column(df, ['Discipline']),
        'doc_no': find_column(df, ['Document No.', 'Document No']),
        'doc_title': find_column(df, ['Document Title']),
        'action_date': find_column(df, ['Action Date']),
        'from_name': find_column(df, ['From Name', 'From_Name', 'From name', 'From']),  # ✅ گسترش یافته
        'to_name': find_column(df, ['To Name', 'To_Name', 'To name', 'To']),  # ✅ گسترش یافته
        'log_status': find_column(df, ['Log Status']),
        'category': find_column(df, ['Category']),
        'ongoing': find_column(df, ['Ongoing']),
        'close_date': find_column(df, ['Close Date']),
        'comment_date': find_column(df, ['Comment Date'])
    }
    
    for c in [col['action_date'], col['close_date'], col['comment_date'], col['date']]:
        if c:
            df[c] = df[c].apply(convert_excel_date)
    data['vendor_history'] = df
    data['vendor_history_cols'] = col
    return data

def calculate_vendor_progress(docs):
    """
    محاسبه پیشرفت وزنی برای مدارک وندور
    - از فیلد weight (که توسط سیستم محاسبه شده) استفاده می‌کند
    - تعدیل ۱۰۰% به ۹۵% را اعمال نمی‌کند
    - همه مدارک (حتی حذف شده) محاسبه می‌شوند
    """
    total_weighted = 0
    total_weight = 0

    for doc in docs:
        weight = doc.get('weight', 0)
        if pd.isna(weight) or weight == 0:
            continue

        progress = doc.get('doc_progress', 0)
        if pd.isna(progress):
            progress = 0

        normalized_progress = normalize_progress(progress)
        total_weighted += weight * normalized_progress
        total_weight += weight

    if total_weight == 0:
        return 0

    return round((total_weighted / total_weight) * 100, 2)

# ==================== Routes ====================
@app.route('/')
def index():
    return render_template('dashboard.html')


@app.route('/leaderboard')
def leaderboard_page():
    return render_template('leaderboard.html')


@app.route('/portfolio')
def portfolio_page():
    return redirect(url_for('leaderboard_page') + '#portfolio')


@app.route('/issue-prediction')
def issue_prediction_page():
    return render_template('issue_prediction.html')


# ==================== پیش‌بینی تاریخ صدور مدرک ====================
PREDICTION_KEY_PREFIX = "issue_pred:"
ISSUE_PREDICTION_ADMIN_USERS = ['احسان برازنده راد', 'admin']
DASHBOARD_ADMIN_USERS = ISSUE_PREDICTION_ADMIN_USERS  # همون لیست؛ اسم عمومی‌تر برای بخش‌های دیگه (مثل دانش دستی)


# ==================== دانش دستی دستیار هوشمند ====================
# یادداشت‌هایی که مدیر پروژه دستی وارد می‌کند تا دستیار در جواب‌ها ازشون
# استفاده کنه. دو دسته: 'fact' (یک واقعیت/نکته که هنگام جست‌وجوی سؤال باهاش
# مچ می‌شه) و 'instruction' (یک قانون رفتاری که همیشه به مدل یادآوری می‌شه،
# مثل لحن پاسخ یا نکته‌ای که باید همیشه رعایت بشه).
KNOWLEDGE_KEY_PREFIX = "manual_knowledge:"


def _save_knowledge_entry(entry_id, record):
    if not redis_client:
        return False
    try:
        redis_client.set(f"{KNOWLEDGE_KEY_PREFIX}{entry_id}", json.dumps(record, default=str))
        return True
    except Exception as e:
        safe_log(f"[knowledge] خطا در ذخیره‌ی یادداشت {entry_id}: {e}", level="warning")
        return False


def _delete_knowledge_entry(entry_id):
    if not redis_client:
        return False
    try:
        redis_client.delete(f"{KNOWLEDGE_KEY_PREFIX}{entry_id}")
        return True
    except Exception as e:
        safe_log(f"[knowledge] خطا در حذف یادداشت {entry_id}: {e}", level="warning")
        return False


def _get_all_knowledge_entries():
    """همه‌ی یادداشت‌ها رو برمی‌گردونه، جدیدترین اول."""
    if not redis_client:
        return []
    try:
        entries = []
        for key in redis_client.scan_iter(f"{KNOWLEDGE_KEY_PREFIX}*"):
            entry_id = key.decode('utf-8').replace(KNOWLEDGE_KEY_PREFIX, '') if isinstance(key, bytes) else key.replace(KNOWLEDGE_KEY_PREFIX, '')
            raw = redis_client.get(key)
            if not raw:
                continue
            try:
                record = json.loads(raw)
            except Exception:
                continue
            record['id'] = entry_id
            entries.append(record)
        entries.sort(key=lambda e: e.get('added_at', ''), reverse=True)
        return entries
    except Exception as e:
        safe_log(f"[knowledge] خطا در خواندن یادداشت‌ها: {e}", level="warning")
        return []


@app.route('/api/knowledge', methods=['GET'])
@update_activity
def list_knowledge():
    try:
        entries = _get_all_knowledge_entries()
        return jsonify({
            'entries': entries,
            'is_admin': session.get('username') in DASHBOARD_ADMIN_USERS,
        })
    except Exception as e:
        safe_log(f"[knowledge] خطا در فهرست یادداشت‌ها: {e}", level="error")
        return jsonify({'error': str(e)}), 500


@app.route('/api/knowledge/add', methods=['POST'])
@update_activity
def add_knowledge():
    try:
        current_user = session.get('username')
        if not current_user:
            return jsonify({'error': 'ابتدا وارد شوید'}), 401
        if current_user not in DASHBOARD_ADMIN_USERS:
            return jsonify({'error': 'فقط ادمین می‌تواند به دستیار هوشمند چیز جدید یاد بدهد'}), 403

        payload = request.get_json(silent=True) or {}
        text = (payload.get('text') or '').strip()
        category = payload.get('category') or 'fact'
        if category not in ('fact', 'instruction'):
            category = 'fact'
        if not text:
            return jsonify({'error': 'متن یادداشت خالی است'}), 400
        if len(text) > 1000:
            return jsonify({'error': 'یادداشت خیلی طولانی است (حداکثر ۱۰۰۰ کاراکتر)'}), 400

        entry_id = f"{int(time.time() * 1000)}_{uuid.uuid4().hex[:6]}"
        record = {
            'text': text,
            'category': category,
            'added_by': current_user,
            'added_at': datetime.now().isoformat(),
        }
        _save_knowledge_entry(entry_id, record)
        record['id'] = entry_id
        return jsonify({'success': True, 'entry': record})
    except Exception as e:
        safe_log(f"[knowledge] خطا در افزودن یادداشت: {e}", level="error")
        safe_log(traceback.format_exc(), level="error")
        return jsonify({'error': str(e)}), 500


@app.route('/api/knowledge/<entry_id>', methods=['DELETE'])
@update_activity
def delete_knowledge(entry_id):
    try:
        current_user = session.get('username')
        if not current_user or current_user not in DASHBOARD_ADMIN_USERS:
            return jsonify({'error': 'فقط ادمین می‌تواند یادداشت حذف کند'}), 403
        _delete_knowledge_entry(entry_id)
        return jsonify({'success': True})
    except Exception as e:
        safe_log(f"[knowledge] خطا در حذف یادداشت: {e}", level="error")
        return jsonify({'error': str(e)}), 500


@app.route('/knowledge')
def knowledge_page():
    return render_template('knowledge.html')


# ==================== اطلاعات مالی پروژه‌ها (قرارداد/الحاقیه/صورت‌وضعیت) ====================
# ویرایش: فقط احسان برازنده راد. دیدن: یک لیست جداگانه و قابل‌ویرایش (پیش‌فرض
# چند نفر) که فقط خودِ احسان می‌تونه بهش اضافه/ازش کم کنه — دقیقاً مثل لیست
# ایمیل‌ها تو /admin/email.
FINANCE_EDITOR_USERS = ['احسان برازنده راد']
FINANCE_VIEWERS_KEY = "finance_viewers_list"
DEFAULT_FINANCE_VIEWERS = ['احسان برازنده راد']

FINANCE_PROJECT_PREFIX = "project_finance:"
FINANCE_PAYMENT_PREFIX = "progress_payment:"


def _get_finance_viewers():
    if not redis_client:
        return list(DEFAULT_FINANCE_VIEWERS)
    try:
        raw = redis_client.get(FINANCE_VIEWERS_KEY)
        if raw:
            return json.loads(raw)
    except Exception:
        pass
    return list(DEFAULT_FINANCE_VIEWERS)


def _save_finance_viewers(viewers):
    if redis_client:
        try:
            redis_client.set(FINANCE_VIEWERS_KEY, json.dumps(viewers, ensure_ascii=False))
        except Exception as e:
            safe_log(f"[finance] خطا در ذخیره‌ی لیست دسترسی: {e}", level="warning")


def _can_view_finance(username):
    return bool(username) and (username in FINANCE_EDITOR_USERS or username in _get_finance_viewers())


def _can_edit_finance(username):
    return username in FINANCE_EDITOR_USERS


def _get_finance_project(project_name):
    if not redis_client:
        return None
    try:
        raw = redis_client.get(f"{FINANCE_PROJECT_PREFIX}{project_name}")
        return json.loads(raw) if raw else None
    except Exception as e:
        safe_log(f"[finance] خطا در خواندن پروژه {project_name}: {e}", level="warning")
        return None


def _save_finance_project(project_name, record):
    if redis_client:
        try:
            redis_client.set(f"{FINANCE_PROJECT_PREFIX}{project_name}", json.dumps(record, ensure_ascii=False, default=str))
            return True
        except Exception as e:
            safe_log(f"[finance] خطا در ذخیره‌ی پروژه {project_name}: {e}", level="warning")
    return False


def _get_all_finance_projects():
    if not redis_client:
        return []
    out = []
    try:
        for key in redis_client.scan_iter(f"{FINANCE_PROJECT_PREFIX}*"):
            raw = redis_client.get(key)
            if raw:
                try:
                    out.append(json.loads(raw))
                except Exception:
                    continue
    except Exception as e:
        safe_log(f"[finance] خطا در خواندن پروژه‌ها: {e}", level="warning")
    return out


def _get_progress_payments(project_name):
    if not redis_client:
        return []
    try:
        raw = redis_client.get(f"{FINANCE_PAYMENT_PREFIX}{project_name}")
        return json.loads(raw) if raw else []
    except Exception as e:
        safe_log(f"[finance] خطا در خواندن صورت‌وضعیت‌های {project_name}: {e}", level="warning")
        return []


def _save_progress_payments(project_name, payments):
    if redis_client:
        try:
            redis_client.set(f"{FINANCE_PAYMENT_PREFIX}{project_name}", json.dumps(payments, ensure_ascii=False, default=str))
            return True
        except Exception as e:
            safe_log(f"[finance] خطا در ذخیره‌ی صورت‌وضعیت‌های {project_name}: {e}", level="warning")
    return False


def get_all_finance_summaries():
    """
    برای هر پروژه‌ای که رکورد مالی داره، summary آماده می‌سازه — استفاده در
    /api/finance/projects و در پاسخ دستیار هوشمند به سؤال‌های مالی.
    """
    import project_finance
    out = {}
    for record in _get_all_finance_projects():
        name = record.get('project_name')
        if not name:
            continue
        payments = _get_progress_payments(name)
        out[name] = project_finance.compute_summary(record, payments)
    return out


@app.route('/api/finance/projects', methods=['GET'])
@update_activity
def list_finance_projects():
    current_user = session.get('username')
    if not _can_view_finance(current_user):
        return jsonify({'error': 'دسترسی به اطلاعات مالی نداری'}), 403
    return jsonify({
        'projects': list(get_all_finance_summaries().values()),
        'can_edit': _can_edit_finance(current_user),
    })


@app.route('/api/finance/projects/<project_name>', methods=['GET'])
@update_activity
def get_finance_project(project_name):
    import project_finance
    current_user = session.get('username')
    if not _can_view_finance(current_user):
        return jsonify({'error': 'دسترسی به اطلاعات مالی نداری'}), 403
    record = _get_finance_project(project_name)
    if not record:
        return jsonify({'error': 'این پروژه هنوز ثبت نشده'}), 404
    payments = _get_progress_payments(project_name)
    summary = project_finance.compute_summary(record, payments)
    return jsonify({
        'project': record,
        'summary': summary,
        'payments': sorted(payments, key=lambda p: p.get('payment_date') or '', reverse=True),
        'can_edit': _can_edit_finance(current_user),
    })


@app.route('/api/finance/projects', methods=['POST'])
@update_activity
def save_finance_project():
    try:
        current_user = session.get('username')
        if not _can_edit_finance(current_user):
            return jsonify({'error': 'فقط احسان برازنده راد می‌تواند این بخش را ویرایش کند'}), 403

        payload = request.get_json(silent=True) or {}
        project_name = (payload.get('project_name') or '').strip()
        if not project_name:
            return jsonify({'error': 'نام پروژه خالی است'}), 400

        project_code = str(payload.get('project_code') or '').strip()
        if project_code:
            for other in _get_all_finance_projects():
                if other.get('project_name') != project_name and str(other.get('project_code') or '').strip() == project_code:
                    return jsonify({'error': f'این کد قبلاً به پروژه‌ی «{other.get("project_name")}» تخصیص داده شده'}), 400

        existing = _get_finance_project(project_name) or {}
        record = {
            'project_name': project_name,
            'project_code': project_code,
            'employer': (payload.get('employer') or '').strip(),
            'contract_name': (payload.get('contract_name') or '').strip(),
            'start_date': payload.get('start_date') or None,
            'original_amount': payload.get('original_amount') or 0,
            'duration_months': payload.get('duration_months') or None,
            'amendments': existing.get('amendments') or [],
            'subcontractors': existing.get('subcontractors') or [],
            'updated_by': current_user,
            'updated_at': datetime.now().isoformat(),
        }
        _save_finance_project(project_name, record)
        return jsonify({'success': True, 'project': record})
    except Exception as e:
        safe_log(f"[finance] خطا در ذخیره‌ی پروژه: {e}", level="error")
        safe_log(traceback.format_exc(), level="error")
        return jsonify({'error': str(e)}), 500


@app.route('/api/finance/projects/<project_name>', methods=['DELETE'])
@update_activity
def delete_finance_project(project_name):
    current_user = session.get('username')
    if not _can_edit_finance(current_user):
        return jsonify({'error': 'فقط احسان برازنده راد می‌تواند این بخش را ویرایش کند'}), 403
    if redis_client:
        try:
            redis_client.delete(f"{FINANCE_PROJECT_PREFIX}{project_name}")
            redis_client.delete(f"{FINANCE_PAYMENT_PREFIX}{project_name}")
        except Exception as e:
            safe_log(f"[finance] خطا در حذف پروژه {project_name}: {e}", level="warning")
    return jsonify({'success': True})


@app.route('/api/finance/projects/<project_name>/amendment', methods=['POST'])
@update_activity
def add_finance_amendment(project_name):
    current_user = session.get('username')
    if not _can_edit_finance(current_user):
        return jsonify({'error': 'فقط احسان برازنده راد می‌تواند این بخش را ویرایش کند'}), 403
    record = _get_finance_project(project_name)
    if not record:
        return jsonify({'error': 'این پروژه هنوز ثبت نشده'}), 404

    payload = request.get_json(silent=True) or {}
    amendment = {
        'id': f"{int(time.time() * 1000)}_{uuid.uuid4().hex[:6]}",
        'registered_date': payload.get('registered_date') or datetime.now().strftime('%Y-%m-%d'),
        'amount_added': payload.get('amount_added') or 0,
        'new_expiry_date': payload.get('new_expiry_date') or None,
        'added_by': current_user,
    }
    record.setdefault('amendments', []).append(amendment)
    record['updated_by'] = current_user
    record['updated_at'] = datetime.now().isoformat()
    _save_finance_project(project_name, record)
    return jsonify({'success': True, 'amendment': amendment})


@app.route('/api/finance/projects/<project_name>/amendment/<amendment_id>', methods=['DELETE'])
@update_activity
def delete_finance_amendment(project_name, amendment_id):
    current_user = session.get('username')
    if not _can_edit_finance(current_user):
        return jsonify({'error': 'فقط احسان برازنده راد می‌تواند این بخش را ویرایش کند'}), 403
    record = _get_finance_project(project_name)
    if not record:
        return jsonify({'error': 'این پروژه هنوز ثبت نشده'}), 404
    record['amendments'] = [a for a in (record.get('amendments') or []) if a.get('id') != amendment_id]
    _save_finance_project(project_name, record)
    return jsonify({'success': True})


@app.route('/api/finance/projects/<project_name>/subcontractor', methods=['POST'])
@update_activity
def add_finance_subcontractor(project_name):
    current_user = session.get('username')
    if not _can_edit_finance(current_user):
        return jsonify({'error': 'فقط احسان برازنده راد می‌تواند این بخش را ویرایش کند'}), 403
    record = _get_finance_project(project_name)
    if not record:
        return jsonify({'error': 'این پروژه هنوز ثبت نشده'}), 404

    payload = request.get_json(silent=True) or {}
    name = (payload.get('name') or '').strip()
    if not name:
        return jsonify({'error': 'نام پیمانکار جز خالی است'}), 400

    subcontractor = {
        'id': f"{int(time.time() * 1000)}_{uuid.uuid4().hex[:6]}",
        'name': name,
        'scope': (payload.get('scope') or '').strip(),
    }
    record.setdefault('subcontractors', []).append(subcontractor)
    record['updated_by'] = current_user
    record['updated_at'] = datetime.now().isoformat()
    _save_finance_project(project_name, record)
    return jsonify({'success': True, 'subcontractor': subcontractor})


@app.route('/api/finance/projects/<project_name>/subcontractor/<sub_id>', methods=['DELETE'])
@update_activity
def delete_finance_subcontractor(project_name, sub_id):
    current_user = session.get('username')
    if not _can_edit_finance(current_user):
        return jsonify({'error': 'فقط احسان برازنده راد می‌تواند این بخش را ویرایش کند'}), 403
    record = _get_finance_project(project_name)
    if not record:
        return jsonify({'error': 'این پروژه هنوز ثبت نشده'}), 404
    record['subcontractors'] = [s for s in (record.get('subcontractors') or []) if s.get('id') != sub_id]
    _save_finance_project(project_name, record)
    return jsonify({'success': True})


def _jalali_to_gregorian(jy, jm, jd):
    """تبدیل تاریخ شمسی به میلادی (الگوریتم استاندارد، بدون نیاز به کتابخانه‌ی خارجی)."""
    jy += 1595
    days = -355668 + (365 * jy) + ((jy // 33) * 8) + (((jy % 33) + 3) // 4) + jd
    if jm < 7:
        days += (jm - 1) * 31
    else:
        days += ((jm - 7) * 30) + 186
    gy = 400 * (days // 146097)
    days %= 146097
    if days > 36524:
        days -= 1
        gy += 100 * (days // 36524)
        days %= 36524
        if days >= 365:
            days += 1
    gy += 4 * (days // 1461)
    days %= 1461
    if days > 365:
        gy += (days - 1) // 365
        days = (days - 1) % 365
    gd = days + 1
    is_leap_g = (gy % 4 == 0 and (gy % 100 != 0 or gy % 400 == 0))
    days_in_month = [0, 31, 29 if is_leap_g else 28, 31, 30, 31, 30, 31, 31, 30, 31, 30, 31]
    gm = 1
    for i in range(1, 13):
        if gd <= days_in_month[i]:
            gm = i
            break
        gd -= days_in_month[i]
    return gy, gm, gd


def _parse_flexible_date(value):
    """
    تاریخ صورت‌وضعیت رو هم به‌صورت میلادی و هم شمسی می‌فهمه. کسی که این
    اکسل رو دستی پر می‌کنه، طبیعیه تاریخ شمسی بنویسه (مثل 1405/05/31)؛
    pandas این رو یا با خطا رد می‌کنه یا (بدتر) به‌اشتباه به‌عنوان یک سال
    میلادی عجیب می‌خونه. اینجا اول تشخیص می‌دیم شمسیه یا میلادی.
    برمی‌گردونه: رشته‌ی 'YYYY-MM-DD' میلادی، یا None اگه نامعتبر بود.
    """
    if value is None:
        return None
    try:
        if pd.isna(value):
            return None
    except (TypeError, ValueError):
        pass

    if isinstance(value, (pd.Timestamp, datetime)):
        return value.strftime('%Y-%m-%d')

    text = str(value).strip()
    m = re.match(r'^(\d{3,4})[/\-.](\d{1,2})[/\-.](\d{1,2})$', text)
    if m:
        y, mo, d = int(m.group(1)), int(m.group(2)), int(m.group(3))
        if 1300 <= y <= 1500 and 1 <= mo <= 12 and 1 <= d <= 31:
            try:
                gy, gm, gd = _jalali_to_gregorian(y, mo, d)
                return f"{gy:04d}-{gm:02d}-{gd:02d}"
            except Exception:
                return None

    try:
        parsed = pd.to_datetime(value, errors='coerce')
    except Exception:
        return None
    if pd.isna(parsed):
        return None
    # اگه با وجود پارس «موفق»، سال خارج از بازه‌ی منطقیِ پروژه‌هاست (مثلاً
    # همون تفسیر اشتباه سال شمسی به‌عنوان میلادی)، به‌جای ثبت یک تاریخ
    # مزخرف، نامعتبر حسابش کن.
    if parsed.year < 1990 or parsed.year > 2100:
        return None
    return parsed.strftime('%Y-%m-%d')


def _normalize_match_text(s):
    return re.sub(r'\s+', ' ', str(s or '').strip()).lower()


def _resolve_finance_project_name(raw_project_text, raw_code, all_projects):
    """
    متنِ ستون «پروژه» و/یا «کد» تو اکسل رو به نام دقیق و از قبل ثبت‌شده‌ی
    پروژه تبدیل می‌کنه — تا فرقِ حروف بزرگ/کوچیک یا فاصله‌ی اضافه باعث
    نشه یک پروژه چند بار با اسم‌های کمی متفاوت ذخیره بشه. کد اولویت داره
    چون قابل‌اتکاتره؛ اگه کد نبود یا مچ نشد، اسم (بدون‌حساسیت به بزرگ/کوچیک)
    چک می‌شه. اگه هیچ‌کدوم مچ نشد، None برمی‌گرده (یعنی این پروژه هنوز تو
    سیستم ثبت نشده).
    """
    code_norm = str(raw_code or '').strip()
    if code_norm:
        for p in all_projects:
            if str(p.get('project_code') or '').strip() == code_norm:
                return p.get('project_name')

    name_norm = _normalize_match_text(raw_project_text)
    if name_norm:
        for p in all_projects:
            if _normalize_match_text(p.get('project_name')) == name_norm:
                return p.get('project_name')

    return None


@app.route('/api/finance/progress-payments/import', methods=['POST'])
@update_activity
def import_progress_payments():
    try:
        current_user = session.get('username')
        if not _can_edit_finance(current_user):
            return jsonify({'error': 'فقط احسان برازنده راد می‌تواند صورت‌وضعیت وارد کند'}), 403

        if 'file' not in request.files:
            return jsonify({'error': 'فایلی ارسال نشده است'}), 400
        file = request.files['file']
        if not file or not file.filename:
            return jsonify({'error': 'فایلی انتخاب نشده است'}), 400

        try:
            df = pd.read_excel(file, engine='openpyxl')
        except Exception as e:
            return jsonify({'error': f'خواندن فایل اکسل ناموفق بود: {e}'}), 400

        df.columns = [str(c).strip() for c in df.columns]
        proj_col = next((c for c in df.columns if 'پروژه' in c), None)
        code_col = next((c for c in df.columns if 'کد' in c), None)
        contractor_col = next((c for c in df.columns if 'پیمانکار' in c), None)
        date_col = next((c for c in df.columns if 'تاریخ' in c), None)
        amount_col = next((c for c in df.columns if 'مبلغ' in c), None)
        period_col = next((c for c in df.columns if ('دوره' in c or 'توضیح' in c)), None)

        if (not proj_col and not code_col) or not date_col or not amount_col:
            return jsonify({'error': 'ستون‌های «پروژه» (یا «کد»)، «تاریخ» و «مبلغ» تو فایل پیدا نشد'}), 400

        all_projects = _get_all_finance_projects()
        if not all_projects:
            return jsonify({'error': 'هنوز هیچ پروژه‌ای تو سیستم ثبت نشده — اول از همین صفحه پروژه رو بساز'}), 400

        grouped = {}
        imported = 0
        skipped = []

        for _, r in df.iterrows():
            raw_project_text = str(r.get(proj_col, '')).strip() if proj_col else ''
            raw_code = r.get(code_col) if code_col else None
            if (not raw_project_text or raw_project_text.lower() == 'nan') and not str(raw_code or '').strip():
                continue

            project_name = _resolve_finance_project_name(raw_project_text, raw_code, all_projects)
            if not project_name:
                skipped.append({
                    'project': raw_project_text or str(raw_code),
                    'reason': 'پروژه‌ای با این نام/کد ثبت نشده — اول از صفحه‌ی پروژه‌ها ثبتش کن',
                })
                continue

            try:
                amount = float(r.get(amount_col))
                if pd.isna(amount):
                    raise ValueError()
            except Exception:
                skipped.append({'project': project_name, 'reason': 'مبلغ نامعتبر'})
                continue

            payment_date = _parse_flexible_date(r.get(date_col))
            if not payment_date:
                skipped.append({'project': project_name, 'reason': 'فرمت تاریخ نامعتبر (میلادی یا شمسی بنویسید، مثل 1405/05/31 یا 2026-08-22)'})
                continue

            contractor = ''
            if contractor_col:
                contractor = str(r.get(contractor_col, '') or '').strip()
                if contractor.lower() == 'nan':
                    contractor = ''

            record = {
                'id': f"{int(time.time() * 1000)}_{uuid.uuid4().hex[:6]}",
                'payer_type': 'subcontractor' if contractor else 'main',
                'subcontractor_name': contractor or None,
                'payment_date': payment_date,
                'amount': amount,
                'period_covered': (str(r.get(period_col, '') or '') if period_col else ''),
                'imported_by': current_user,
                'imported_at': datetime.now().isoformat(),
            }
            grouped.setdefault(project_name, []).append(record)
            imported += 1

        for project_name, records in grouped.items():
            existing = _get_progress_payments(project_name)
            existing.extend(records)
            _save_progress_payments(project_name, existing)

        return jsonify({
            'success': True,
            'imported': imported,
            'projects_touched': list(grouped.keys()),
            'skipped_count': len(skipped),
            'skipped': skipped[:50],
        })
    except Exception as e:
        safe_log(f"[finance] خطا در ایمپورت صورت‌وضعیت: {e}", level="error")
        safe_log(traceback.format_exc(), level="error")
        return jsonify({'error': str(e)}), 500


@app.route('/api/finance/projects/<project_name>/contract-file', methods=['POST'])
@update_activity
def upload_contract_file(project_name):
    import contract_knowledge
    current_user = session.get('username')
    if not _can_edit_finance(current_user):
        return jsonify({'error': 'فقط احسان برازنده راد می‌تواند فایل قرارداد اضافه کند'}), 403

    if 'file' not in request.files:
        return jsonify({'error': 'فایلی ارسال نشده است'}), 400
    file = request.files['file']
    if not file or not file.filename:
        return jsonify({'error': 'فایلی انتخاب نشده است'}), 400

    allowed_ext = {'.pdf', '.docx', '.xlsx', '.xls', '.txt'}
    ext = os.path.splitext(file.filename)[1].lower()
    if ext not in allowed_ext:
        return jsonify({'error': 'فقط PDF/Word/اکسل مجاز است'}), 400

    d = contract_knowledge.ensure_project_dir(project_name)
    safe_name = f"{int(time.time())}_{os.path.basename(file.filename)}"
    file.save(os.path.join(str(d), safe_name))
    contract_knowledge.build_index(project_name, force=True)
    return jsonify({'success': True, 'filename': safe_name})


@app.route('/api/finance/projects/<project_name>/contract-files', methods=['GET'])
@update_activity
def list_contract_files(project_name):
    import contract_knowledge
    current_user = session.get('username')
    if not _can_view_finance(current_user):
        return jsonify({'error': 'دسترسی به اطلاعات مالی نداری'}), 403
    return jsonify({'files': contract_knowledge.list_files(project_name)})


@app.route('/api/finance/viewers', methods=['GET'])
@update_activity
def get_finance_viewers_route():
    current_user = session.get('username')
    if not _can_edit_finance(current_user):
        return jsonify({'error': 'فقط احسان برازنده راد می‌تواند لیست دسترسی را ببیند'}), 403
    return jsonify({'viewers': _get_finance_viewers()})


@app.route('/api/finance/viewers', methods=['POST'])
@update_activity
def add_finance_viewer():
    current_user = session.get('username')
    if not _can_edit_finance(current_user):
        return jsonify({'error': 'فقط احسان برازنده راد می‌تواند لیست دسترسی را ویرایش کند'}), 403
    payload = request.get_json(silent=True) or {}
    name = (payload.get('username') or '').strip()
    if not name:
        return jsonify({'error': 'نام خالی است'}), 400
    viewers = _get_finance_viewers()
    if name not in viewers:
        viewers.append(name)
        _save_finance_viewers(viewers)
    return jsonify({'success': True, 'viewers': viewers})


@app.route('/api/finance/viewers/<username>', methods=['DELETE'])
@update_activity
def remove_finance_viewer(username):
    current_user = session.get('username')
    if not _can_edit_finance(current_user):
        return jsonify({'error': 'فقط احسان برازنده راد می‌تواند لیست دسترسی را ویرایش کند'}), 403
    viewers = [v for v in _get_finance_viewers() if v != username]
    _save_finance_viewers(viewers)
    return jsonify({'success': True, 'viewers': viewers})


@app.route('/finance')
def finance_page():
    return render_template('finance.html')


def _get_prediction_record(doc_no):
    if not redis_client:
        return None
    try:
        raw = redis_client.get(f"{PREDICTION_KEY_PREFIX}{doc_no}")
        return json.loads(raw) if raw else None
    except Exception as e:
        safe_log(f"[issue-prediction] خطا در خواندن پیش‌بینی {doc_no}: {e}", level="warning")
        return None


def _save_prediction_record(doc_no, record):
    if not redis_client:
        return False
    try:
        redis_client.set(f"{PREDICTION_KEY_PREFIX}{doc_no}", json.dumps(record, default=str))
        return True
    except Exception as e:
        safe_log(f"[issue-prediction] خطا در ذخیره‌ی پیش‌بینی {doc_no}: {e}", level="warning")
        return False


def _get_all_prediction_records():
    """همه‌ی رکوردهای پیش‌بینی رو یک‌جا برمی‌گردونه (برای جدول اصلی و امتیازدهی)."""
    if not redis_client:
        return {}
    try:
        records = {}
        for key in redis_client.scan_iter(f"{PREDICTION_KEY_PREFIX}*"):
            doc_no = key.decode('utf-8').replace(PREDICTION_KEY_PREFIX, '') if isinstance(key, bytes) else key.replace(PREDICTION_KEY_PREFIX, '')
            raw = redis_client.get(key)
            if raw:
                try:
                    records[doc_no] = json.loads(raw)
                except Exception:
                    continue
        return records
    except Exception as e:
        safe_log(f"[issue-prediction] خطا در خواندن همه‌ی پیش‌بینی‌ها: {e}", level="warning")
        return {}


def _build_issue_prediction_rows(master_df, project, category, discipline):
    """
    مدارک فعال (فیلترشده) را به‌همراه وضعیت پیش‌بینی فعلی هرکدام برمی‌گرداند.
    project/category/discipline هرکدام می‌توانند لیست باشند (مولتی‌سلکت).
    خروجی: (rows, projects, categories, disciplines) — سه‌تای آخر بدون فیلتر
    (برای پرکردن گزینه‌های فیلتر تو UI) محاسبه می‌شوند.
    """
    active = issue_prediction.get_active_documents(master_df, project, category, discipline)

    all_active_unfiltered = issue_prediction.get_active_documents(master_df)
    projects = sorted(set(all_active_unfiltered['Project'].dropna().astype(str).str.strip()))
    categories = sorted(set(all_active_unfiltered['Category'].dropna().astype(str).str.strip()))
    disciplines = sorted(set(all_active_unfiltered['Discipline'].dropna().astype(str).str.strip()))

    all_predictions = _get_all_prediction_records()

    rows = []
    for idx, row in active.reset_index(drop=True).iterrows():
        doc_no = str(row['Document No.'])
        cycle_anchor = issue_prediction.compute_cycle_anchor(row)
        pred = all_predictions.get(doc_no)

        predicted_date = None
        is_changed = False
        can_edit = True
        if pred and pred.get('cycle_anchor') == cycle_anchor:
            predicted_date = pred.get('predicted_date')
            is_changed = pred.get('edit_count', 0) >= 1
            can_edit = pred.get('edit_count', 0) < 2  # یک‌بار ثبت + یک‌بار ویرایش = حداکثر ۲

        days_overdue = None
        if predicted_date:
            try:
                pd_date = pd.to_datetime(predicted_date).date()
                days_overdue = issue_prediction.compute_days_overdue(pd_date)
            except Exception:
                pass

        rows.append({
            'id': idx + 1,
            'document_no': doc_no,
            'document_title': row.get('Document Title'),
            'category': row.get('Category'),
            'document_progress': row.get('Document Progress'),
            'project': row.get('Project'),
            'discipline': row.get('Discipline'),
            'predicted_issue_date': predicted_date,
            'is_changed': is_changed,
            'can_edit': can_edit,
            'days_overdue': days_overdue,
        })

    return rows, projects, categories, disciplines


@app.route('/api/issue-predictions')
@update_activity
def get_issue_predictions():
    try:
        # هر سه فیلتر مولتی‌سلکت هستن: فرانت چند بار project=... رو با همون
        # اسم می‌فرسته (مثل project=A&project=B). اگه هیچی نیومده، یعنی «همه».
        project = [v.strip() for v in request.args.getlist('project') if v.strip()] or ['همه']
        category = [v.strip() for v in request.args.getlist('category') if v.strip()] or ['همه']
        discipline = [v.strip() for v in request.args.getlist('discipline') if v.strip()] or ['همه']

        data_store.initialize(find_files())
        master_df = data_store.get_dataframe('master')
        if master_df is None or master_df.empty:
            return jsonify({'rows': [], 'projects': [], 'categories': [], 'disciplines': []})

        rows, projects, categories, disciplines = _build_issue_prediction_rows(
            master_df, project, category, discipline
        )

        return jsonify({
            'rows': clean_for_json(rows),
            'projects': projects,
            'categories': categories,
            'disciplines': disciplines,
            'user_discipline': _load_person_discipline_map().get(session.get('username')),
            'is_admin': session.get('username') in ISSUE_PREDICTION_ADMIN_USERS
        })

    except Exception as e:
        safe_log(f"❌ خطا در /api/issue-predictions: {e}")
        safe_log(traceback.format_exc())
        return jsonify({'error': str(e), 'rows': []}), 500


@app.route('/api/issue-predictions/set', methods=['POST'])
@update_activity
def set_issue_prediction():
    try:
        current_user = session.get('username')
        if not current_user:
            return jsonify({'error': 'ابتدا وارد شوید'}), 401

        payload = request.get_json(force=True) or {}
        doc_no = str(payload.get('document_no', '')).strip()
        predicted_date_str = str(payload.get('predicted_date', '')).strip()
        if not doc_no or not predicted_date_str:
            return jsonify({'error': 'document_no و predicted_date الزامی هستند'}), 400

        try:
            predicted_date = pd.to_datetime(predicted_date_str).strftime('%Y-%m-%d')
        except Exception:
            return jsonify({'error': 'فرمت تاریخ نامعتبر است'}), 400

        data_store.initialize(find_files())
        master_df = data_store.get_dataframe('master')
        latest = issue_prediction.get_latest_snapshot(master_df)
        doc_row = latest[latest['Document No.'].astype(str) == doc_no]
        if doc_row.empty:
            return jsonify({'error': 'مدرک یافت نشد'}), 404
        doc_row = doc_row.iloc[0]

        # ===== احراز هویت: کاربر فقط برای دیسیپلین خودش می‌تونه پیش‌بینی بزنه =====
        # استثنا: ادمین (احسان برازنده راد) می‌تونه برای همه‌ی دیسیپلین‌ها
        # ثبت/ویرایش کنه، بدون محدودیت تطبیق دیسیپلین.
        ADMIN_USERS = ISSUE_PREDICTION_ADMIN_USERS
        is_admin = current_user in ADMIN_USERS

        if not is_admin:
            user_discipline_map = _load_person_discipline_map()
            user_discipline = user_discipline_map.get(current_user)
            doc_discipline = str(doc_row.get('Discipline', '')).strip()
            if not user_discipline or normalize_name(user_discipline) != normalize_name(doc_discipline):
                return jsonify({'error': f'شما فقط می‌توانید برای دیسیپلین «{user_discipline or "نامشخص"}» پیش‌بینی ثبت کنید، نه «{doc_discipline}»'}), 403

        cycle_anchor = issue_prediction.compute_cycle_anchor(doc_row)
        existing = _get_prediction_record(doc_no)

        if existing and existing.get('cycle_anchor') == cycle_anchor:
            edit_count = existing.get('edit_count', 0)
            if edit_count >= 2:
                return jsonify({'error': 'دیگر امکان تغییر پیش‌بینی این مدرک وجود ندارد (یک‌بار ثبت + یک‌بار ویرایش استفاده شده)'}), 400
            history = existing.get('history', [])
        else:
            # چرخه‌ی جدید (یا اولین‌بار) - از صفر شروع می‌شه
            edit_count = 0
            history = []

        history.append({'predicted_date': predicted_date, 'set_by': current_user, 'set_at': datetime.now().isoformat()})

        record = {
            'predicted_date': predicted_date,
            'cycle_anchor': cycle_anchor,
            'edit_count': edit_count + 1,
            'set_by': current_user,
            'set_at': datetime.now().isoformat(),
            'history': history,
        }
        _save_prediction_record(doc_no, record)

        return jsonify({'success': True, 'predicted_date': predicted_date, 'is_changed': (edit_count + 1) >= 2})

    except Exception as e:
        safe_log(f"❌ خطا در /api/issue-predictions/set: {e}")
        safe_log(traceback.format_exc())
        return jsonify({'error': str(e)}), 500


@app.route('/api/issue-predictions/export')
@update_activity
def export_issue_predictions():
    """
    خروجی اکسل از مدارک فعال (با همون فیلترهای مولتی‌سلکت پروژه/دسته/دیسیپلین)
    برای پرکردن دستی «تاریخ پیش‌بینی صدور» توسط دیسیپلین‌ها و ایمپورت دوباره
    از طریق /api/issue-predictions/import.
    """
    try:
        project = [v.strip() for v in request.args.getlist('project') if v.strip()] or ['همه']
        category = [v.strip() for v in request.args.getlist('category') if v.strip()] or ['همه']
        discipline = [v.strip() for v in request.args.getlist('discipline') if v.strip()] or ['همه']

        data_store.initialize(find_files())
        master_df = data_store.get_dataframe('master')
        if master_df is None or master_df.empty:
            return jsonify({'error': 'داده‌ای برای خروجی گرفتن وجود ندارد'}), 400

        rows, _, _, _ = _build_issue_prediction_rows(master_df, project, category, discipline)
        if not rows:
            return jsonify({'error': 'با این فیلترها مدرکی یافت نشد'}), 404

        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Protection
        from openpyxl.utils import get_column_letter

        wb = Workbook()
        ws = wb.active
        ws.title = 'پیش‌بینی صدور'
        ws.sheet_view.rightToLeft = True

        headers = ['ردیف', 'پروژه', 'دیسیپلین', 'دسته', 'شماره مدرک', 'عنوان مدرک',
                   'پیشرفت (%)', 'تاریخ پیش‌بینی فعلی', 'تاریخ پیش‌بینی جدید (این ستون را پر کنید)']
        col_widths = [6, 16, 14, 12, 26, 40, 10, 18, 30]

        header_fill = PatternFill(start_color='1A237E', end_color='1A237E', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True, size=11)
        editable_fill = PatternFill(start_color='FFF9C4', end_color='FFF9C4', fill_type='solid')
        locked_protection = Protection(locked=True)
        unlocked_protection = Protection(locked=False)

        for col_idx, title in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=title)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.protection = locked_protection
            ws.column_dimensions[get_column_letter(col_idx)].width = col_widths[col_idx - 1]

        EDITABLE_COL = 9  # تاریخ پیش‌بینی جدید — تنها ستونی که قفل نیست

        for r, row in enumerate(rows, start=2):
            values = [
                row['id'], row.get('project') or '', row.get('discipline') or '',
                row.get('category') or '', row.get('document_no') or '',
                row.get('document_title') or '',
                row.get('document_progress') if row.get('document_progress') is not None else '',
                row.get('predicted_issue_date') or '',
                '',
            ]
            for col_idx, val in enumerate(values, start=1):
                cell = ws.cell(row=r, column=col_idx, value=val)
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=(col_idx in (5, 6)))
                if col_idx == EDITABLE_COL:
                    cell.protection = unlocked_protection
                    cell.fill = editable_fill
                    cell.number_format = 'yyyy-mm-dd'
                else:
                    cell.protection = locked_protection

        ws.freeze_panes = 'A2'
        # قفل کل شیت به‌جز ستون تاریخ جدید — بدون رمز، فقط جلوی جابه‌جایی/پاک
        # شدن تصادفیِ شماره مدرک و بقیه‌ی ستون‌ها رو می‌گیره.
        ws.protection.sheet = True

        help_ws = wb.create_sheet('راهنما')
        help_ws.sheet_view.rightToLeft = True
        help_lines = [
            'راهنمای استفاده:',
            '۱) فقط ستون زرد «تاریخ پیش‌بینی جدید» رو پر کنید (فرمت تاریخ میلادی، مثلاً 2026-09-15).',
            '۲) ستون «شماره مدرک» کلید اصلی برای ایمپورت دوباره است — تغییرش ندید و ردیف جابه‌جا نکنید.',
            '۳) ردیف‌هایی که پر نشوند، دست‌نخورده باقی می‌مانند (هیچ تاریخی برایشان ثبت نمی‌شود).',
            '۴) فایل نهایی را از همان صفحه‌ی «تاریخ پیش‌بینی صدور» با دکمه‌ی «ایمپورت اکسل» آپلود کنید.',
        ]
        for i, line in enumerate(help_lines, start=1):
            c = help_ws.cell(row=i, column=1, value=line)
            c.alignment = Alignment(horizontal='right')
        help_ws.column_dimensions['A'].width = 95

        export_dir = os.path.join(BASE_DIR, 'exports')
        os.makedirs(export_dir, exist_ok=True)
        # پاک‌سازی فایل‌های خروجی قدیمی (بیش از ۲۴ ساعت) تا فضای دیسک پر نشود
        try:
            cutoff = time.time() - 86400
            for old_f in glob.glob(os.path.join(export_dir, 'issue_prediction_export_*.xlsx')):
                if os.path.getmtime(old_f) < cutoff:
                    os.remove(old_f)
        except Exception:
            pass

        filename = f"issue_prediction_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        filepath = os.path.join(export_dir, filename)
        wb.save(filepath)

        return send_file(filepath, as_attachment=True, download_name=filename)

    except Exception as e:
        safe_log(f"❌ خطا در /api/issue-predictions/export: {e}")
        safe_log(traceback.format_exc())
        return jsonify({'error': str(e)}), 500


@app.route('/api/issue-predictions/import', methods=['POST'])
@update_activity
def import_issue_predictions():
    """
    ایمپورت فایل اکسلی که از /api/issue-predictions/export گرفته شده و توسط
    دیسیپلین‌ها دستی پر شده. چون این یک عملیات دسته‌جمعی روی مدارک همه‌ی
    دیسیپلین‌هاست (نه فقط دیسیپلین خودِ کاربر)، فقط ادمین مجاز است.
    قوانین همون قوانین ثبت تکی رو رعایت می‌کنه: هر مدرک حداکثر یک‌بار ثبت +
    یک‌بار ویرایش در هر چرخه (cycle_anchor)، و ردیف‌های خالی نادیده گرفته می‌شن.
    """
    try:
        current_user = session.get('username')
        if not current_user:
            return jsonify({'error': 'ابتدا وارد شوید'}), 401
        if current_user not in ISSUE_PREDICTION_ADMIN_USERS:
            return jsonify({'error': 'فقط ادمین می‌تواند ایمپورت دسته‌جمعی انجام دهد'}), 403

        if 'file' not in request.files:
            return jsonify({'error': 'فایلی ارسال نشده است'}), 400
        file = request.files['file']
        if not file or not file.filename:
            return jsonify({'error': 'فایلی انتخاب نشده است'}), 400

        try:
            df = pd.read_excel(file, engine='openpyxl', sheet_name=0)
        except Exception as e:
            return jsonify({'error': f'خواندن فایل اکسل ناموفق بود: {e}'}), 400

        df.columns = [str(c).strip() for c in df.columns]
        doc_col = next((c for c in df.columns if 'شماره مدرک' in c), None)
        date_col = next((c for c in df.columns if 'جدید' in c), None)
        if not doc_col or not date_col:
            return jsonify({'error': 'ستون «شماره مدرک» یا «تاریخ پیش‌بینی جدید» در فایل پیدا نشد — از همون فایلی که از دکمه‌ی خروجی اکسل گرفتید استفاده کنید'}), 400

        data_store.initialize(find_files())
        master_df = data_store.get_dataframe('master')
        if master_df is None or master_df.empty:
            return jsonify({'error': 'داده‌ی مدارک در دسترس نیست'}), 400
        latest = issue_prediction.get_latest_snapshot(master_df)

        imported = 0
        skipped = []

        for _, r in df.iterrows():
            doc_no = str(r.get(doc_col, '')).strip()
            date_val = r.get(date_col)

            if not doc_no or doc_no.lower() == 'nan':
                continue
            if pd.isna(date_val) or str(date_val).strip() == '':
                continue  # ردیف پر نشده — نادیده گرفته می‌شه، خطا حساب نمی‌شه

            try:
                predicted_date = pd.to_datetime(date_val).strftime('%Y-%m-%d')
            except Exception:
                skipped.append({'document_no': doc_no, 'reason': 'فرمت تاریخ نامعتبر'})
                continue

            doc_row_df = latest[latest['Document No.'].astype(str) == doc_no]
            if doc_row_df.empty:
                skipped.append({'document_no': doc_no, 'reason': 'مدرک در داده‌ی فعلی یافت نشد (شاید Finished/حذف شده)'})
                continue
            doc_row = doc_row_df.iloc[0]

            cycle_anchor = issue_prediction.compute_cycle_anchor(doc_row)
            existing = _get_prediction_record(doc_no)

            if existing and existing.get('cycle_anchor') == cycle_anchor:
                edit_count = existing.get('edit_count', 0)
                if edit_count >= 2:
                    skipped.append({'document_no': doc_no, 'reason': 'ظرفیت ثبت/ویرایش این مدرک قبلاً تمام شده'})
                    continue
                history = existing.get('history', [])
            else:
                edit_count = 0
                history = []

            history.append({
                'predicted_date': predicted_date, 'set_by': current_user,
                'set_at': datetime.now().isoformat(), 'via': 'excel_import',
            })

            record = {
                'predicted_date': predicted_date,
                'cycle_anchor': cycle_anchor,
                'edit_count': edit_count + 1,
                'set_by': current_user,
                'set_at': datetime.now().isoformat(),
                'history': history,
            }
            _save_prediction_record(doc_no, record)
            imported += 1

        return jsonify({
            'success': True,
            'imported': imported,
            'skipped_count': len(skipped),
            'skipped': skipped[:50],  # جلوگیری از پاسخ خیلی بزرگ در ایمپورت‌های حجیم
        })

    except Exception as e:
        safe_log(f"❌ خطا در /api/issue-predictions/import: {e}")
        safe_log(traceback.format_exc())
        return jsonify({'error': str(e)}), 500


def _load_holidays_set():
    """
    فایل تعطیلات (Holidays.xlsx) رو از کنار app.py می‌خونه، اگه وجود نداشته
    باشه یه مجموعه‌ی خالی برمی‌گردونه (یعنی فقط پنج‌شنبه/جمعه حساب می‌شه).
    """
    holidays_path = os.path.join(BASE_DIR, 'Holidays.xlsx')
    if not os.path.exists(holidays_path):
        return set()
    try:
        return due_date_engine.load_holidays(holidays_path)
    except Exception as e:
        safe_log(f"[leaderboard] خطا در خواندن فایل تعطیلات: {e}", level="warning")
        return set()


def _load_person_discipline_map():
    """
    یک دیکشنری {نام فارسیِ کاربر (همون که تو session.get('username') میاد)
    -> دیسیپلین} می‌سازه. چون Role.xlsx با اسم انگلیسی و سیستم لاگین با اسم
    فارسی کار می‌کنه، از NAME_MAPPING به‌عنوان پل بین این دو استفاده می‌شه.
    """
    role_path = os.path.join(BASE_DIR, 'Role.xlsx')
    if not os.path.exists(role_path):
        return {}
    try:
        df = pd.read_excel(role_path, engine='openpyxl')
        name_col = discipline_col = None
        for c in df.columns:
            if str(c).strip().lower() in ('person name', 'name', 'نام'):
                name_col = c
            if str(c).strip().lower() in ('discipline', 'دیسیپلین'):
                discipline_col = c
        if not name_col or not discipline_col:
            return {}

        # نام انگلیسی (بدون کد نقش) -> نام فارسی، از روی NAME_MAPPING
        english_to_persian = {}
        for key, persian in NAME_MAPPING.items():
            plain_english = normalize_name(key)
            english_to_persian[plain_english] = persian

        result = {}
        for _, row in df.iterrows():
            eng_name = row.get(name_col)
            discipline = row.get(discipline_col)
            if pd.isna(eng_name) or pd.isna(discipline):
                continue
            normalized_eng = normalize_name(eng_name)
            persian_name = english_to_persian.get(normalized_eng)
            if persian_name:
                result[persian_name] = str(discipline).strip()
        return result
    except Exception as e:
        safe_log(f"[issue-prediction] خطا در ساخت جدول دیسیپلین کاربران: {e}", level="warning")
        return {}


def _load_role_name_map():
    """
    فایل Role.xlsx (کنار app.py) رو می‌خونه و یه دیکشنری
    {نام نرمال‌شده -> نام رسمیِ proper-case} می‌سازه.
    این فایل مرجع اصلیِ اسم افراد برای تخته امتیازات محسوب می‌شه.
    """
    role_path = os.path.join(BASE_DIR, 'Role.xlsx')
    if not os.path.exists(role_path):
        return {}
    try:
        df = pd.read_excel(role_path, engine='openpyxl')
        name_col = None
        for c in df.columns:
            if str(c).strip().lower() in ('person name', 'name', 'نام'):
                name_col = c
                break
        if name_col is None:
            return {}

        mapping = {}
        for raw in df[name_col].dropna():
            canon = str(raw).strip()
            if not canon:
                continue
            proper = _to_proper_case(canon)
            mapping[normalize_name(canon)] = proper
        return mapping
    except Exception as e:
        safe_log(f"[leaderboard] خطا در خواندن Role.xlsx: {e}", level="warning")
        return {}


def _to_proper_case(name: str) -> str:
    """حرف اول هر کلمه بزرگ، بقیه کوچک (برای اسم‌های لاتین). روی متن فارسی بی‌اثره."""
    return ' '.join(w[:1].upper() + w[1:].lower() if w else w for w in str(name).strip().split())


def _make_name_resolver():
    """
    فایل Role.xlsx رو فقط یک‌بار می‌خونه و یک تابع resolve_name برمی‌گردونه.

    نکته‌ی مهم: فقط کسایی که تو Role.xlsx هستن (یعنی کارکنان فعلی) وارد
    سیستم امتیازدهی می‌شن. هرکسی که تو تاریخچه هست ولی تو Role.xlsx نیست
    (مثلاً کسی که از شرکت رفته) به‌طور کامل از محاسبات حذف می‌شه — نه فقط
    اسمش عوض نمی‌شه، بلکه اصلاً دیگه تو تخته امتیازات دیده نمی‌شه.
    """
    role_map = _load_role_name_map()

    def resolve(raw_name):
        if not raw_name:
            return None
        raw_name = str(raw_name).strip()
        normalized = normalize_name(raw_name)
        return role_map.get(normalized)  # None یعنی تو Role.xlsx نیست → حذف می‌شه

    return resolve


def _serialize_details(durations_df):
    """durations دیتافریم رو به یک دیکشنری {("person","role"): [رکورد, ...]} تبدیل می‌کنه."""
    if durations_df.empty:
        return {}

    def to_str(v):
        if pd.isna(v):
            return None
        if hasattr(v, 'strftime'):
            return v.strftime('%Y-%m-%d')
        return str(v)

    grouped = {}
    for _, row in durations_df.iterrows():
        key = f"{row['person']}|{row['role']}"
        grouped.setdefault(key, []).append({
            'document_no': row['Document No.'],
            'revision': row['revision_group'],
            'duration_days': round(float(row['duration_days']), 2) if pd.notna(row['duration_days']) else None,
            'role_avg_duration': round(float(row['role_avg_duration']), 2) if pd.notna(row.get('role_avg_duration')) else None,
            'role_budget_days': round(float(row['role_budget_days']), 2) if pd.notna(row.get('role_budget_days')) else None,
            'T0': to_str(row.get('T0')),
            'due_date': to_str(row.get('due_date')),
            'close_date': to_str(row.get('close_date')),
            'delay_days': (int(row['delay_days']) if pd.notna(row.get('delay_days')) else None),
            'is_late': bool(row.get('is_late')) if pd.notna(row.get('is_late')) else False,
            'is_primary_delay_cause': bool(row.get('is_primary_delay_cause')),
            'total_revisions_for_doc': (int(row['total_revisions_for_doc']) if pd.notna(row.get('total_revisions_for_doc')) else None),
            'is_combined_role': bool(row.get('is_combined_role')),
        })
    return grouped


def _serialize_distribute_details(distribute_df):
    """distribute_df رو به دیکشنری {"person": [رکورد, ...]} تبدیل می‌کنه (برای مودال جزئیات)."""
    if distribute_df is None or distribute_df.empty:
        return {}

    def to_str(v):
        if pd.isna(v):
            return None
        if hasattr(v, 'strftime'):
            return v.strftime('%Y-%m-%d')
        return str(v)

    grouped = {}
    for _, row in distribute_df.iterrows():
        person = row['to_discipline_person']
        grouped.setdefault(person, []).append({
            'document_no': row['Document No.'],
            'from_person': row['from_discipline_person'],
            'distribute_date': to_str(row.get('distribute_date')),
            'response_date': to_str(row.get('response_date')),
            'response_days': round(float(row['response_days']), 2) if pd.notna(row.get('response_days')) else None,
            'points': int(row['points']) if pd.notna(row.get('points')) else 0,
            'outcome': row.get('outcome'),
        })
    return grouped


def _filter_history_by_project(history_df, project):
    """فیلتر history_df بر اساس پروژه (اگه 'همه' بود، بدون تغییر برمی‌گرده)."""
    if not project or project == 'همه':
        return history_df
    project_col = find_column(history_df, ['Project'])
    if not project_col:
        return history_df
    mask = history_df[project_col].apply(lambda p: normalize_project(p) == normalize_project(project))
    return history_df[mask]


def _get_project_list(history_df):
    project_col = find_column(history_df, ['Project'])
    if not project_col:
        return []
    return sorted(set(
        str(p).strip() for p in history_df[project_col].dropna().unique()
        if str(p).strip()
    ))


def _load_org_roster():
    import org_structure
    return org_structure.load_org_roster(
        os.path.join(BASE_DIR, 'Role.xlsx'),
        NAME_MAPPING,
        normalize_name,
        _to_proper_case,
    )


def _role_code_label(code):
    mapping = {
        'MDJ': 'کارشناس',
        'MDS': 'مهندس ارشد',
        'EM': 'مدیر مهندسی',
        'PROCUREMENT': 'تدارکات',
        'PMO': 'دفتر مدیریت پروژه',
        'CEO': 'مدیرعامل',
        'DCC': 'کنترل مدارک',
        'ENGINEER': 'مهندس',
        'CTR': 'پیمانکار',
        'specialist': 'کارشناس',
        'senior': 'مهندس ارشد',
        'manager': 'مدیر مهندسی',
    }
    if not code:
        return ''
    return mapping.get(str(code).strip(), str(code).strip())


def _build_people_directory(scores=None):
    """نمایه افراد از Role.xlsx + NAME_MAPPING + نقش واقعی در تاریخچه پروژه‌ها."""
    from collections import defaultdict
    from revision_metrics import extract_role_code
    import org_structure

    directory = {}

    def ensure(name):
        key = normalize_name(name)
        if not key:
            return None
        person = directory.get(key)
        if person is None:
            person = {
                'name': _to_proper_case(name),
                'persian_name': '',
                'aliases': [],
                'discipline': '',
                'org_role': '',
                'org_role_label': '',
                'score_role': '',
                'score_role_label': '',
                'total_score': 0,
                'on_time_score': None,
                'quality_score': None,
                'speed_score': None,
                'volume_score': None,
                'n_revisions_touched': 0,
                'delay_causes_count': 0,
                'avg_revisions_needed': None,
                'avg_duration_days': None,
                'peer_avg_duration': None,
                'role_budget_days': None,
                'max_revisions_in_role': 0,
                'is_combined_role': False,
                'peer_group': '',
                'shared_distribute_points': 0,
                'shared_prediction_points': 0,
                'assignments': [],
            }
            directory[key] = person
        return person

    def add_alias(person, alias):
        if not person or not alias:
            return
        text = str(alias).strip()
        if text and text not in person['aliases'] and text != person['name'] and text != person.get('persian_name'):
            person['aliases'].append(text)

    for raw, persian in NAME_MAPPING.items():
        person = ensure(raw)
        if not person:
            continue
        person['persian_name'] = person['persian_name'] or persian
        add_alias(person, raw)
        add_alias(person, persian)

    for item in _load_org_roster():
        person = ensure(item.get('name'))
        if not person:
            continue
        person['name'] = item.get('name') or person['name']
        if item.get('persian_name'):
            person['persian_name'] = item['persian_name']
        if item.get('discipline') and item.get('discipline') != 'نامشخص' and not org_structure.is_excluded_discipline(item['discipline']):
            person['discipline'] = item['discipline']
        if item.get('role'):
            person['org_role'] = item['role']
            person['org_role_label'] = _role_code_label(item['role'])
        add_alias(person, item.get('persian_name'))
        if item.get('project') and item.get('role'):
            person['assignments'].append({
                'project': item['project'],
                'role_code': str(item['role']).upper(),
                'role_label': _role_code_label(item['role']),
            })

    resolver = _make_name_resolver()
    history_df = data_store.get_dataframe('history')
    if history_df is not None and not history_df.empty:
        to_col = find_column(history_df, ['To Name'])
        from_col = find_column(history_df, ['From Name', 'From_Name', 'From name', 'From'])
        proj_col = find_column(history_df, ['Project'])
        seen_assign = set()
        for col in (to_col, from_col):
            if not col:
                continue
            use_cols = [col] + ([proj_col] if proj_col else [])
            pairs = history_df[use_cols].dropna(subset=[col]).drop_duplicates()
            for rec in pairs.itertuples(index=False):
                raw = rec[0]
                if pd.isna(raw) or not str(raw).strip():
                    continue
                project = str(rec[1]).strip() if proj_col and len(rec) > 1 else ''
                plain, code = extract_role_code(str(raw))
                resolved = resolver(plain) or resolver(raw)
                person = ensure(resolved or plain)
                if not person:
                    continue
                add_alias(person, plain)
                if not project or not code:
                    continue
                stamp = (person['name'], project, str(code).upper())
                if stamp in seen_assign:
                    continue
                seen_assign.add(stamp)
                person['assignments'].append({
                    'project': project,
                    'role_code': str(code).upper(),
                    'role_label': _role_code_label(code),
                })

    role_rank = {'EM': 0, 'manager': 0, 'MDS': 1, 'senior': 1, 'MDJ': 2, 'specialist': 2}
    for person in directory.values():
        uniq = []
        seen_roles = set()
        for item in person['assignments']:
            stamp = (item.get('project'), str(item.get('role_code') or '').upper())
            if stamp in seen_roles:
                continue
            seen_roles.add(stamp)
            uniq.append(item)
        uniq.sort(key=lambda a: (role_rank.get(a.get('role_code'), 9), a.get('project') or ''))
        person['assignments'] = uniq[:12]

    for row in scores or []:
        person = ensure(row.get('person'))
        if not person:
            continue
        person['score_role'] = row.get('role') or person['score_role']
        person['score_role_label'] = _role_code_label(person['score_role'])
        try:
            person['total_score'] = max(person['total_score'], int(round(float(row.get('total_score') or 0))))
        except (TypeError, ValueError):
            pass
        person['core_score'] = row.get('core_score')
        person['on_time_score'] = row.get('on_time_score')
        person['quality_score'] = row.get('quality_score')
        person['speed_score'] = row.get('speed_score')
        person['volume_score'] = row.get('volume_score')
        person['n_revisions_touched'] = row.get('n_revisions_touched') or 0
        person['delay_causes_count'] = row.get('delay_causes_count') or 0
        person['avg_revisions_needed'] = row.get('avg_revisions_needed')
        person['avg_duration_days'] = row.get('avg_duration_days')
        person['peer_avg_duration'] = row.get('peer_avg_duration')
        person['role_budget_days'] = row.get('role_budget_days')
        person['max_revisions_in_role'] = row.get('max_revisions_in_role') or 0
        person['is_combined_role'] = bool(row.get('is_combined_role'))
        person['peer_group'] = row.get('peer_group') or person.get('peer_group') or ''
        person['shared_distribute_points'] = row.get('shared_distribute_points') or 0
        person['shared_prediction_points'] = row.get('shared_prediction_points') or 0
        disc = row.get('discipline') or person['discipline']
        if disc and not org_structure.is_excluded_discipline(disc):
            person['discipline'] = disc
        elif row.get('role') == 'manager':
            person['discipline'] = 'مدیر مهندسی'

    people = list(directory.values())
    people.sort(key=lambda p: p.get('total_score') or 0, reverse=True)
    return people


def _compute_prediction_discipline_points():
    """
    امتیاز پیش‌بینی فقط اگر حداقل ۵ روز قبل از صدور ثبت شده باشد.
    هر مدرک ±۵ می‌گیرد و این مقدار یک‌بار به دیسیپلین همان مدرک اضافه می‌شود؛
    بین اعضای دیسیپلین ضرب نمی‌شود.
    """
    from collections import defaultdict
    import org_structure
    predictions = _get_all_prediction_records()
    if not predictions:
        return {}

    try:
        master_df = data_store.get_dataframe('master')
        if master_df is None or master_df.empty:
            return {}
        latest = issue_prediction.get_latest_snapshot(master_df)
        latest['Document No.'] = latest['Document No.'].astype(str)
        latest_by_doc = latest.set_index('Document No.')
    except Exception as e:
        safe_log(f"[issue-prediction] خطا در محاسبه‌ی امتیاز دقت پیش‌بینی: {e}", level="warning")
        return {}

    by_disc = defaultdict(lambda: {'points': 0, 'docs': []})
    for doc_no, record in predictions.items():
        if str(doc_no) not in latest_by_doc.index:
            continue
        row = latest_by_doc.loc[str(doc_no)]
        if isinstance(row, pd.DataFrame):
            row = row.iloc[0]
        issued_date = row.get('Issued Date')
        if pd.isna(issued_date):
            continue

        predicted_date_str = record.get('predicted_date')
        if not predicted_date_str:
            continue
        registered_at = issue_prediction.prediction_registered_at(record)
        try:
            pts = issue_prediction.compute_prediction_accuracy_points(
                pd.to_datetime(predicted_date_str).date(),
                pd.to_datetime(issued_date).date(),
                registered_at=registered_at,
            )
        except Exception:
            continue
        if pts == 0:
            continue

        discipline = org_structure.canonicalize_discipline(row.get('Discipline') or 'نامشخص')
        if org_structure.is_excluded_discipline(discipline):
            continue
        by_disc[discipline]['points'] += pts
        by_disc[discipline]['docs'].append({
            'document_no': str(doc_no),
            'points': pts,
            'predicted_date': predicted_date_str,
            'issued_date': str(pd.to_datetime(issued_date).date()),
        })
    return dict(by_disc)


def _org_portfolio(project, scores=None):
    import org_structure
    try:
        people = _build_people_directory(scores or [])
        return org_structure.build_project_portfolios(people, scores or [], project_name=project or 'همه')
    except Exception as e:
        safe_log(f"[leaderboard] خطا در پورتفولیو سازمانی: {e}", level="warning")
        return {'project': project or 'همه', 'projects': [], 'managers': []}


# ==================== تنظیمات امتیازدهی (وزن‌ها + قوانین بونوس سفارشی) ====================
# ویرایش‌شان فقط برای ادمین است چون مستقیماً روی رتبه‌بندی واقعی افراد اثر
# می‌گذارد. وزن‌ها و بونوس‌ها در Redis ذخیره می‌شوند و هر بار که leaderboard
# محاسبه می‌شود (چه امروز، چه گذشته) دوباره اعمال می‌شوند — یعنی هیچ اسنپ‌شاتِ
# قدیمیِ ثابتی وجود ندارد که نیاز به migration داشته باشد.
SCORING_WEIGHTS_KEY = "scoring_config:weights"
SCORING_BONUS_RULES_KEY = "scoring_config:bonus_rules"

BONUS_EVENT_TYPES = {
    'issue': {'label': 'خروج از اینباکس (Issue کردن مدرک)', 'log_status': 'Issue'},
    'assign': {'label': 'دریافت مدرک (Assign)', 'log_status': 'Assign'},
    'comment': {'label': 'ثبت کامنت (Comment)', 'log_status': 'Comment'},
    'distribute_action': {'label': 'ارسال Distribute به همکار', 'log_status': 'Distribute'},
}


def _get_scoring_weights():
    if not redis_client:
        return dict(scoring.SCORE_WEIGHTS)
    try:
        raw = redis_client.get(SCORING_WEIGHTS_KEY)
        if raw:
            return scoring._normalize_weights(json.loads(raw))
    except Exception as e:
        safe_log(f"[scoring-config] خطا در خواندن وزن‌ها: {e}", level="warning")
    return dict(scoring.SCORE_WEIGHTS)


def _save_scoring_weights(weights):
    normalized = scoring._normalize_weights(weights)
    if redis_client:
        try:
            redis_client.set(SCORING_WEIGHTS_KEY, json.dumps(normalized))
        except Exception as e:
            safe_log(f"[scoring-config] خطا در ذخیره‌ی وزن‌ها: {e}", level="warning")
    return normalized


def _get_bonus_rules():
    if not redis_client:
        return []
    try:
        raw = redis_client.get(SCORING_BONUS_RULES_KEY)
        return json.loads(raw) if raw else []
    except Exception as e:
        safe_log(f"[scoring-config] خطا در خواندن قوانین بونوس: {e}", level="warning")
        return []


def _save_bonus_rules(rules):
    if redis_client:
        try:
            redis_client.set(SCORING_BONUS_RULES_KEY, json.dumps(rules, ensure_ascii=False))
        except Exception as e:
            safe_log(f"[scoring-config] خطا در ذخیره‌ی قوانین بونوس: {e}", level="warning")


def _get_scoring_config_version() -> str:
    """
    یه اثرانگشت کوتاه از وضعیت فعلیِ وزن‌ها + قوانین بونوس. هر جای کد که
    نتیجه‌ی امتیازدهی (leaderboard یا خلاصه‌ی هوش مصنوعی) رو کش می‌کنه،
    این توکن باید تو کلید کش باشه — وگرنه بعد از تغییر تنظیمات تو
    /scoring-settings، همون نتیجه‌ی قدیمی از کش برمی‌گرده و تغییر اصلاً
    دیده نمی‌شه.
    """
    import hashlib
    try:
        payload = json.dumps(
            {'weights': _get_scoring_weights(), 'rules': _get_bonus_rules()},
            sort_keys=True, default=str,
        )
        return hashlib.md5(payload.encode('utf-8')).hexdigest()[:10]
    except Exception as e:
        safe_log(f"[scoring-config] خطا در ساخت نسخه‌ی تنظیمات: {e}", level="warning")
        return "noversion"


def _compute_event_counts(df, log_status, date_from=None, date_to=None):
    """
    تعداد دفعاتی که هر نفر یک نوع رویداد خاص (مثلاً Issue) رو انجام داده،
    از روی ستون From Name شمارش می‌کنه — چون تو این تراکنش‌ها، From Name
    یعنی «چه کسی این کار رو انجام داد» (همون قراردادی که تو revision_metrics
    برای تشخیص نقش‌ها هم استفاده شده).
    """
    import person_activity
    if df is None or df.empty:
        return {}
    needed = ['Action Date', 'From Name', 'Log Status']
    if any(c not in df.columns for c in needed):
        return {}

    x = df[needed].copy()
    x['Action Date'] = pd.to_datetime(x['Action Date'], errors='coerce')
    x = x.dropna(subset=['Action Date'])
    x = x[x['Log Status'].astype(str).str.strip() == log_status]

    if date_from is not None:
        x = x[x['Action Date'] >= pd.Timestamp(date_from)]
    if date_to is not None:
        x = x[x['Action Date'] < pd.Timestamp(date_to) + pd.Timedelta(days=1)]

    x['person'] = x['From Name'].apply(person_activity._extract_name)
    x = x.dropna(subset=['person'])
    if x.empty:
        return {}
    return x.groupby('person').size().to_dict()


def _compute_custom_bonus_points(df, date_from=None, date_to=None, resolve_name=None):
    """
    مجموع امتیاز همه‌ی قوانین بونوس فعال (سفارشی، فردی — نه تیمی) رو به
    ازای هر نفر حساب می‌کنه: {person: extra_points}.
    """
    rules = [r for r in _get_bonus_rules() if r.get('enabled', True)]
    if not rules:
        return {}

    totals = defaultdict(float)
    for rule in rules:
        event = BONUS_EVENT_TYPES.get(rule.get('event_type'))
        if not event:
            continue
        counts = _compute_event_counts(df, event['log_status'], date_from=date_from, date_to=date_to)
        try:
            points_per_unit = float(rule.get('points_per_unit') or 0)
        except (TypeError, ValueError):
            continue
        for person, count in counts.items():
            name = resolve_name(person) if resolve_name else person
            if not name:
                continue
            totals[name] += count * points_per_unit

    return dict(totals)


@app.route('/api/scoring-config/weights', methods=['GET'])
@update_activity
def get_scoring_weights_route():
    if session.get('username') not in DASHBOARD_ADMIN_USERS:
        return jsonify({'error': 'این بخش فقط برای ادمین است'}), 403
    return jsonify({
        'weights': _get_scoring_weights(),
        'defaults': dict(scoring.SCORE_WEIGHTS),
        'can_edit': True,
    })


@app.route('/api/scoring-config/weights', methods=['POST'])
@update_activity
def save_scoring_weights_route():
    current_user = session.get('username')
    if current_user not in DASHBOARD_ADMIN_USERS:
        return jsonify({'error': 'فقط ادمین می‌تواند وزن‌های امتیازدهی را تغییر دهد'}), 403
    payload = request.get_json(silent=True) or {}
    weights = payload.get('weights') or {}
    if not isinstance(weights, dict) or not weights:
        return jsonify({'error': 'مقادیر وزن نامعتبر است'}), 400
    normalized = _save_scoring_weights(weights)
    return jsonify({'success': True, 'weights': normalized})


@app.route('/api/scoring-config/bonus-rules', methods=['GET'])
@update_activity
def get_bonus_rules_route():
    if session.get('username') not in DASHBOARD_ADMIN_USERS:
        return jsonify({'error': 'این بخش فقط برای ادمین است'}), 403
    return jsonify({
        'rules': _get_bonus_rules(),
        'event_types': [{'key': k, 'label': v['label']} for k, v in BONUS_EVENT_TYPES.items()],
        'can_edit': True,
    })


@app.route('/api/scoring-config/bonus-rules', methods=['POST'])
@update_activity
def add_bonus_rule_route():
    try:
        current_user = session.get('username')
        if current_user not in DASHBOARD_ADMIN_USERS:
            return jsonify({'error': 'فقط ادمین می‌تواند قانون بونوس اضافه کند'}), 403

        payload = request.get_json(silent=True) or {}
        name = (payload.get('name') or '').strip()
        event_type = payload.get('event_type')
        try:
            points_per_unit = float(payload.get('points_per_unit'))
        except (TypeError, ValueError):
            return jsonify({'error': 'امتیاز به‌ازای هر واحد نامعتبر است'}), 400
        if not name:
            return jsonify({'error': 'نام قانون خالی است'}), 400
        if event_type not in BONUS_EVENT_TYPES:
            return jsonify({'error': 'نوع رویداد نامعتبر است'}), 400

        rules = _get_bonus_rules()
        rule = {
            'id': f"{int(time.time() * 1000)}_{uuid.uuid4().hex[:6]}",
            'name': name,
            'event_type': event_type,
            'points_per_unit': points_per_unit,
            'enabled': True,
            'added_by': current_user,
            'added_at': datetime.now().isoformat(),
        }
        rules.append(rule)
        _save_bonus_rules(rules)
        return jsonify({'success': True, 'rule': rule})
    except Exception as e:
        safe_log(f"[scoring-config] خطا در افزودن قانون بونوس: {e}", level="error")
        safe_log(traceback.format_exc(), level="error")
        return jsonify({'error': str(e)}), 500


@app.route('/api/scoring-config/bonus-rules/<rule_id>', methods=['PATCH'])
@update_activity
def update_bonus_rule_route(rule_id):
    current_user = session.get('username')
    if current_user not in DASHBOARD_ADMIN_USERS:
        return jsonify({'error': 'فقط ادمین می‌تواند قانون بونوس را ویرایش کند'}), 403
    payload = request.get_json(silent=True) or {}
    rules = _get_bonus_rules()
    found = False
    for r in rules:
        if r.get('id') == rule_id:
            found = True
            if 'enabled' in payload:
                r['enabled'] = bool(payload['enabled'])
            if 'points_per_unit' in payload:
                try:
                    r['points_per_unit'] = float(payload['points_per_unit'])
                except (TypeError, ValueError):
                    pass
            if 'name' in payload and str(payload['name']).strip():
                r['name'] = str(payload['name']).strip()
    if not found:
        return jsonify({'error': 'قانون پیدا نشد'}), 404
    _save_bonus_rules(rules)
    return jsonify({'success': True, 'rules': rules})


@app.route('/api/scoring-config/bonus-rules/<rule_id>', methods=['DELETE'])
@update_activity
def delete_bonus_rule_route(rule_id):
    current_user = session.get('username')
    if current_user not in DASHBOARD_ADMIN_USERS:
        return jsonify({'error': 'فقط ادمین می‌تواند قانون بونوس را حذف کند'}), 403
    rules = [r for r in _get_bonus_rules() if r.get('id') != rule_id]
    _save_bonus_rules(rules)
    return jsonify({'success': True, 'rules': rules})


@app.route('/scoring-settings')
def scoring_settings_page():
    return render_template('scoring_settings.html')


def _build_leaderboard_result(project, date_from=None, date_to=None):
    import org_structure

    data_store.initialize(find_files())
    history_df = data_store.get_dataframe('history')
    empty_portfolio = {'project': project, 'projects': [], 'managers': []}
    if history_df is None or history_df.empty:
        return {'scores': [], 'details': {}, 'distribute_details': {}, 'projects': [],
                'disciplines': [], 'portfolio': empty_portfolio,
                'message': 'داده‌ای برای محاسبه یافت نشد'}

    project_list = _get_project_list(history_df)
    filtered_df = _filter_history_by_project(history_df, project)
    if filtered_df.empty:
        return {'scores': [], 'details': {}, 'distribute_details': {}, 'projects': project_list,
                'disciplines': [], 'portfolio': _org_portfolio(project, []),
                'message': 'مدرکی برای این پروژه یافت نشد'}

    holidays = _load_holidays_set()
    scores_df, durations_df, distribute_df = scoring.compute_all_metrics(
        filtered_df, holidays=holidays, due_days=5,
        return_details=True, resolve_name=_make_name_resolver(),
        date_from=date_from, date_to=date_to,
        weights=_get_scoring_weights(),
    )
    if scores_df.empty:
        return {'scores': [], 'details': {}, 'distribute_details': {}, 'projects': project_list,
                'disciplines': [], 'portfolio': _org_portfolio(project, []),
                'message': 'موردی برای امتیازدهی در این بازه یافت نشد'}

    scores = scores_df.to_dict(orient='records')
    roster = _load_org_roster()
    prediction_by_disc = _compute_prediction_discipline_points()
    scores, _ = org_structure.apply_shared_team_bonuses(
        scores, distribute_df, prediction_by_disc, roster, resolve_name=_make_name_resolver()
    )

    # ===== بونوس‌های سفارشیِ فردی (از /scoring-settings) — جمعی/تیمی نیستن،
    # مستقیم به امتیاز خودِ همون فرد اضافه می‌شن، جدا از بونوس‌های مشترکِ
    # Distribute/پیش‌بینی که در سطح دیسیپلین تقسیم می‌شن.
    custom_bonus = _compute_custom_bonus_points(
        filtered_df, date_from=date_from, date_to=date_to, resolve_name=_make_name_resolver()
    )
    for row in scores:
        extra = custom_bonus.get(row.get('person'), 0)
        row['custom_bonus_points'] = round(extra, 1)
        if extra:
            row['total_score'] = round((row.get('total_score') or 0) + extra, 1)

    if any(custom_bonus.values()):
        # rank_in_role قبلاً یک‌بار (بدون بونوس سفارشی) تو apply_shared_team_bonuses
        # محاسبه شده؛ چون total_score عوض شد، باید دوباره حساب بشه وگرنه رتبه‌ها
        # با عددِ نهایی که نمایش داده می‌شه هم‌خوانی ندارن.
        for role in {r.get('role') for r in scores}:
            role_rows = [r for r in scores if r.get('role') == role]
            role_rows.sort(key=lambda r: r.get('total_score', 0), reverse=True)
            for idx, row in enumerate(role_rows, start=1):
                row['rank_in_role'] = idx

    keep = (
        'person', 'role', 'org_role', 'discipline', 'total_score', 'core_score',
        'shared_distribute_points', 'shared_prediction_points', 'custom_bonus_points', 'rank_in_role',
        'on_time_score', 'quality_score', 'speed_score', 'volume_score',
        'n_revisions_touched', 'delay_causes_count', 'avg_revisions_needed',
        'avg_duration_days', 'peer_avg_duration', 'role_budget_days',
        'max_revisions_in_role', 'is_combined_role', 'peer_group',
    )
    compact = []
    for row in scores:
        item = {}
        for key in keep:
            value = row.get(key)
            try:
                if value is not None and pd.isna(value):
                    value = None
            except (TypeError, ValueError):
                pass
            item[key] = value
        compact.append(item)
    scores = compact
    disciplines = org_structure.build_discipline_groups(scores)
    portfolio = _org_portfolio(project, scores)
    return {
        'scores': scores,
        'details': _serialize_details(durations_df),
        'distribute_details': _serialize_distribute_details(distribute_df),
        'prediction_by_discipline': prediction_by_disc,
        'projects': project_list,
        'disciplines': disciplines,
        'portfolio': portfolio,
    }


def _compute_prediction_accuracy_bonus():
    """سازگاری قدیمی: دیگر استفاده نمی‌شود؛ امتیاز در سطح دیسیپلین حساب می‌شود."""
    return {}


@app.route('/api/competitive-scores')
@update_activity
def get_competitive_scores():
    try:
        project = (request.args.get('project', 'همه') or 'همه').strip()
        date_from_str = request.args.get('date_from', '').strip()
        date_to_str = request.args.get('date_to', '').strip()
        date_from = pd.to_datetime(date_from_str).date() if date_from_str else None
        date_to = pd.to_datetime(date_to_str).date() if date_to_str else None

        data_version = data_store.get_data_version()
        scoring_version = _get_scoring_config_version()
        cache_key = f"leaderboard_v14:{data_version}:{scoring_version}:{project}:{date_from_str or 'all'}:{date_to_str or 'all'}"
        cached_item = cache_get(cache_key)
        if cached_item is not None:
            return jsonify({
                'scores': cached_item.get('scores', []),
                'projects': cached_item.get('projects', []),
                'disciplines': cached_item.get('disciplines', []),
                'portfolio': cached_item.get('portfolio') or {},
            })

        result = _build_leaderboard_result(project, date_from, date_to)
        cache_set(cache_key, result)
        return jsonify({
            'scores': result.get('scores', []),
            'projects': result.get('projects', []),
            'disciplines': result.get('disciplines', []),
            'portfolio': result.get('portfolio') or {},
            'message': result.get('message'),
        })

    except Exception as e:
        safe_log(f"❌ خطا در /api/competitive-scores: {e}")
        safe_log(traceback.format_exc())
        return jsonify({'error': str(e), 'scores': []}), 500


@app.route('/api/competitive-scores/details')
@update_activity
def get_competitive_score_details():
    """
    جزئیات مدرک‌به‌مدرکِ امتیاز یک نفر (برای نمایش تو مودال جزئیات).
    پارامترها: person, role, project (اختیاری)
    """
    try:
        person = request.args.get('person', '').strip()
        role = request.args.get('role', '').strip()
        project = (request.args.get('project', 'همه') or 'همه').strip()
        date_from_str = request.args.get('date_from', '').strip()
        date_to_str = request.args.get('date_to', '').strip()
        date_from = pd.to_datetime(date_from_str).date() if date_from_str else None
        date_to = pd.to_datetime(date_to_str).date() if date_to_str else None
        if not person or not role:
            return jsonify({'error': 'person و role الزامی هستند', 'details': []}), 400

        data_version = data_store.get_data_version()
        scoring_version = _get_scoring_config_version()
        cache_key = f"leaderboard_v14:{data_version}:{scoring_version}:{project}:{date_from_str or 'all'}:{date_to_str or 'all'}"
        cached_item = cache_get(cache_key)

        if cached_item is None:
            cached_item = _build_leaderboard_result(project, date_from, date_to)
            cache_set(cache_key, cached_item)

        all_details = cached_item.get('details', {})
        key = f"{person}|{role}"
        docs = all_details.get(key, [])
        docs = sorted(docs, key=lambda d: (not d.get('is_primary_delay_cause'), d.get('close_date') or ''), reverse=False)

        score_row = next((s for s in (cached_item.get('scores') or []) if s.get('person') == person and s.get('role') == role), {})
        discipline = score_row.get('discipline')
        distribute_docs = cached_item.get('distribute_details', {}).get(person, [])
        if discipline:
            for member in (cached_item.get('scores') or []):
                if member.get('discipline') == discipline:
                    extra = cached_item.get('distribute_details', {}).get(member.get('person'), [])
                    for event in extra:
                        if event not in distribute_docs:
                            distribute_docs.append(event)

        pred_docs = (cached_item.get('prediction_by_discipline') or {}).get(discipline, {}).get('docs', []) if discipline else []

        return jsonify({
            'person': person,
            'role': role,
            'discipline': discipline,
            'core_score': score_row.get('core_score'),
            'shared_distribute_points': score_row.get('shared_distribute_points', 0),
            'shared_prediction_points': score_row.get('shared_prediction_points', 0),
            'total_score': score_row.get('total_score'),
            'details': docs,
            'distribute_details': distribute_docs,
            'prediction_details': pred_docs,
        })

    except Exception as e:
        safe_log(f"❌ خطا در /api/competitive-scores/details: {e}")
        safe_log(traceback.format_exc())
        return jsonify({'error': str(e), 'details': []}), 500


@app.route('/api/data')
@update_activity
def get_data():
    project = request.args.get('project', 'همه')
    period = int(request.args.get('period', '30'))
    doc_type = request.args.get('type', 'همه')

    data_version = data_store.get_data_version()
    cache_key = f"data:{data_version}:{project}-{period}-{doc_type}"
    cached_item = cache_get(cache_key)
    if cached_item is not None:
        return jsonify(cached_item)

    # ----------------------- Local safe helpers -----------------------
    def to_ts(v):
        t = pd.to_datetime(v, errors='coerce')
        return t if pd.notna(t) else pd.NaT

    def to_date(v):
        t = to_ts(v)
        return t.date() if pd.notna(t) else None

    def is_yes(v):
        return str(v).strip().lower() == 'yes'

    def norm_str(v):
        return str(v).strip().lower() if v is not None else ''

    def safe_doc_type(doc):
        return str(doc.get('doc_type', '')).strip().upper()

    def safe_doc_no(doc):
        return str(doc.get('document_no', '')).strip()

    def safe_disc(doc):
        return (get_doc_discipline(doc) or 'نامشخص').strip()

    def typed_key_of(doc):
        dt = safe_doc_type(doc) or 'MASTER'
        dn = safe_doc_no(doc)
        return f"{dt}::{dn}" if dn else None

    def project_match(doc_project, project_filter):
        if project_filter == 'همه':
            return True
        return norm_str(doc_project) == norm_str(project_filter)

    def doc_type_match(doc, doc_type_filter):
        dt = safe_doc_type(doc)
        if doc_type_filter == 'اصلی':
            return dt == 'MASTER'
        if doc_type_filter == 'وندور':
            return dt == 'VENDOR'
        return True

    def filter_overdue(overdue_list, project_filter, doc_type_filter):
        out = []
        for d in (overdue_list or []):
            if not isinstance(d, dict):
                continue
            if not project_match(d.get('project'), project_filter):
                continue
            if not doc_type_match(d, doc_type_filter):
                continue
            out.append(d)
        return out

    def _safe_weight(v, default=0.0):
        try:
            x = float(v or 0)
            if pd.isna(x) or x <= 0:
                return default
            return x
        except Exception:
            return default

    def vendor_weight_with_fallback(doc):
        w = _safe_weight(doc.get('weight'), 0.0)
        if w > 0:
            return w
        ew = _safe_weight(doc.get('eng_weight'), 0.0)
        if ew > 0:
            return ew
        return 1.0

    def eng_weight(doc):
        return _safe_weight(doc.get('eng_weight'), 0.0)

    def normalize_progress(v):
        try:
            x = float(v)
            if pd.isna(x):
                return 0.0
            return max(0.0, min(100.0, x))
        except Exception:
            return 0.0

    def safe_change(current, old):
        try:
            c = float(current or 0)
            o = float(old or 0)
            if pd.isna(c): c = 0.0
            if pd.isna(o): o = 0.0
            return round(c - o, 2)
        except Exception:
            return 0.0

    def resolve_baseline_date(all_dates_sorted, latest_date, period_days, ref_date=None):
        """
        تاریخ baseline = نزدیک‌ترین تاریخ <= (latest_date - period_days)
        اگر نبود => None
        """
        if not all_dates_sorted:
            return None

        target = latest_date - timedelta(days=period_days)
        candidates = [d for d in all_dates_sorted if d <= target]
        if candidates:
            return max(candidates)

        # fallback: اگر تاریخ قدیمی نداریم، None بده تا change=0 شود
        return None

    def build_current_weight_maps(latest_docs):
        eng_w = {}
        ven_w = {}
        for d in latest_docs:
            tk = typed_key_of(d)
            if not tk:
                continue
            dt = safe_doc_type(d)
            if dt == 'MASTER':
                w = eng_weight(d)
                if w > 0:
                    eng_w[tk] = w
            elif dt == 'VENDOR':
                w = vendor_weight_with_fallback(d)
                if w > 0:
                    ven_w[tk] = w
        return eng_w, ven_w, {**eng_w, **ven_w}

    def build_discipline_weight_map(disc_latest_docs):
        """وزن ثابت فعلی برای مقایسه تاریخی (راهکار A)"""
        weights = {}
        for d in disc_latest_docs:
            tk = typed_key_of(d)
            if not tk:
                continue
            dt = safe_doc_type(d)
            if dt == 'MASTER':
                w = eng_weight(d)
            elif dt == 'VENDOR':
                w = vendor_weight_with_fallback(d)
            else:
                continue
            if w > 0:
                weights[tk] = w
        return weights
    def split_typed_weights(all_weights, doc_type):
        prefix = f"{doc_type}::"
        return {tk: w for tk, w in all_weights.items() if tk.startswith(prefix)}
    def get_latest_typed_docs_on_or_before(versions, target_date):
        """آخرین نسخه هر مدرک تا target_date — کلید: MASTER::xxx / VENDOR::xxx"""
        latest = {}
        if target_date is None:
            return latest
        for doc in versions:
            tk = typed_key_of(doc)
            if not tk:
                continue
            dts = to_ts(doc.get('date'))
            if pd.isna(dts) or dts.date() > target_date:
                continue
            prev = latest.get(tk)
            if prev is None:
                latest[tk] = doc
                continue
            prev_ts = to_ts(prev.get('date'))
            if pd.isna(prev_ts) or dts > prev_ts:
                latest[tk] = doc
        return latest
    # ----------------------- Main -----------------------
    try:
        # ---------- Load & integrate (یک‌بار به ازای هر نسخه داده) ----------
        data, integrated = get_processed_snapshot()
        cached_after_snapshot = cache_get(cache_key)
        if cached_after_snapshot is not None:
            return jsonify(cached_after_snapshot)

        overall_progress_eng = integrated.get('engineering_progress', 0)
        overall_progress_vendor = integrated.get('vendor_progress', 0)

        # ---------- Base docs ----------
        all_docs_raw = (integrated.get('master_docs', []) or []) + (integrated.get('vendor_docs', []) or [])
        all_docs_raw = [d for d in all_docs_raw if isinstance(d, dict) and safe_doc_no(d)]

        all_docs_filtered = [d for d in all_docs_raw if project_match(d.get('project'), project)]
        all_docs_filtered = [d for d in all_docs_filtered if doc_type_match(d, doc_type)]

        # ---------- Overdue ----------
        overdue_docs_filtered = filter_overdue(integrated.get('overdue_docs', []), project, doc_type)
        overdue_client_filtered = filter_overdue(integrated.get('overdue_client', []), project, doc_type)
        overdue_contractor_filtered = filter_overdue(integrated.get('overdue_contractor', []), project, doc_type)

        # ---------- Latest record per typed key ----------
        latest_records = {}
        for doc in all_docs_filtered:
            tk = typed_key_of(doc)
            if not tk:
                continue
            dts = to_ts(doc.get('date'))
            if pd.isna(dts):
                continue

            prev = latest_records.get(tk)
            if prev is None:
                latest_records[tk] = doc
                continue

            prev_ts = to_ts(prev.get('date'))
            if pd.isna(prev_ts) or dts > prev_ts:
                latest_records[tk] = doc

        all_latest_docs = list(latest_records.values())

        deleted_docs = [d for d in all_latest_docs if is_yes(d.get('deleted', ''))]
        active_docs = [d for d in all_latest_docs if not is_yes(d.get('deleted', ''))]

        # ---------- All valid dates ----------
        all_dates = sorted(set(
            dd for dd in (to_date(d.get('date')) for d in all_docs_filtered) if dd is not None
        ))

        # ---- دیسیپلین‌ها ----
        disc_docs_map = defaultdict(lambda: {'MASTER': [], 'VENDOR': []})
        for d in all_latest_docs:
            disc = safe_disc(d)
            dt = safe_doc_type(d)
            if dt in ('MASTER', 'VENDOR'):
                disc_docs_map[disc][dt].append(d)

        disc_stats = {}
        import org_structure
        for disc, bucket in disc_docs_map.items():
            if org_structure.is_excluded_discipline(disc):
                continue
            eng_docs = bucket.get('MASTER', [])
            ven_docs = bucket.get('VENDOR', [])

            # ===== محاسبه جداگانه پیشرفت =====
            eng_progress = calculate_adjusted_progress(eng_docs) if eng_docs else 0.0
            ven_progress = calculate_vendor_progress(ven_docs) if ven_docs else 0.0
            
            # پیشرفت ترکیبی (میانگین وزنی بر اساس تعداد مدارک)
            total_docs = len(eng_docs) + len(ven_docs)
            if total_docs > 0:
                combined_progress = round(
                    (eng_progress * len(eng_docs) + ven_progress * len(ven_docs)) / total_docs, 
                    2
                )
            else:
                combined_progress = 0.0

            active_disc_docs = [x for x in (eng_docs + ven_docs) if not is_yes(x.get('deleted', ''))]

            not_issued_disc = sum(
                1 for x in active_disc_docs
                if get_responsible_category(x.get('responsible')) == 'not_issued'
                and pd.isna(x.get('issued_date'))
            )
            with_customer_disc = sum(
                1 for x in active_disc_docs
                if get_responsible_category(x.get('responsible')) == 'with_customer'
                and pd.isna(x.get('comment_date'))
            )
            approved_disc = sum(
                1 for x in active_disc_docs
                if get_responsible_category(x.get('responsible')) == 'approved'
                and normalize_progress(x.get('doc_progress', 0)) == 100
            )
            deleted_disc = sum(1 for x in (eng_docs + ven_docs) if is_yes(x.get('deleted', '')))

            disc_stats[disc] = {
                'total': len(eng_docs) + len(ven_docs),
                'active': len(active_disc_docs),
                'deleted': deleted_disc,
                'with_customer': with_customer_disc,
                'approved': approved_disc,
                'not_issued': not_issued_disc,
                'progress': combined_progress,
                'progress_engineering': round(eng_progress, 2),
                'progress_vendor': round(ven_progress, 2),
                'progress_history': {}
            }

        # ===== تاریخچه پیشرفت کلی =====
        progress_history = {}
        combined_current_weights = {}
        latest_date = None

        if all_dates:
            latest_date = max(all_dates)
            # ===== محاسبه وزن‌های فعلی برای پیشرفت کلی (با استفاده از تابع موجود) =====
            current_eng_weights, current_vendor_weights, combined_current_weights = build_current_weight_maps(all_latest_docs)

        # ===== تاریخچه پیشرفت برای هر دیسیپلین (با تفکیک مهندسی/وندور) =====
        # با استفاده از منطق ساده (مشابه API جدید) برای جلوگیری از تغییرات کاذب

        for disc_name in disc_stats.keys():
            disc_latest_docs = [d for d in all_latest_docs if safe_disc(d) == disc_name]
            if not disc_latest_docs:
                continue

            # تفکیک مدارک بر اساس نوع
            eng_latest = [d for d in disc_latest_docs if safe_doc_type(d) == 'MASTER']
            ven_latest = [d for d in disc_latest_docs if safe_doc_type(d) == 'VENDOR']
            
            disc_all_versions = [d for d in all_docs_filtered if safe_disc(d) == disc_name]
            
            # ===== پیشرفت فعلی با استفاده از calculate_adjusted_progress =====
            current_disc_progress = calculate_adjusted_progress(disc_latest_docs)
            current_eng_progress = calculate_adjusted_progress(eng_latest) if eng_latest else 0.0
            current_ven_progress = calculate_vendor_progress(ven_latest) if ven_latest else 0.0

            # ===== تاریخچه برای هر دوره =====
            disc_history = {}
            
            for period_days in [7, 30]:
                # اگر all_dates وجود نداشته باشد، نمی‌توانیم تاریخچه محاسبه کنیم
                if not all_dates or latest_date is None:
                    disc_history[f'last_{period_days}_days'] = {
                        'date': datetime.now().strftime('%Y-%m-%d'),
                        'progress': round(current_disc_progress, 2),
                        'progress_engineering': round(current_eng_progress, 2),
                        'progress_vendor': round(current_ven_progress, 2),
                        'change': 0.0,
                        'change_engineering': 0.0,
                        'change_vendor': 0.0
                    }
                    continue
                
                baseline_date = resolve_baseline_date(all_dates, latest_date, period_days)
                
                if baseline_date is None:
                    disc_history[f'last_{period_days}_days'] = {
                        'date': latest_date.strftime('%Y-%m-%d'),
                        'progress': round(current_disc_progress, 2),
                        'progress_engineering': round(current_eng_progress, 2),
                        'progress_vendor': round(current_ven_progress, 2),
                        'change': 0.0,
                        'change_engineering': 0.0,
                        'change_vendor': 0.0
                    }
                    continue
                
                # ===== محدودیت ۳ روز برای جلوگیری از تغییرات کاذب =====
                #target_date = latest_date - timedelta(days=period_days)
                #days_diff = abs((baseline_date - target_date).days)
                #if days_diff > 3:
                    #disc_history[f'last_{period_days}_days'] = {
                        #'date': latest_date.strftime('%Y-%m-%d'),
                        #'progress': round(current_disc_progress, 2),
                        #'progress_engineering': round(current_eng_progress, 2),
                        #'progress_vendor': round(current_ven_progress, 2),
                        #'change': 0.0,
                        #'change_engineering': 0.0,
                        #'change_vendor': 0.0
                    #}
                    #continue
                
                # ===== وزن ثابت فعلی (A) + cohort (B) =====
                disc_weights = build_discipline_weight_map(disc_latest_docs)
                eng_weights = split_typed_weights(disc_weights, 'MASTER')
                ven_weights = split_typed_weights(disc_weights, 'VENDOR')
                # پیشرفت فعلی با وزن ثابت
                current_disc_fw = (
                    get_progress_at_date_fixed_weight(disc_all_versions, latest_date, disc_weights)
                    if disc_weights and latest_date else current_disc_progress
                )
                current_eng_fw = (
                    get_progress_at_date_fixed_weight(disc_all_versions, latest_date, eng_weights)
                    if eng_weights and latest_date else current_eng_progress
                )
                current_ven_fw = (
                    get_progress_at_date_fixed_weight(disc_all_versions, latest_date, ven_weights)
                    if ven_weights and latest_date else current_ven_progress
                )
                # پیشرفت در baseline با همان وزن‌های فعلی
                old_disc_fw = (
                    get_progress_at_date_fixed_weight(disc_all_versions, baseline_date, disc_weights)
                    if disc_weights else 0.0
                )
                old_eng_fw = (
                    get_progress_at_date_fixed_weight(disc_all_versions, baseline_date, eng_weights)
                    if eng_weights else 0.0
                )
                old_ven_fw = (
                    get_progress_at_date_fixed_weight(disc_all_versions, baseline_date, ven_weights)
                    if ven_weights else 0.0
                )
                # --- Cohort: فقط مدارکی که در baseline_date وجود داشتند (B) ---
                baseline_latest_map = get_latest_typed_docs_on_or_before(
                    disc_all_versions, baseline_date
                )
                cohort_keys = set(baseline_latest_map.keys())
                current_latest_map = get_latest_typed_docs_on_or_before(
                    disc_all_versions, latest_date
                )
                new_doc_keys = set(current_latest_map.keys()) - cohort_keys
                cohort_weights = {tk: w for tk, w in disc_weights.items() if tk in cohort_keys}
                cohort_eng_weights = {tk: w for tk, w in eng_weights.items() if tk in cohort_keys}
                cohort_ven_weights = {tk: w for tk, w in ven_weights.items() if tk in cohort_keys}
                current_cohort = (
                    get_progress_at_date_fixed_weight(disc_all_versions, latest_date, cohort_weights)
                    if cohort_weights and latest_date else current_disc_fw
                )
                old_cohort = (
                    get_progress_at_date_fixed_weight(disc_all_versions, baseline_date, cohort_weights)
                    if cohort_weights else old_disc_fw
                )
                change_full = safe_change(current_disc_fw, old_disc_fw)
                change_cohort = safe_change(current_cohort, old_cohort)
                disc_history[f'last_{period_days}_days'] = {
                    'date': baseline_date.strftime('%Y-%m-%d'),
                    # همان فیلدهای قبلی — برای سازگاری با UI
                    'progress': round(old_disc_fw, 2),
                    'progress_engineering': round(old_eng_fw, 2),
                    'progress_vendor': round(old_ven_fw, 2),
                    'change': change_full,
                    'change_engineering': safe_change(current_eng_fw, old_eng_fw),
                    'change_vendor': safe_change(current_ven_fw, old_ven_fw),
                    # فیلدهای جدید (B)
                    'change_cohort': change_cohort,
                    'new_docs_count': len(new_doc_keys),
                    'new_docs_impact': round(change_full - change_cohort, 2),
                }
            disc_stats[disc_name]['progress_history'] = disc_history
            disc_stats[disc_name]['progress'] = round(current_disc_progress, 2)
            disc_stats[disc_name]['progress_engineering'] = round(current_eng_progress, 2)
            disc_stats[disc_name]['progress_vendor'] = round(current_ven_progress, 2)

        # ===== محاسبه پیشرفت کلی (با استفاده از وزن‌های فعلی) =====
        current_overall_progress = (
            get_progress_at_date_fixed_weight(all_docs_filtered, latest_date, combined_current_weights)
            if combined_current_weights and latest_date else 0.0
        )

        # ===== تاریخچه پیشرفت کلی =====
        if all_dates and latest_date is not None:
            for period_days in [7, 30]:
                baseline_date = resolve_baseline_date(all_dates, latest_date, period_days)

                if baseline_date is None:
                    progress_history[f'last_{period_days}_days'] = {
                        'date': latest_date.strftime('%Y-%m-%d'),
                        'progress': round(current_overall_progress, 2),
                        'change': 0.0,
                        'change_cohort': 0.0,
                        'new_docs_count': 0,
                        'new_docs_impact': 0.0,
                        'progress_cohort': round(current_overall_progress, 2),
                        'current_cohort': round(current_overall_progress, 2),
                        'current_progress': round(current_overall_progress, 2),
                    }
                    continue

                # --- A: fixed-weight (همان وزن‌های فعلی) ---
                old_overall_fw = (
                    get_progress_at_date_fixed_weight(
                        all_docs_filtered, baseline_date, combined_current_weights
                    )
                    if combined_current_weights else 0.0
                )
                current_overall_fw = (
                    get_progress_at_date_fixed_weight(
                        all_docs_filtered, latest_date, combined_current_weights
                    )
                    if combined_current_weights else current_overall_progress
                )
                change_full = safe_change(current_overall_fw, old_overall_fw)

                # --- B: cohort (فقط مدارکی که در baseline_date وجود داشتند) ---
                baseline_latest_map = get_latest_typed_docs_on_or_before(
                    all_docs_filtered, baseline_date
                )
                current_latest_map = get_latest_typed_docs_on_or_before(
                    all_docs_filtered, latest_date
                )
                cohort_keys = set(baseline_latest_map.keys())
                new_doc_keys = set(current_latest_map.keys()) - cohort_keys

                cohort_weights = {
                    tk: w for tk, w in combined_current_weights.items()
                    if tk in cohort_keys
                }

                current_cohort = (
                    get_progress_at_date_fixed_weight(
                        all_docs_filtered, latest_date, cohort_weights
                    )
                    if cohort_weights else current_overall_fw
                )
                old_cohort = (
                    get_progress_at_date_fixed_weight(
                        all_docs_filtered, baseline_date, cohort_weights
                    )
                    if cohort_weights else old_overall_fw
                )
                change_cohort = safe_change(current_cohort, old_cohort)

                progress_history[f'last_{period_days}_days'] = {
                    'date': baseline_date.strftime('%Y-%m-%d'),
                    # کل (fixed-weight)
                    'progress': round(old_overall_fw, 2),
                    'current_progress': round(current_overall_fw, 2),
                    'change': change_full,
                    # cohort (مدارک موجود)
                    'progress_cohort': round(old_cohort, 2),
                    'current_cohort': round(current_cohort, 2),
                    'change_cohort': change_cohort,
                    'new_docs_count': len(new_doc_keys),
                    'new_docs_impact': round(change_full - change_cohort, 2),
                }

        # ---------- Overall stats ----------
        overall_progress_simple = calculate_project_average_progress(all_latest_docs, integrated.get('projects', []))
        overall_progress = overall_progress_simple

        stats = {
            'total': len(active_docs),
            'not_issued': sum(
                1 for d in active_docs
                if get_responsible_category(d.get('responsible')) == 'not_issued' and pd.isna(d.get('issued_date'))
            ),
            'with_customer': sum(
                1 for d in active_docs
                if get_responsible_category(d.get('responsible')) == 'with_customer' and pd.isna(d.get('comment_date'))
            ),
            'approved': sum(
                1 for d in active_docs
                if get_responsible_category(d.get('responsible')) == 'approved'
                and normalize_progress(d.get('doc_progress', 0)) == 100
            ),
            'hold': sum(1 for d in active_docs if is_yes(d.get('hold', ''))),
            'deleted': len(deleted_docs),
            'overall_progress': round(overall_progress, 2),
            'overall_progress_engineering': round(overall_progress_eng, 2),
            'overall_progress_vendor': round(overall_progress_vendor, 2),
            'progress_history': progress_history
        }

        # ---------- Inbox ----------
        inbox_stats = calculate_inbox_stats(
            data.get('history'),
            data.get('history_cols', {}),
            data.get('vendor_history'),
            data.get('vendor_history_cols', {}),
            data.get('persons'),
            project_filter=project if project != 'همه' else None
        )

        # ---------- Package stats ----------
        package_stats = {}
        filtered_docs = active_docs

        if isinstance(integrated.get('package_stats'), dict):
            try:
                for proj, pkgs in integrated['package_stats'].items():
                    if project != 'همه' and normalize_project(proj) != normalize_project(project):
                        continue
                    package_stats.setdefault(proj, {})
                    for pkg_norm, st in pkgs.items():
                        package_stats[proj][pkg_norm] = {
                            'count': int(st.get('count', 0)),
                            'progress': float(st.get('progress', 0)),
                            'vendor': str(st.get('vendor', 'نامشخص')),
                            'weight_sum': float(st.get('weight_sum', 0)),
                            'weighted_progress': float(st.get('weighted_progress', 0)),
                            'original_name': str(st.get('original_name', pkg_norm)),
                            'original_vendor': str(st.get('original_vendor', st.get('vendor', 'نامشخص')))
                        }
            except Exception:
                package_stats = {}

        if not package_stats:
            for doc in filtered_docs:
                pkg = doc.get('package_name')
                if not pkg:
                    continue

                proj = doc.get('project', 'نامشخص')
                if not proj or pd.isna(proj):
                    proj = 'نامشخص'
                if project != 'همه' and normalize_project(proj) != normalize_project(project):
                    continue

                pkg_norm = normalize_package(pkg)
                vendor_name = doc.get('vendor_name', 'نامشخص')
                vendor_norm = normalize_vendor(vendor_name)

                package_stats.setdefault(proj, {})
                package_stats[proj].setdefault(pkg_norm, {
                    'count': 0,
                    'progress': 0,
                    'vendor': vendor_norm,
                    'vendors': set(),
                    'weight_sum': 0.0,
                    'weighted_progress': 0.0,
                    'original_name': pkg,
                    'original_vendor': vendor_name
                })

                entry = package_stats[proj][pkg_norm]
                entry['count'] += 1

                w = _safe_weight(doc.get('weight'), 0.0)
                p = normalize_progress(doc.get('doc_progress', 0))
                if w > 0:
                    entry['weight_sum'] += w
                    entry['weighted_progress'] += (w * p)

                entry['vendors'].add(vendor_norm)

            for proj_key, pkgs in package_stats.items():
                for pkg_key, st in pkgs.items():
                    st['progress'] = round((st['weighted_progress'] / st['weight_sum']), 2) if st['weight_sum'] > 0 else 0
                    st['vendors'] = list(st['vendors'])

        response_data = {
            'projects': integrated.get('projects'),
            'disciplines': integrated.get('disciplines'),
            'projects_details': integrated.get('projects_details'),
            'stats': stats,
            'discipline_stats': disc_stats,
            'latest_overall_progress': integrated.get('latest_overall_progress'),
            'inbox_stats': inbox_stats,
            'package_stats': package_stats,
            'overdue_docs': overdue_docs_filtered,
            'overdue_client': overdue_client_filtered,
            'overdue_contractor': overdue_contractor_filtered,
            'all_users': integrated.get('all_users')
        }

        cache_set(cache_key, response_data)

        return jsonify(response_data)

    except Exception as e:
        import traceback
        safe_log(f"Error in get_data: {e}", level="error")
        safe_log(traceback.format_exc(), level="error")
        return jsonify({'error': str(e)}), 500

    
@app.route('/api/upload', methods=['POST'])
def upload_file():
    try:
        if 'file' not in request.files:
            return jsonify({'error': 'فایلی انتخاب نشده است'}), 400
        file = request.files['file']
        if file.filename == '':
            return jsonify({'error': 'نام فایل خالی است'}), 400
        filepath = os.path.join(UPLOAD_FOLDER, file.filename)
        file.save(filepath)

        invalidate_processed_snapshot()
        data_store.clear_cache()
        try:
            df = pd.read_excel(filepath, engine='openpyxl', nrows=5)
            
        except Exception as e:
            safe_log(f" فایل آپلود شد اما خواندن آن با مشکل مواجه شد: {e}")

        # پردازش سنگین رو همین الان توی پس‌زمینه انجام بده تا کاربرِ بعدی
        # منتظر محاسبه نمونه (کش از قبل آماده باشد)
        start_cache_warm_up()

        return jsonify({'success': True, 'message': f'فایل {file.filename} با موفقیت آپلود شد'})
    except Exception as e:
        safe_log(f" خطا در آپلود: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/export')
def export_data():
    try:
        data = load_all_data()
        data = process_master(data)
        data = process_history(data)
        data = process_vendor_master(data)
        data = process_vendor_history(data)
        integrated = integrate_data(data)
        output_file = os.path.join(BASE_DIR, 'report_export.xlsx')
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            if integrated['master_docs']:
                pd.DataFrame(integrated['master_docs']).to_excel(writer, sheet_name='مدارک اصلی', index=False)
            if integrated['vendor_docs']:
                pd.DataFrame(integrated['vendor_docs']).to_excel(writer, sheet_name='مدارک وندور', index=False)
            summary = {
                'نوع': ['کل مدارک', 'مدارک اصلی', 'مدارک وندور', 'دست کارفرما', 'صادر نشده', 'تایید شده', 'هولد شده', 'حذف شده', 'میانگین پیشرفت'],
                'تعداد': [
                    integrated['stats']['total'],
                    integrated['stats']['master_count'],
                    integrated['stats']['vendor_count'],
                    integrated['stats']['with_customer'],
                    integrated['stats']['not_issued'],
                    integrated['stats']['approved'],
                    integrated['stats']['hold'],
                    integrated['stats']['deleted'],
                    integrated['stats']['avg_progress']
                ]
            }
            pd.DataFrame(summary).to_excel(writer, sheet_name='خلاصه', index=False)
        return send_file(output_file, as_attachment=True)
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# ==================== Routes چت ====================

@app.route('/api/chat/clear-all-private', methods=['POST'])
def clear_all_private_chats():
    """پاک کردن همه چت‌های خصوصی (فقط ادمین)"""
    try:
        # ===== بررسی ادمین =====
        current_user = session.get('username')
        if not current_user:
            return jsonify({'error': 'لطفاً وارد شوید'}), 401
        
        admin_names = ['admin', 'مدیر', 'ادمین', 'ehsan', 'barazande']
        if current_user.lower() not in admin_names:
            return jsonify({'error': 'تنها ادمین می‌تواند همه چت‌ها را پاک کند'}), 403
        
        # ===== پاک کردن همه چت‌های خصوصی =====
        import shutil
        if os.path.exists(PRIVATE_CHAT_DIR):
            # شمارش فایل‌ها قبل از حذف
            files = os.listdir(PRIVATE_CHAT_DIR)
            file_count = len(files)
            
            # حذف پوشه و ایجاد مجدد
            shutil.rmtree(PRIVATE_CHAT_DIR)
            os.makedirs(PRIVATE_CHAT_DIR, exist_ok=True)
            
            safe_log(f"🗑️ ادمین {current_user} همه {file_count} چت خصوصی را پاک کرد")
            
            return jsonify({
                'success': True,
                'message': f'همه {file_count} چت خصوصی با موفقیت پاک شدند',
                'count': file_count
            })
        else:
            return jsonify({'success': True, 'message': 'هیچ چت خصوصی برای پاک کردن وجود ندارد', 'count': 0})
        
    except Exception as e:
        safe_log(f"❌ خطا در پاک کردن چت‌ها: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/chat/clear-private', methods=['POST'])
def clear_private_chat():
    """پاک کردن چت خصوصی بین کاربر فعلی و کاربر دیگر"""
    try:
        # ===== دریافت کاربر فعلی =====
        current_user = session.get('username')
        if not current_user:
            return jsonify({'error': 'لطفاً وارد شوید'}), 401
        
        data = request.get_json()
        if not data:
            return jsonify({'error': 'داده‌ای ارسال نشده است'}), 400
        
        other_user = data.get('other_user', '').strip()
        if not other_user:
            return jsonify({'error': 'نام کاربر مقابل را وارد کنید'}), 400
        
        # ===== پیدا کردن فایل چت =====
        chat_file = get_private_chat_file(current_user, other_user)
        
        if os.path.exists(chat_file):
            # حذف فایل
            os.remove(chat_file)
            safe_log(f"🗑️ کاربر {current_user} چت با {other_user} را پاک کرد")
            return jsonify({
                'success': True,
                'message': f'چت با {other_user} با موفقیت پاک شد'
            })
        else:
            return jsonify({
                'success': True,
                'message': 'هیچ چتی برای پاک کردن وجود ندارد'
            })
        
    except Exception as e:
        safe_log(f"❌ خطا در پاک کردن چت: {e}")
        return jsonify({'error': str(e)}), 500

def _prepare_full_dataset():
    """
    آماده‌سازی داده‌ی خام برای محاسبه‌ی اینباکس.
    این تابع سنگین‌ترین بخش کاره (load + پردازش ۴ تا دیتافریم)،
    برای همین باید فقط یک‌بار صدا زده بشه و نتیجه‌اش برای همه‌ی
    افراد (توی گرم‌کردن کش) یا برای یک درخواست تکی استفاده بشه.
    """
    data = load_all_data()
    data = process_master(data)
    data = process_history(data)
    data = process_vendor_master(data)
    data = process_vendor_history(data)
    return data


_processed_snapshot = {'version': None, 'data': None, 'integrated': None}
_processed_snapshot_lock = threading.Lock()


def invalidate_processed_snapshot():
    with _processed_snapshot_lock:
        _processed_snapshot['version'] = None
        _processed_snapshot['data'] = None
        _processed_snapshot['integrated'] = None


def get_processed_snapshot(force=False):
    """
    لایه ۲ کش: نتیجه load + process + integrate فقط یک‌بار
    به ازای هر نسخه داده در RAM همین پردازش نگه داشته می‌شود.
    """
    with _processed_snapshot_lock:
        version = data_store.get_data_version()
        if (
            not force
            and _processed_snapshot['version'] == version
            and _processed_snapshot['data'] is not None
            and _processed_snapshot['integrated'] is not None
        ):
            return _processed_snapshot['data'], _processed_snapshot['integrated']

        data = _prepare_full_dataset()
        integrated = integrate_data(data)
        version = data_store.get_data_version()
        _processed_snapshot['version'] = version
        _processed_snapshot['data'] = data
        _processed_snapshot['integrated'] = integrated
        return data, integrated


_activity_index_cache = {'version': None, 'df': None}
_activity_index_lock = threading.Lock()


def get_activity_index():
    """
    ایندکس فعالیت افراد (برای پرسش‌های باز مثل «فلانی چیکار کرده؟» یا
    «فلانی چند درصد پیشرفت داشته تو پروژه X؟») را می‌سازد و به ازای هر
    نسخه‌ی داده فقط یک‌بار در حافظه (نه Redis، چون DataFrame است) کش می‌کند
    — دقیقاً همون الگوی get_processed_snapshot.
    """
    import person_activity

    with _activity_index_lock:
        version = data_store.get_data_version()
        if _activity_index_cache['version'] == version and _activity_index_cache['df'] is not None:
            return _activity_index_cache['df']

        data, _ = get_processed_snapshot()
        df = person_activity.build_activity_index(
            data.get('history'), data.get('vendor_history'),
            data.get('master'), data.get('vendor_master'),
        )
        _activity_index_cache['version'] = version
        _activity_index_cache['df'] = df
        return df


def _build_inbox_result(data, person_name, source_filter, project_filter):
    """
    با استفاده از داده‌ی از قبل آماده‌شده (خروجی _prepare_full_dataset)،
    جزئیات اینباکس یک فرد خاص رو می‌سازه. هیچ کار سنگینی (لود/پردازش فایل)
    اینجا انجام نمی‌شه، فقط فیلتر کردن.
    """
    person_key_norm = normalize_name(person_name)

    def extract_docs(df, cols, source_type):
        if df is None or cols is None:
            return []

        to_name_col = cols.get('to_name')
        log_status_col = cols.get('log_status')
        ongoing_col = cols.get('ongoing')
        close_date_col = cols.get('close_date')
        doc_no_col = cols.get('doc_no')
        doc_title_col = cols.get('doc_title')
        discipline_col = cols.get('discipline')
        action_date_col = cols.get('action_date')
        project_col = cols.get('project')

        if not to_name_col or not log_status_col:
            return []

        valid_statuses = ['Assign', 'Issue', 'Distribute']
        latest = {}

        for idx, row in df.iterrows():
            to_name = row.get(to_name_col)
            if pd.isna(to_name):
                continue
            if normalize_name(to_name) != person_key_norm:
                continue

            if project_filter is not None and project_col:
                proj = row.get(project_col)
                if pd.isna(proj):
                    continue
                if normalize_project(proj) != normalize_project(project_filter):
                    continue

            doc_no = row.get(doc_no_col) if doc_no_col else None
            if pd.isna(doc_no):
                continue

            log_status = row.get(log_status_col, '')
            if pd.isna(log_status):
                continue
            log_status_str = str(log_status).strip()
            if log_status_str not in valid_statuses:
                continue

            ongoing = row.get(ongoing_col, '')
            if pd.notna(ongoing) and str(ongoing).lower().strip() not in ['yes', 'بله', 'true']:
                continue

            close_date = row.get(close_date_col) if close_date_col else None
            if pd.notna(close_date):
                continue

            action_date = row.get(action_date_col) if action_date_col else None
            if pd.isna(action_date):
                continue

            if doc_no not in latest or (
                action_date is not None and
                (latest[doc_no].get('action_date') is None or
                 action_date > latest[doc_no].get('action_date'))
            ):
                latest[doc_no] = {
                    'doc_no': doc_no,
                    'doc_title': row.get(doc_title_col) if doc_title_col else None,
                    'discipline': row.get(discipline_col) if discipline_col else None,
                    'log_status': log_status_str,
                    'action_date': action_date,
                    'project': row.get(project_col) if project_col else None,
                    'source': source_type
                }
        return list(latest.values())

    if source_filter == 'مهندسی':
        history_items = extract_docs(data.get('history'), data.get('history_cols', {}), 'مهندسی')
        vendor_items = []
    elif source_filter == 'وندور':
        history_items = []
        vendor_items = extract_docs(data.get('vendor_history'), data.get('vendor_history_cols', {}), 'وندور')
    else:
        history_items = extract_docs(data.get('history'), data.get('history_cols', {}), 'مهندسی')
        vendor_items = extract_docs(data.get('vendor_history'), data.get('vendor_history_cols', {}), 'وندور')

    all_items = history_items + vendor_items

    unique = {}
    for item in all_items:
        doc_no = item['doc_no']
        if doc_no not in unique or (item['action_date'] is not None and (unique[doc_no].get('action_date') is None or item['action_date'] > unique[doc_no]['action_date'])):
            unique[doc_no] = item

    today = datetime.now().date()
    docs_list = []
    for doc_no, info in unique.items():
        action_date = info['action_date']
        if isinstance(action_date, pd.Timestamp):
            action_date = action_date.date()
        elif isinstance(action_date, datetime):
            action_date = action_date.date()
        days = (today - action_date).days if action_date else 0

        docs_list.append({
            'document_no': doc_no,
            'document_title': info.get('doc_title') or '-',
            'discipline': info.get('discipline') or '-',
            'log_status': info['log_status'],
            'days': days,
            'project': info.get('project') or '-',
            'source': info.get('source', 'نامشخص'),
            'action_date': action_date.strftime('%Y-%m-%d') if action_date else None
        })

    grouped = {}
    for doc in docs_list:
        status = doc['log_status']
        if status not in grouped:
            grouped[status] = []
        grouped[status].append(doc)

    for status in grouped:
        grouped[status].sort(key=lambda x: x['document_no'])

    person_df = data.get('persons')
    display_name = person_name
    if person_df is not None:
        person_col = find_column(person_df, ['Person Name'])
        if person_col:
            for name in person_df[person_col].dropna():
                if normalize_name(name) == person_key_norm:
                    display_name = str(name).strip()
                    break

    return {
        'person': display_name,
        'total': len(docs_list),
        'grouped': grouped
    }


@app.route('/api/inbox-details/<person_name>')
def get_inbox_details(person_name):
    source_filter = request.args.get('source', None)
    project_filter = request.args.get('project', None)
    
    if project_filter == 'همه' or not project_filter:
        project_filter = None
    
    try:
        person_name = unquote(person_name)

        # ---- کش: اگر قبلاً برای همین فرد/فیلتر محاسبه شده، همون رو برگردون ----
        data_version = data_store.get_data_version()
        cache_key = f"inbox:{data_version}:{person_name}:{source_filter}:{project_filter}"
        cached_item = cache_get(cache_key)
        if cached_item is not None:
            return jsonify(cached_item)

        data, _ = get_processed_snapshot()
        result = _build_inbox_result(data, person_name, source_filter, project_filter)
        cache_set(cache_key, result)
        return jsonify(result)

    except Exception as e:
        safe_log(f" خطا در دریافت جزئیات اینباکس: {e}")
        safe_log(traceback.format_exc())
        return jsonify({'error': str(e)}), 500

# ==================== Helper functions for changed-docs ====================

def _safe_float(v, default=0.0):
    try:
        if v is None:
            return default
        if isinstance(v, float) and math.isnan(v):
            return default
        s = str(v).strip().replace('%', '')
        return float(s) if s else default
    except Exception:
        return default


def _to_date(v):
    """Normalize mixed date values (str/datetime/date/pandas timestamp) -> date or None"""
    if v is None:
        return None
    if isinstance(v, date) and not isinstance(v, datetime):
        return v
    if isinstance(v, datetime):
        return v.date()

    try:
        if hasattr(v, "to_pydatetime"):
            return v.to_pydatetime().date()
    except Exception:
        pass

    s = str(v).strip()
    if not s or s.lower() == "nan":
        return None

    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%d-%m-%Y", "%d/%m/%Y"):
        try:
            return datetime.strptime(s[:10], fmt).date()
        except Exception:
            pass

    try:
        return datetime.fromisoformat(s.replace("Z", "")).date()
    except Exception:
        return None


def _first_nonempty_value(doc, keys):
    for key in keys:
        val = doc.get(key)
        if val is not None and val != '':
            return val
    return None


def _normalize_doc_record(doc):
    if not isinstance(doc, dict):
        return doc

    normalized = dict(doc)

    field_aliases = {
        'discipline': ['discipline', 'Discipline', 'DISCIPLINE', 'discipline_name', 'Discipline Name'],
        'document_no': ['document_no', 'Document_No', 'Document No', 'doc_no', 'Doc No', 'DOC_NO'],
        'doc_type': ['doc_type', 'Doc Type', 'DOC_TYPE'],
        'weight': ['weight', 'Weight'],
        'eng_weight': ['eng_weight', 'Eng Weight', 'ENGINEERING_WEIGHT'],
        'status': ['status', 'Status'],
        'date': ['date', 'Date', 'DATE'],
        'document_title': ['document_title', 'Document Title', 'title', 'Title'],
        'deleted': ['deleted', 'Deleted', 'DELETED'],
        'hold': ['hold', 'Hold', 'HOLD'],
        'responsible': ['responsible', 'Responsible', 'RESPONSIBLE'],
        'comment': ['comment', 'Comment', 'COMMENT'],
        'doc_progress': ['doc_progress', 'Document Progress', 'doc progress'],
    }

    for canonical, aliases in field_aliases.items():
        if canonical not in normalized or normalized.get(canonical) in (None, ''):
            val = _first_nonempty_value(doc, aliases)
            if val is not None:
                normalized[canonical] = val

    if normalized.get('discipline') is not None:
        normalized['discipline'] = str(normalized['discipline']).strip()
    if normalized.get('doc_type') is not None:
        normalized['doc_type'] = str(normalized['doc_type']).strip().upper()
    if normalized.get('date') is not None:
        normalized['date'] = _to_date(normalized['date'])
    if normalized.get('doc_progress') is not None:
        try:
            normalized['doc_progress'] = float(normalized['doc_progress'])
        except:
            normalized['doc_progress'] = 0.0

    return normalized


def get_doc_discipline(doc):
    import org_structure
    val = _first_nonempty_value(doc, ['discipline', 'Discipline', 'DISCIPLINE', 'discipline_name', 'Discipline Name'])
    if val is None:
        return ''
    return org_structure.canonicalize_discipline(val)


def _get_latest_version_on_or_before(versions, target_date):
    """Get the latest version of a document with date <= target_date"""
    best = None
    best_date = None

    for v in versions:
        d = v.get('date')
        if d is None:
            continue
        if isinstance(d, pd.Timestamp):
            d = d.date()
        elif isinstance(d, datetime):
            d = d.date()

        if d <= target_date and (best_date is None or d > best_date):
            best = v
            best_date = d

    return best


def _is_deleted(doc):
    return str(doc.get('deleted', '')).strip().lower() in ('yes', 'true', '1')


def _is_hold(doc):
    return str(doc.get('hold', '')).strip().lower() in ('yes', 'true', '1')


def load_all_docs_from_store():
    """بارگذاری همه مدارک از تمام منابع (master, vendor_master, history, vendor_history)"""
    frames = []
    sources = [
        ('master', 'MASTER', 'master'),
        ('history', 'MASTER', 'history'),
        ('vendor_master', 'VENDOR', 'vendor_master'),
        ('vendor_history', 'VENDOR', 'vendor_history'),
    ]

    for name, default_doc_type, source_name in sources:
        df = data_store.get_dataframe(name)
        if df is None or df.empty:
            continue

        df = df.copy()
        df['_source'] = source_name

        if 'doc_type' not in df.columns:
            df['doc_type'] = default_doc_type
        else:
            df['doc_type'] = df['doc_type'].fillna(default_doc_type)

        frames.append(df)

    if not frames:
        return []

    combined = pd.concat(frames, ignore_index=True, sort=False)
    records = combined.to_dict('records')
    return [_normalize_doc_record(r) for r in records]


def get_active_latest_docs_from_store():
    """بارگذاری آخرین نسخه فعال هر مدرک از master و vendor_master"""
    latest_docs = []
    sources = [
        ('master', 'MASTER'),
        ('vendor_master', 'VENDOR'),
    ]

    for name, default_doc_type in sources:
        df = data_store.get_dataframe(name)
        if df is None or df.empty:
            continue

        df = df.copy()
        if 'doc_type' not in df.columns:
            df['doc_type'] = default_doc_type
        else:
            df['doc_type'] = df['doc_type'].fillna(default_doc_type)

        records = df.to_dict('records')
        latest_docs.extend(_normalize_doc_record(r) for r in records)

    return latest_docs


def _group_docs_by_number(all_docs):
    """گروه‌بندی مدارک بر اساس document_no"""
    grouped = {}
    for doc in all_docs:
        doc_no = doc.get('document_no')
        if not doc_no:
            continue
        grouped.setdefault(str(doc_no).strip(), []).append(doc)
    return grouped


def _get_all_dates(docs):
    """دریافت همه تاریخ‌های موجود در مدارک"""
    dates = set()
    for doc in docs:
        d = doc.get('date')
        if d is not None:
            if isinstance(d, pd.Timestamp):
                d = d.date()
            elif isinstance(d, datetime):
                d = d.date()
            dates.add(d)
    return sorted(dates)


# ==================== ROUTE: /api/changed-docs ====================

def _build_changed_docs_result(data, period, discipline, project, doc_type, only_changed):
    """
    با استفاده از داده‌ی از قبل آماده‌شده (خروجی _prepare_full_dataset)،
    جدول «مدارک تغییریافته» رو می‌سازه. هیچ کار سنگینی (لود/پردازش فایل)
    اینجا انجام نمی‌شه، فقط محاسبه روی داده‌ی آماده.
    """
    days_map = {'day': 1, 'week': 7, 'month': 30}
    days = days_map.get(period, 7)

    integrated = integrate_data(data)
    all_docs = integrated['master_docs'] + integrated['vendor_docs']

    if discipline:
        discipline_lower = discipline.strip().lower()
        all_docs = [d for d in all_docs if get_doc_discipline(d).lower() == discipline_lower]

    if not all_docs:
        return {
            'docs': [],
            'count': 0,
            'period': period,
            'target_date': None,
            'latest_date': None,
            'discipline': discipline or None,
            'projects': [],
            'message': 'هیچ مدرکی با این مشخصات یافت نشد'
        }

    all_dates = _get_all_dates(all_docs)
    if not all_dates:
        return {
            'docs': [],
            'count': 0,
            'period': period,
            'target_date': None,
            'latest_date': None,
            'discipline': discipline or None,
            'projects': [],
            'message': 'هیچ تاریخی در داده‌ها یافت نشد'
        }

    latest_date = max(all_dates)
    target_date = latest_date - timedelta(days=days)

    available_dates = [d for d in all_dates if d <= target_date]
    if available_dates:
        closest_date = max(available_dates)
    else:
        closest_date = min(all_dates)

    grouped = _group_docs_by_number(all_docs)

    changed_docs = []
    total_docs = 0
    total_changed = 0

    for doc_no, versions in grouped.items():
        if not versions:
            continue
        total_docs += 1

        current_doc = _get_latest_version_on_or_before(versions, latest_date)
        if not current_doc or _is_deleted(current_doc) or _is_hold(current_doc):
            continue

        old_doc = _get_latest_version_on_or_before(versions, closest_date)

        current_progress = get_adjusted_progress(current_doc)
        old_progress = get_adjusted_progress(old_doc) if old_doc else 0.0
        change = round(current_progress - old_progress, 2)

        if only_changed and change == 0:
            continue

        if project and project != 'همه':
            proj = current_doc.get('project', '')
            if proj and str(proj).strip().lower() != project.lower():
                continue

        if doc_type and doc_type != 'همه':
            mapping = {
                'internal': ['MASTER'],
                'vendor': ['VENDOR'],
                'master': ['MASTER'],
                'engineering': ['MASTER']
            }
            expected_types = mapping.get(doc_type.lower(), [doc_type.upper()])
            dt = current_doc.get('doc_type', '').upper()
            if dt not in expected_types:
                continue

        total_changed += 1

        changed_docs.append({
            'document_no': str(doc_no),
            'document_title': current_doc.get('document_title', ''),
            'discipline': get_doc_discipline(current_doc),
            'project': current_doc.get('project', ''),
            'doc_type': current_doc.get('doc_type', ''),
            'status': current_doc.get('status', ''),
            'responsible': current_doc.get('responsible', ''),
            'old_progress': old_progress,
            'new_progress': current_progress,
            'change': change,
            'is_adjusted': (current_progress != current_doc.get('doc_progress', 0)),
            'effective_date': latest_date.strftime('%Y-%m-%d'),
            'old_date': closest_date.strftime('%Y-%m-%d')
        })

    all_projects = sorted(set(
        d.get('project', '') for d in all_docs
        if d.get('project') and str(d.get('project')).strip()
    ))

    changed_docs.sort(key=lambda x: abs(x['change']), reverse=True)

    average_change = round(
        sum(d['change'] for d in changed_docs) / len(changed_docs), 2
    ) if changed_docs else 0

    return {
        'docs': changed_docs,
        'count': len(changed_docs),
        'total_docs': total_docs,
        'total_changed': total_changed,
        'average_change': average_change,
        'period': period,
        'target_date': closest_date.strftime('%Y-%m-%d'),
        'latest_date': latest_date.strftime('%Y-%m-%d'),
        'discipline': discipline or None,
        'project': project or None,
        'doc_type': doc_type,
        'projects': all_projects
    }


@app.route('/api/changed-docs', methods=['GET'])
def get_changed_docs():
    try:
        # ===== دریافت پارامترها =====
        period = request.args.get('period', 'week').strip().lower()
        discipline = request.args.get('discipline', '').strip()
        project = request.args.get('project', '').strip()
        doc_type = request.args.get('doc_type', 'همه').strip()
        only_changed = request.args.get('only_changed', 'true').strip().lower() in ('1', 'true', 'yes', 'on')

        # ---- کش: اگر قبلاً برای همین ترکیب فیلترها محاسبه شده، همون رو برگردون ----
        data_version = data_store.get_data_version()
        cache_key = f"changed:{data_version}:{period}:{discipline}:{project}:{doc_type}:{only_changed}"
        cached_item = cache_get(cache_key)
        if cached_item is not None:
            return jsonify(cached_item)

        data, _ = get_processed_snapshot()
        result = _build_changed_docs_result(data, period, discipline, project, doc_type, only_changed)
        cache_set(cache_key, result)
        return jsonify(result)

    except Exception as e:
        safe_log(f"❌ خطا در /api/changed-docs: {e}")
        safe_log(traceback.format_exc())
        return jsonify({'error': str(e), 'docs': []}), 500

@app.route('/api/private-chat/users')
@limiter.limit("600 per minute", override_defaults=True)
@update_activity
def get_online_users():
    try:
        current_user = session.get('username')
        if not current_user:
            return jsonify({'error': 'لطفاً ابتدا وارد شوید'}), 401
        
        now = datetime.now()
        users_with_status = []
        
        for username in USERS_EMAILS.keys():
            if username == current_user:
                continue
            
            last_activity = get_user_activity(username)
            is_online = False
            last_activity_str = None
            if last_activity:
                is_online = (now - last_activity).total_seconds() < ONLINE_TIMEOUT_SECONDS
                last_activity_str = last_activity.isoformat(timespec='seconds')
            
            users_with_status.append({
                'username': username,
                'is_online': is_online,
                'last_activity': last_activity_str  # ✅ اضافه شد
            })
        
        users_with_status.sort(key=lambda x: (not x['is_online'], x['username']))
        
        return jsonify({
            'users': [u['username'] for u in users_with_status],
            'status': {u['username']: u['is_online'] for u in users_with_status},
            'last_activity': {u['username']: u['last_activity'] for u in users_with_status}  # ✅ اضافه شد
        })
        
    except Exception as e:
        safe_log(f"❌ خطا در دریافت لیست کاربران: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/private-chat/messages')
@limiter.limit("600 per minute", override_defaults=True)
@update_activity
def get_private_messages():
    try:
        # ===== دریافت کاربر واقعی از Session =====
        current_user = session.get('username')
        if not current_user:
            return jsonify({'error': 'لطفاً ابتدا وارد شوید'}), 401
        
        # ===== بررسی اینکه کاربر در سیستم ثبت شده است =====
        if current_user not in USERS_EMAILS:
            return jsonify({'error': 'کاربر در سیستم ثبت نشده است'}), 403
        
        user2 = request.args.get('user2', '')
        if not user2:
            return jsonify({'error': 'نام کاربر مقصد مشخص نیست'}), 400
        
        # ===== بررسی اینکه کاربر مقصد در سیستم ثبت شده است =====
        if user2 not in USERS_EMAILS:
            return jsonify({'error': 'کاربر مقصد در سیستم ثبت نشده است'}), 404
        
        # ===== دریافت پیام‌ها =====
        messages = load_private_messages(current_user, user2)
        return jsonify(messages)
        
    except Exception as e:
        safe_log(f"❌ خطا در دریافت پیام‌های خصوصی: {e}")
        return jsonify({'error': str(e)}), 500


# ==================== سیستم احراز هویت ====================

@app.route('/api/auth/logout', methods=['POST'])
def logout():
    session.pop('username', None)
    return jsonify({'success': True, 'message': 'خروج موفق'})

@app.route('/api/auth/current-user')
def current_user():
    username = session.get('username')
    if username:
        return jsonify({'username': username})
    return jsonify({'username': None})


@app.route('/api/private-chat/send', methods=['POST'])
@update_activity
def send_private_message():
    try:
        # ===== دریافت کاربر از Session =====
        current_user = session.get('username')
        if not current_user:
            return jsonify({'error': 'لطفاً ابتدا وارد شوید'}), 401
        
        # ===== بررسی وجود کاربر در سیستم =====
        if current_user not in USERS_EMAILS:
            return jsonify({'error': 'کاربر در سیستم ثبت نشده است'}), 403
        
        data = request.get_json()
        if not data:
            return jsonify({'error': 'داده‌ای ارسال نشده است'}), 400
        
        # ===== دریافت اطلاعات از داده (بدون from) =====
        to_user = data.get('to', '').strip()
        text = data.get('text', '').strip()
        
        if not to_user:
            return jsonify({'error': 'نام گیرنده را وارد کنید'}), 400
        
        if not text:
            return jsonify({'error': 'متن پیام نمی‌تواند خالی باشد'}), 400
        
        # ===== بررسی وجود کاربر مقصد =====
        if to_user not in USERS_EMAILS:
            return jsonify({'error': 'کاربر مقصد در سیستم ثبت نشده است'}), 404
        
        # ===== تشخیص ادمین =====
        admin_names = ['admin', 'مدیر', 'ادمین', 'ehsan', 'barazande']
        is_admin = current_user.lower() in admin_names
        
        # ===== ساخت پیام =====
        message = {
            'id': int(datetime.now().timestamp() * 1000),
            'from': current_user,      # ✅ از Session
            'to': to_user,
            'text': text,
            'time': datetime.now().isoformat(),
            'is_admin': is_admin,
            'status': 'sent'           # ✅ وضعیت اولیه
        }
        
        # ===== ذخیره پیام =====
        save_private_message(current_user, to_user, message)
        
        return jsonify({
            'success': True,
            'message': message
        })
        
    except Exception as e:
        safe_log(f"❌ خطا در ارسال پیام خصوصی: {e}")
        return jsonify({'error': str(e)}), 500

# ====== APIهای مدیریت ایمیل ======

from flask import session, redirect, url_for, make_response

@app.route('/admin/email', methods=['GET', 'POST'])
def admin_email():
    # اگر کاربر قبلاً لاگین کرده، صفحه مدیریت را نمایش بده
    if session.get('admin_logged_in'):
        resp = make_response(render_template('admin_email.html'))
        resp.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate, max-age=0'
        resp.headers['Pragma'] = 'no-cache'
        resp.headers['Expires'] = '0'
        return resp

    # درخواست POST = ارسال فرم ورود
    if request.method == 'POST':
        password = request.form.get('password', '')
        if password == 'admin123':  # رمز عبور صحیح
            session['admin_logged_in'] = True
            return redirect(url_for('admin_email'))
        else:
            return render_login_form(error='رمز عبور اشتباه است. لطفاً دوباره تلاش کنید.')

    # درخواست GET و لاگین نشده: نمایش فرم ورود
    return render_login_form()

def render_login_form(error=None):
    """نمایش فرم ورود با پیام خطای اختیاری"""
    error_html = f'<p style="color:#c62828; font-size:14px; margin-bottom:12px;">{error}</p>' if error else ''
    return f'''
    <!DOCTYPE html>
    <html>
    <head>
        <meta charset="UTF-8">
        <title>ورود به مدیریت ایمیل</title>
        <style>
            body {{ font-family: 'Segoe UI', sans-serif; background: #f0f2f5; display: flex; justify-content: center; align-items: center; height: 100vh; margin: 0; }}
            .login-box {{ background: white; padding: 40px; border-radius: 16px; box-shadow: 0 4px 20px rgba(0,0,0,0.1); text-align: center; width: 350px; }}
            .login-box h2 {{ color: #1a237e; margin-bottom: 20px; }}
            .login-box input {{ width: 100%; padding: 12px; border: 1px solid #ddd; border-radius: 8px; font-size: 16px; margin-bottom: 16px; }}
            .login-box button {{ width: 100%; padding: 12px; background: #1a237e; color: white; border: none; border-radius: 8px; font-size: 16px; cursor: pointer; }}
            .login-box button:hover {{ background: #0d47a1; }}
            .login-box .error {{ color: #c62828; font-size: 14px; margin-bottom: 12px; }}
        </style>
    </head>
    <body>
        <div class="login-box">
            <h2>مدیریت ایمیل</h2>
            {error_html}
            <form method="post" action="/admin/email">
                <input type="password" name="password" placeholder="رمز عبور را وارد کنید..." required autofocus>
                <button type="submit">ورود</button>
            </form>
            <p style="margin-top:12px; color:#999; font-size:12px;">تماس با ادمین برای دریافت رمز عبور</p>
        </div>
    </body>
    </html>
    '''

@app.route('/admin/logout')
def admin_logout():
    """خروج از بخش مدیریت"""
    session.pop('admin_logged_in', None)
    return redirect(url_for('admin_email'))


@app.route('/api/email/test', methods=['POST'])
def test_email():
    """ارسال ایمیل تست"""
    try:
        data = request.get_json()
        test_email = data.get('test_email', '')
        if not test_email:
            return jsonify({'error': 'ایمیل تست وارد نشده است'}), 400
        
        html_content = '''
        <h2 style="color:#1a237e;">📧 تست ارسال ایمیل</h2>
        <p>این یک ایمیل تست از <strong>داشبورد مدیریت مدارک آفام</strong> است.</p>
        <p>تنظیمات SMTP به درستی کار می‌کند.</p>
        <hr>
        <p style="color:#999; font-size:12px;">تاریخ: ''' + datetime.now().strftime('%Y-%m-%d %H:%M:%S') + '''</p>
        '''
        
        logo_path = os.path.join(BASE_DIR, 'static', 'logo.png')
        send_email([test_email], '📧 تست ایمیل - داشبورد آفام', html_content, logo_path)
        
        return jsonify({'success': True, 'message': f'ایمیل تست به {test_email} ارسال شد'})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/email/preview')
def email_preview():
    """پیش‌نمایش قالب ایمیل"""
    try:
        update_date = datetime.now().strftime('%Y-%m-%d')

        # دریافت آمار — پایپ‌لاین کامل (همون چیزی که send_update_email
        # واقعاً می‌فرسته)، نه فقط process_master/process_vendor_master،
        # وگرنه overdue وندور و بقیه‌ی placeholderها تو پیش‌نمایش خالی می‌مونن
        all_data = load_all_data()
        all_data = process_master(all_data)
        all_data = process_history(all_data)
        all_data = process_vendor_master(all_data)
        all_data = process_vendor_history(all_data)
        integrated = integrate_data(all_data)

        digest = collect_email_digest(integrated, update_date)
        html = fill_email_template(digest)
        if not html:
            return jsonify({'html': '<p style="color:#999;">قالب ایمیل یافت نشد</p>'})

        # جایگزینی لوگو (برای پیش‌نمایش داخل مرورگر)
        html = html.replace('cid:logo', '/static/logo.png')

        return jsonify({'html': html})
    except Exception as e:
        return jsonify({'html': f'<p style="color:#c62828;">خطا: {str(e)}</p>'})

@app.route('/api/email/config', methods=['GET', 'POST'])
def email_config():
    """دریافت یا ذخیره تنظیمات SMTP"""
    if request.method == 'GET':
        config = load_email_config()
        # حذف پسورد برای امنیت
        if 'sender_password' in config:
            config['sender_password'] = '******'
        return jsonify(config)
    
    try:
        data = request.get_json()
        config = load_email_config()
        config.update({
            'smtp_server': data.get('smtp_server', ''),
            'smtp_port': data.get('smtp_port', 587),
            'sender_email': data.get('sender_email', ''),
            'sender_password': data.get('sender_password', '')
        })
        save_email_config(config)
        return jsonify({'success': True, 'message': 'تنظیمات با موفقیت ذخیره شد'})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/email/list', methods=['GET', 'POST', 'DELETE'])
def email_list():
    """مدیریت لیست ایمیل‌ها"""
    if request.method == 'GET':
        return jsonify({'emails': get_email_list()})
    
    if request.method == 'POST':
        data = request.get_json()
        email = data.get('email', '').strip()
        if not email:
            return jsonify({'error': 'ایمیل معتبر نیست'}), 400
        if add_email(email):
            return jsonify({'success': True, 'message': 'ایمیل با موفقیت اضافه شد'})
        return jsonify({'error': 'ایمیل قبلاً در لیست وجود دارد'}), 400
    
    if request.method == 'DELETE':
        data = request.get_json()
        email = data.get('email', '').strip()
        if remove_email(email):
            return jsonify({'success': True, 'message': 'ایمیل با موفقیت حذف شد'})
        return jsonify({'error': 'ایمیل در لیست یافت نشد'}), 404

@app.route('/api/email/send-update', methods=['POST'])
def send_update_email():
    try:
        data = request.get_json() or {}
        update_date = data.get('update_date', datetime.now().strftime('%Y-%m-%d'))
        subject = f"آپدیت داشبورد مدیریت مدارک آفام - تاریخ {update_date}"
        
        cc_emails = data.get('cc_emails', [])
        if cc_emails and isinstance(cc_emails, str):
            cc_emails = [cc_emails]
        
        # ===== دریافت آمار =====
        # نکته: قبلاً اینجا فقط process_master/process_vendor_master صدا زده
        # می‌شد (بدون process_history/process_vendor_history)، در نتیجه
        # calculate_vendor_overdue داخل integrate_data ستون‌های لازم
        # (vendor_history_cols) را نداشت و overdue واقعی محاسبه نمی‌شد.
        # همچنین قالب ایمیل فقط با ۴ متغیر دستی پر می‌شد و بقیه‌ی
        # placeholderها (headline, hotspot_block, دکمه‌های ورود و...) به
        # همون شکل {{ ... }} خام تو ایمیل واقعی باقی می‌موند.
        all_data = load_all_data()
        all_data = process_master(all_data)
        all_data = process_history(all_data)
        all_data = process_vendor_master(all_data)
        all_data = process_vendor_history(all_data)
        integrated = integrate_data(all_data)

        digest = collect_email_digest(integrated, update_date)

        # ===== لاگ برای دیباگ =====
        safe_log(
            f"📊 آمار ایمیل: total={digest['total_docs']}, "
            f"not_issued={digest['not_issued']}, progress={digest['overall_progress']}"
        )

        # ===== ساخت HTML نهایی از روی قالب (پر کردن همه‌ی placeholderها) =====
        html_content = fill_email_template(digest)
        if not html_content:
            return jsonify({'error': 'قالب ایمیل یافت نشد'}), 404

        # ===== ارسال ایمیل =====
        emails = get_email_list()
        if not emails:
            return jsonify({'error': 'لیست ایمیل‌ها خالی است'}), 400
        
        logo_path = os.path.join(BASE_DIR, 'static', 'logo.png')
        send_email(emails, subject, html_content, logo_path, cc_emails)
        
        return jsonify({
            'success': True,
            'message': f'ایمیل آپدیت برای {len(emails)} نفر ارسال شد',
            'count': len(emails),
            'cc_count': len(cc_emails) if cc_emails else 0,
            'stats': {
                'total_docs': digest['total_docs'],
                'not_issued': digest['not_issued'],
                'overall_progress': digest['overall_progress']
            }
        })
        
    except Exception as e:
        safe_log(f"❌ خطا در ارسال ایمیل آپدیت: {e}")
        return jsonify({'error': str(e)}), 500
    
# ====== مدیریت CC ======

@app.route('/api/email/cc/list', methods=['GET'])
def get_cc_list():
    """دریافت لیست ایمیل‌های CC"""
    config = load_email_config()
    return jsonify({'emails': config.get('cc_emails', [])})

@app.route('/api/email/cc', methods=['POST', 'DELETE'])
def manage_cc():
    """مدیریت ایمیل‌های CC"""
    config = load_email_config()
    if 'cc_emails' not in config:
        config['cc_emails'] = []
    
    if request.method == 'POST':
        data = request.get_json()
        email = data.get('email', '').strip()
        if not email:
            return jsonify({'error': 'ایمیل معتبر نیست'}), 400
        if email not in config['cc_emails']:
            config['cc_emails'].append(email)
            save_email_config(config)
            return jsonify({'success': True, 'message': 'ایمیل CC با موفقیت اضافه شد'})
        return jsonify({'error': 'ایمیل CC قبلاً در لیست وجود دارد'}), 400
    
    if request.method == 'DELETE':
        data = request.get_json()
        if data.get('clear_all'):
            config['cc_emails'] = []
            save_email_config(config)
            return jsonify({'success': True, 'message': 'همه ایمیل‌های CC پاک شدند'})
        
        email = data.get('email', '').strip()
        if email in config['cc_emails']:
            config['cc_emails'].remove(email)
            save_email_config(config)
            return jsonify({'success': True, 'message': 'ایمیل CC با موفقیت حذف شد'})
        return jsonify({'error': 'ایمیل CC در لیست یافت نشد'}), 404

#---------- اضافه کردن API آپلود فایل---------    
@app.route('/api/chat/upload', methods=['POST'])
def upload_chat_file():
    try:
        if 'file' not in request.files:
            return jsonify({'error': 'فایلی انتخاب نشده'}), 400
        file = request.files['file']
        if file.filename == '':
            return jsonify({'error': 'نام فایل خالی است'}), 400
        
        # ذخیره فایل
        filename = secure_filename(file.filename)
        filename = f"{datetime.now().strftime('%Y%m%d%H%M%S')}_{file.filename}"
        filepath = os.path.join(UPLOAD_FOLDER_CHAT, filename)
        file.save(filepath)
        
        # ارسال پیام با لینک فایل
        from_user = request.form.get('from', 'کاربر')
        to_user = request.form.get('to', '')
        file_url = f"/chat_uploads/{filename}"
        
        # ذخیره پیام
        message = {
            'id': int(datetime.now().timestamp() * 1000),
            'from': from_user,
            'to': to_user,
            'text': f"📎 {file.filename}",
            'file_url': file_url,
            'file_name': file.filename,
            'time': datetime.now().isoformat(),
            'is_admin': False,
            'status': 'sent'
        }
        save_private_message(from_user, to_user, message)
        
        return jsonify({'success': True, 'message': message, 'file_url': file_url})
    except Exception as e:
        return jsonify({'error': str(e)}), 500
# -------مسیر استاتیک برای فایل‌ها تنظیم------
@app.route('/chat_uploads/<filename>')
def chat_uploaded_file(filename):
    return send_from_directory(UPLOAD_FOLDER_CHAT, filename)

@app.route('/api/email/test-template')
def test_template():
    """تست نمایش قالب با متغیرهای نمونه"""
    try:
        template_path = os.path.join(BASE_DIR, 'templates', 'email_template.html')
        with open(template_path, 'r', encoding='utf-8') as f:
            html = f.read()
        
        # جایگزینی با مقادیر نمونه
        html = html.replace('{{ update_date }}', '2026-08-01')
        html = html.replace('{{ total_docs }}', '1234')
        html = html.replace('{{ not_issued }}', '56')
        html = html.replace('{{ overall_progress }}', '78.5')
        
        return html  # نمایش مستقیم در مرورگر
    except Exception as e:
        return f"خطا: {e}"


@app.route('/api/auth/send-code', methods=['POST'])
def send_auth_code():
    try:
        data = request.get_json()
        email = data.get('email', '').strip().lower()
        
        if not email:
            return jsonify({'error': 'لطفاً ایمیل خود را وارد کنید'}), 400
        
        # بررسی اینکه ایمیل در لیست کاربران وجود دارد
        if email not in EMAIL_TO_USERNAME:
            return jsonify({'error': 'این ایمیل در سیستم ثبت نشده است'}), 404
        
        username = EMAIL_TO_USERNAME[email]
        
        # تولید کد ۶ رقمی
        code = ''.join(secrets.choice('0123456789') for _ in range(6))
        
        # ذخیره کد با زمان انقضا (۵ دقیقه)
        temp_codes[email] = {
            'code': code,
            'expires': datetime.now() + timedelta(minutes=5),
            'username': username,
            'attempts': 0
        }
        
        # ===== ارسال ایمیل حاوی کد =====
        subject = f"کد ورود به داشبورد آفام - {username}"
        html_content = f"""
        <html>
        <body style="font-family: Arial, sans-serif; direction: rtl; text-align: center; padding: 20px;">
            <h2 style="color: #1a237e;">کد ورود شما</h2>
            <p style="font-size: 16px; color: #333;">
                {username} عزیز،
            </p>
            <p style="font-size: 16px; color: #333;">
                کد ورود یکبارمصرف شما به داشبورد مدیریت مدارک آفام:
            </p>
            <div style="background: #e3f2fd; padding: 20px; border-radius: 12px; display: inline-block; margin: 20px auto;">
                <span style="font-size: 32px; font-weight: bold; color: #1a237e; letter-spacing: 10px;">
                    {code}
                </span>
            </div>
            <p style="font-size: 14px; color: #999;">
                این کد به مدت ۵ دقیقه معتبر است.
            </p>
            <p style="font-size: 14px; color: #999;">
                اگر درخواستی برای ورود نداده‌اید، این پیام را نادیده بگیرید.
            </p>
            <hr style="border: 1px solid #eee; margin: 20px 0;">
            <p style="font-size: 12px; color: #bbb;">
                داشبورد مدیریت مدارک آفام - سیستم یکپارچه مدیریت مدارک
            </p>
        </body>
        </html>
        """
        
        # استفاده از تابع send_email موجود
        send_email([email], subject, html_content)
        
        safe_log(f"📧 کد تأیید برای {email} ارسال شد: {code}")
        
        return jsonify({
            'success': True,
            'message': f'کد تأیید به ایمیل {email} ارسال شد',
            'email': email
        })
        
    except Exception as e:
        safe_log(f"❌ خطا در ارسال کد تأیید: {e}")
        return jsonify({'error': str(e)}), 500


@app.route('/api/auth/verify-code', methods=['POST'])
def verify_auth_code():
    try:
        data = request.get_json()
        email = data.get('email', '').strip().lower()
        code = data.get('code', '').strip()
        
        if not email or not code:
            return jsonify({'error': 'ایمیل و کد را وارد کنید'}), 400
        
        # بررسی وجود کد برای این ایمیل
        if email not in temp_codes:
            return jsonify({'error': 'کد تأیید یافت نشد. لطفاً دوباره درخواست کنید.'}), 404
        
        code_data = temp_codes[email]
        
        # بررسی انقضای کد
        if datetime.now() > code_data['expires']:
            # حذف کد منقضی‌شده
            del temp_codes[email]
            return jsonify({'error': 'کد تأیید منقضی شده است. لطفاً دوباره درخواست کنید.'}), 401
        
        # بررسی تعداد تلاش‌ها (حداکثر ۳ بار)
        if code_data['attempts'] >= 3:
            del temp_codes[email]
            return jsonify({'error': 'تعداد تلاش‌های ناموفق بیش از حد مجاز است. لطفاً دوباره درخواست کنید.'}), 401
        
        # بررسی صحت کد
        if code_data['code'] != code:
            temp_codes[email]['attempts'] += 1
            remaining = 3 - temp_codes[email]['attempts']
            return jsonify({'error': f'کد اشتباه است. {remaining} تلاش دیگر باقی مانده.'}), 401
        
        # ===== ورود موفق =====
        username = code_data['username']
        session['username'] = username
        session.permanent = True

        # حذف کد استفاده‌شده
        del temp_codes[email]
        
        safe_log(f"✅ کاربر {username} با موفقیت وارد شد (ایمیل: {email})")
        
        return jsonify({
            'success': True,
            'username': username,
            'message': f'خوش آمدید {username}'
        })
        
    except Exception as e:
        safe_log(f"❌ خطا در تأیید کد: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/private-chat/update-status', methods=['POST'])
@limiter.limit("600 per minute", override_defaults=True)
def update_message_status():
    """به‌روزرسانی وضعیت پیام‌ها (delivered/seen)"""
    try:
        current_user = session.get('username')
        if not current_user:
            return jsonify({'error': 'لطفاً وارد شوید'}), 401
        
        data = request.get_json()
        other_user = data.get('other_user', '').strip()
        message_ids = data.get('message_ids', [])  # لیست ID پیام‌ها
        status = data.get('status', 'delivered')  # delivered یا seen
        
        if not other_user or not message_ids:
            return jsonify({'error': 'اطلاعات ناقص است'}), 400
        
        # بارگذاری پیام‌ها
        messages = load_private_messages(current_user, other_user)
        updated = 0
        
        for msg in messages:
            if msg['id'] in message_ids:
                # فقط پیام‌هایی که از طرف مقابل ارسال شده‌اند را به‌روز کن
                if msg['from'] == other_user and msg['to'] == current_user:
                    if msg.get('status') != status:
                        msg['status'] = status
                        updated += 1
        
        # ذخیره مجدد
        if updated > 0:
            filepath = get_private_chat_file(current_user, other_user)
            with open(filepath, 'w', encoding='utf-8') as f:
                json.dump(messages, f, ensure_ascii=False, indent=2)
        
        return jsonify({'success': True, 'updated': updated})
        
    except Exception as e:
        safe_log(f"❌ خطا در به‌روزرسانی وضعیت پیام: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/chat/broadcast', methods=['POST'])
def broadcast_message():
    """ارسال پیام همگانی به همه کاربران (فقط ادمین)"""
    try:
        # ===== بررسی ادمین =====
        current_user = session.get('username')
        if not current_user:
            return jsonify({'error': 'لطفاً وارد شوید'}), 401
        
        admin_names = ['admin', 'مدیر', 'ادمین', 'ehsan', 'barazande']
        if current_user.lower() not in admin_names:
            return jsonify({'error': 'تنها ادمین می‌تواند پیام همگانی ارسال کند'}), 403
        
        data = request.get_json()
        if not data:
            return jsonify({'error': 'داده‌ای ارسال نشده است'}), 400
        
        message_text = data.get('message', '').strip()
        if not message_text:
            return jsonify({'error': 'متن پیام را وارد کنید'}), 400
        
        # ===== ارسال پیام به همه کاربران =====
        all_users = list(USERS_EMAILS.keys())
        sent_count = 0
        
        for username in all_users:
            if username == current_user:
                continue  # خودش را نادیده بگیر
            
            # ساخت پیام
            message = {
                'id': int(datetime.now().timestamp() * 1000) + sent_count,
                'from': 'سیستم',
                'to': username,
                'text': f"📢 {message_text}",
                'time': datetime.now().isoformat(),
                'is_admin': True,
                'is_system': True,
                'status': 'delivered'  # مستقیماً تحویل داده شده
            }
            
            # ذخیره در چت خصوصی
            save_private_message('سیستم', username, message)
            sent_count += 1
        
        safe_log(f"📢 ادمین {current_user} پیام همگانی به {sent_count} نفر ارسال کرد")
        
        return jsonify({
            'success': True,
            'message': f'پیام همگانی به {sent_count} کاربر ارسال شد',
            'count': sent_count
        })
        
    except Exception as e:
        safe_log(f"❌ خطا در ارسال پیام همگانی: {e}")
        return jsonify({'error': str(e)}), 500


@app.route('/api/debug-distribute-names')
def debug_distribute_names():
    """نمایش تمام نام‌های موجود در رکوردهای Distribute برای دیباگ"""
    data = load_all_data()
    data = process_history(data)
    data = process_vendor_history(data)
    
    history_df = data.get('history')
    history_cols = data.get('history_cols', {})
    
    names = set()
    if history_df is not None:
        to_name_col = history_cols.get('to_name')
        log_status_col = history_cols.get('log_status')
        if to_name_col and log_status_col:
            distribute_df = history_df[history_df[log_status_col].astype(str).str.strip() == 'Distribute']
            names.update(distribute_df[to_name_col].dropna().unique())
    
    vendor_history_df = data.get('vendor_history')
    vendor_history_cols = data.get('vendor_history_cols', {})
    if vendor_history_df is not None:
        to_name_col = vendor_history_cols.get('to_name')
        log_status_col = vendor_history_cols.get('log_status')
        if to_name_col and log_status_col:
            distribute_df = vendor_history_df[vendor_history_df[log_status_col].astype(str).str.strip() == 'Distribute']
            names.update(distribute_df[to_name_col].dropna().unique())
    
    return jsonify({
        'names': sorted(list(names)),
        'count': len(names)
    })

@app.route('/api/debug-doc-history/<doc_no>')
def debug_doc_history(doc_no):
    """نمایش تاریخچه یک مدرک خاص برای دیباگ"""
    try:
        data = load_all_data()
        data = process_history(data)
        df = data.get('history')
        cols = data.get('history_cols', {})
        doc_no_col = cols.get('doc_no')
        
        if df is None or doc_no_col is None:
            return jsonify({'error': 'داده‌های history یافت نشد'}), 404
        
        # فیلتر بر اساس شماره مدرک
        doc_history = df[df[doc_no_col].astype(str).str.strip() == doc_no]
        
        if doc_history.empty:
            return jsonify({'error': f'مدرک {doc_no} در history یافت نشد'}), 404
        
        # تبدیل تاریخ‌ها به رشته
        records = []
        for _, row in doc_history.iterrows():
            record = {}
            for col in doc_history.columns:
                val = row[col]
                if isinstance(val, (pd.Timestamp, datetime)):
                    if pd.isna(val):
                        record[col] = None
                    else:
                        record[col] = val.strftime('%Y-%m-%d %H:%M:%S')
                elif isinstance(val, (pd.Timedelta,)):
                    record[col] = str(val) if not pd.isna(val) else None
                else:
                    record[col] = val if not pd.isna(val) else None
            records.append(record)
        
        return jsonify({
            'doc_no': doc_no,
            'count': len(records),
            'records': records,
            'columns': list(doc_history.columns)
        })
        
    except Exception as e:
        safe_log(f"❌ خطا در دیباگ تاریخچه: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/private-chat/unread-counts')
@limiter.limit("600 per minute", override_defaults=True)
@update_activity
def get_unread_counts():
    """
    دریافت تعداد پیام‌های نخوانده برای کاربر جاری از همه کاربران دیگر
    """
    try:
        current_user = session.get('username')
        if not current_user:
            return jsonify({'error': 'لطفاً وارد شوید'}), 401
        
        unread_counts = {}
        total_unread = 0
        
        # بررسی همه فایل‌های چت در پوشه private_chats
        if not os.path.exists(PRIVATE_CHAT_DIR):
            return jsonify({'unread_counts': {}, 'total': 0})
        
        for filename in os.listdir(PRIVATE_CHAT_DIR):
            if not filename.endswith('.json'):
                continue
            
            # استخراج نام‌های دو طرف از نام فایل
            # فرمت: chat_user1_user2.json
            parts = filename.replace('.json', '').split('_')
            if len(parts) < 3:
                continue
            
            user1 = parts[1]
            user2 = parts[2]
            
            # اگر کاربر جاری در این مکالمه نیست، رد کن
            if current_user not in [user1, user2]:
                continue
            
            # تعیین طرف مقابل
            other_user = user2 if user1 == current_user else user1
            
            # بارگذاری پیام‌ها
            filepath = os.path.join(PRIVATE_CHAT_DIR, filename)
            try:
                with open(filepath, 'r', encoding='utf-8') as f:
                    messages = json.load(f)
            except:
                continue
            
            # شمارش پیام‌های نخوانده از طرف مقابل
            count = 0
            for msg in messages:
                # پیام از طرف مقابل است و به کاربر جاری ارسال شده و دیده نشده است
                if (msg.get('from') == other_user and 
                    msg.get('to') == current_user and 
                    msg.get('status') != 'seen'):
                    count += 1
            
            if count > 0:
                unread_counts[other_user] = count
                total_unread += count
        
        return jsonify({
            'unread_counts': unread_counts,
            'total': total_unread
        })
        
    except Exception as e:
        safe_log(f"❌ خطا در دریافت تعداد پیام‌های نخوانده: {e}")
        return jsonify({'error': str(e)}), 500
#----------------debug gapgpt---------------------------

@app.route('/api/debug-normalized-names')
def debug_normalized_names():
    result = get_distribute_docs()
    names = list(result.keys())
    return jsonify({
        'normalized_names': names,
        'count': len(names),
        'email_mapping': {name: get_email_for_person(name) for name in names}
    })

@app.route('/debug_discipline_changes')
def debug_discipline_changes():
    import pandas as pd
    from datetime import datetime, timedelta
    from flask import request, jsonify

    discipline = request.args.get('discipline', '').strip()
    period = request.args.get('period', 'week').strip().lower()  # week|month
    doc_type_filter = request.args.get('doc_type', '').strip().upper()  # VENDOR|MASTER|''
    only_changed = request.args.get('only_changed', 'false').strip().lower() in ('1', 'true', 'yes')

    if not discipline:
        return jsonify({'error': 'discipline parameter is required'}), 400

    if doc_type_filter and doc_type_filter not in ('VENDOR', 'MASTER'):
        return jsonify({'error': 'doc_type must be VENDOR or MASTER'}), 400

    days = 7 if period == 'week' else 30
    today = datetime.now().date()
    target_date = today - timedelta(days=days)

    # -------- helpers --------
    def _norm_col_name(c):
        return str(c).strip().lower().replace('.', '').replace('_', '').replace(' ', '')

    def normalize_df(df, default_doc_type):
        if df is None or len(df) == 0:
            return pd.DataFrame(columns=[
                'document_no', 'document_title', 'discipline', 'status', 'progress',
                'weight', 'eng_weight', 'date', 'doc_type', 'deleted', 'hold',
                'responsible', 'comment'
            ])

        x = df.copy()
        rename_map = {}

        for col in x.columns:
            n = _norm_col_name(col)

            if n in ('documentno', 'docno', 'mdrno'):
                rename_map[col] = 'document_no'
            elif n in ('documenttitle', 'title'):
                rename_map[col] = 'document_title'
            elif n == 'discipline':
                rename_map[col] = 'discipline'
            elif n in ('status', 'logstatus'):
                rename_map[col] = 'status'
            elif n in ('documentprogress', 'progress', 'docprogress'):
                rename_map[col] = 'progress'
            elif n in ('weight',):
                rename_map[col] = 'weight'
            elif n in ('engweight', 'engineeringweight'):
                rename_map[col] = 'eng_weight'
            elif n in ('date', 'reportdate'):
                rename_map[col] = 'date'
            elif n == 'deleted':
                rename_map[col] = 'deleted'
            elif n == 'hold':
                rename_map[col] = 'hold'
            elif n == 'responsible':
                rename_map[col] = 'responsible'
            elif n == 'comment':
                rename_map[col] = 'comment'
            elif n == 'doctype':
                rename_map[col] = 'doc_type'

        x = x.rename(columns=rename_map)

        # doc_type
        if 'doc_type' not in x.columns:
            x['doc_type'] = default_doc_type
        x['doc_type'] = x['doc_type'].astype(str).str.upper().replace('NAN', default_doc_type)

        # aliases for compatibility with get_adjusted_progress
        if 'progress' in x.columns:
            x['Document Progress'] = x['progress']
            x['document_progress'] = x['progress']

        # type safety
        if 'date' in x.columns:
            x['date'] = pd.to_datetime(x['date'], errors='coerce')

        for c in ['progress', 'weight', 'eng_weight']:
            if c in x.columns:
                x[c] = pd.to_numeric(x[c], errors='coerce').fillna(0.0)
            else:
                x[c] = 0.0

        for c in ['document_no', 'document_title', 'discipline', 'status', 'deleted', 'hold', 'responsible', 'comment']:
            if c not in x.columns:
                x[c] = ''
            x[c] = x[c].astype(str)

        # clean NaN string
        x = x.replace({'nan': '', 'NaN': '', 'None': ''})

        return x

    def get_adjusted_progress_safe(row):
        # row: pandas Series
        d = row.to_dict()

        # 1) try original business logic
        p = None
        try:
            p = get_adjusted_progress(d)
            p = float(p) if pd.notna(p) else None
        except Exception:
            p = None

        # 2) fallback raw progress
        raw = None
        for k in ('progress', 'Document Progress', 'document_progress'):
            v = d.get(k, None)
            if v is not None and pd.notna(v):
                try:
                    raw = float(v)
                    break
                except Exception:
                    pass

        if p is None:
            return raw if raw is not None else 0.0

        # اگر منطق اصلی 0 داد ولی raw معتبر داشتیم، raw را بگیر
        if p == 0.0 and raw is not None:
            return raw

        return p

    def latest_per_doc(df):
        if df.empty:
            return df
        z = df[df['document_no'].astype(str).str.strip() != ''].copy()
        z = z.sort_values('date', ascending=True, na_position='last')
        z = z.groupby(['document_no', 'doc_type'], as_index=False).tail(1)
        return z

    # -------- load data --------
    try:
        data_store.initialize(app.config.get('UPLOAD_FILES', {}))
    except Exception:
        pass

    master_df = normalize_df(data_store.get_dataframe('master'), 'MASTER')
    vendor_df = normalize_df(data_store.get_dataframe('vendor_master'), 'VENDOR')
    history_df = normalize_df(data_store.get_dataframe('history'), 'MASTER')
    vendor_history_df = normalize_df(data_store.get_dataframe('vendor_history'), 'VENDOR')

    # -------- select by doc_type --------
    if doc_type_filter == 'VENDOR':
        current_pool = vendor_df
        history_pool = vendor_history_df
    elif doc_type_filter == 'MASTER':
        current_pool = master_df
        history_pool = history_df
    else:
        current_pool = pd.concat([master_df, vendor_df], ignore_index=True)
        history_pool = pd.concat([history_df, vendor_history_df], ignore_index=True)

    # discipline filter
    disc_norm = discipline.strip().lower()
    current_pool = current_pool[current_pool['discipline'].str.strip().str.lower() == disc_norm].copy()
    history_pool = history_pool[history_pool['discipline'].str.strip().str.lower() == disc_norm].copy()

    # current snapshot: latest per doc from current pool
    current_latest = latest_per_doc(current_pool)

    # old snapshot candidate = history + current (تا target_date)
    old_hist = history_pool.copy()
    old_cur = current_pool.copy()

    old_hist = old_hist[old_hist['date'].notna() & (old_hist['date'].dt.date <= target_date)]
    old_cur = old_cur[old_cur['date'].notna() & (old_cur['date'].dt.date <= target_date)]

    old_union = pd.concat([old_hist, old_cur], ignore_index=True)

    # اگر تاریخ در current وجود ندارد/NaT است، fallback: از history تنها استفاده کن
    if old_union.empty:
        old_union = old_hist.copy()

    old_latest = latest_per_doc(old_union)


    # build keyed tables
    cur = current_latest.copy()
    old = old_latest.copy()

    # adjusted progress
    if not cur.empty:
        cur['current_progress'] = cur.apply(get_adjusted_progress_safe, axis=1)
    else:
        cur['current_progress'] = []

    if not old.empty:
        old['old_progress'] = old.apply(get_adjusted_progress_safe, axis=1)
    else:
        old['old_progress'] = []

    cur_cols = ['document_no', 'doc_type', 'document_title', 'discipline', 'status',
                'weight', 'eng_weight', 'current_progress']
    old_cols = ['document_no', 'doc_type', 'old_progress']

    for c in cur_cols:
        if c not in cur.columns:
            cur[c] = 0.0 if c in ('weight', 'eng_weight', 'current_progress') else ''
    for c in old_cols:
        if c not in old.columns:
            old[c] = 0.0 if c == 'old_progress' else ''

    cur = cur[cur_cols].copy()
    old = old[old_cols].copy()

    merged = cur.merge(old, on=['document_no', 'doc_type'], how='outer', indicator=True)

    # fill defaults
    for c in ['current_progress', 'old_progress', 'weight', 'eng_weight']:
        if c in merged.columns:
            merged[c] = pd.to_numeric(merged[c], errors='coerce').fillna(0.0)

    for c in ['document_title', 'discipline', 'status', 'document_no', 'doc_type']:
        if c in merged.columns:
            merged[c] = merged[c].fillna('').astype(str)

    # row status
    def _row_kind(ind):
        if ind == 'left_only':
            return 'NEW'
        if ind == 'right_only':
            return 'REMOVED'
        return 'UNCHANGED_OR_UPDATED'

    merged['row_kind'] = merged['_merge'].apply(_row_kind)

    # progress for removed docs
    # removed: current=0, old=old_progress
    merged.loc[merged['_merge'] == 'right_only', 'current_progress'] = 0.0

    # change
    merged['change'] = (merged['current_progress'] - merged['old_progress']).round(2)

    # final state label
    def _state(r):
        if r['_merge'] == 'left_only':
            return 'NEW'
        if r['_merge'] == 'right_only':
            return 'REMOVED'
        return 'UPDATED' if abs(r['change']) > 0 else 'UNCHANGED'

    merged['state'] = merged.apply(_state, axis=1)

    # weight selection
    def _pick_weight(r):
        dt = str(r.get('doc_type', '')).upper()
        if dt == 'MASTER':
            return float(r.get('eng_weight', 0.0) or 0.0)
        # VENDOR: اگر weight ندارید، فعلاً 0
        return float(r.get('weight', 0.0) or 0.0)

    merged['weight_used'] = merged.apply(_pick_weight, axis=1)
    merged['weighted_delta'] = (merged['weight_used'] * merged['change'] / 100.0).round(6)
    merged['weighted_delta'] = merged['weighted_delta'].replace([pd.NA], 0.0).fillna(0.0)

    # only_changed option
    if only_changed:
        merged = merged[merged['change'].abs() > 0].copy()
    # sort
    merged = merged.sort_values(by=['state', 'document_no'], ascending=[True, True])

    documents = []
    for _, r in merged.iterrows():
        documents.append({
            'document_no': r['document_no'],
            'document_title': r['document_title'] or '-',
            'discipline': r['discipline'] or discipline,
            'doc_type': r['doc_type'],
            'status': r['status'] or '',
            'state': r['state'],  # NEW / REMOVED / UPDATED / UNCHANGED
            'weight': float(r['weight_used']),
            'old_progress': float(round(r['old_progress'], 2)),
            'current_progress': float(round(r['current_progress'], 2)),
            'change': float(round(r['change'], 2)),
            'weighted_delta': float(round(r['weighted_delta'], 6)),
        })

    total_weighted_delta = round(
        sum(float(d.get('weighted_delta', 0.0) or 0.0) for d in documents),
        6
    )

    return jsonify({
        'discipline': discipline,
        'period': period,
        'doc_type': doc_type_filter or 'ALL',
        'only_changed': only_changed,
        'target_date': target_date.strftime('%Y-%m-%d'),
        'count': len(documents),
        'total_weighted_delta': total_weighted_delta,
        'documents': documents
    })

@app.route('/debug_disciplines')
def debug_disciplines():
    all_docs = load_all_docs_from_store()
    active_latest_docs = get_active_latest_docs_from_store()

    discipline_counts = {}

    for doc in active_latest_docs:
        disc = get_doc_discipline(doc) or 'نامشخص'
        if not disc:
            continue

        if disc not in discipline_counts:
            discipline_counts[disc] = 0
        discipline_counts[disc] += 1

    sorted_disciplines = sorted(discipline_counts.keys())

    return jsonify({
        'all_docs_count': len(all_docs),
        'active_latest_docs_count': len(active_latest_docs),
        'discipline_count': len(sorted_disciplines),
        'sample_disciplines': sorted_disciplines[:20],
        'discipline_counts': [
            {'discipline': disc, 'count': discipline_counts[disc]}
            for disc in sorted_disciplines[:50]
        ]
    })

# ==================== تحلیل هوشمند پروژه (Qwen) ====================
def _normalize_intel_type(doc_type):
    value = (doc_type or 'اصلی').strip()
    if value in ('وندور', 'vendor', 'VENDOR'):
        return 'وندور'
    return 'اصلی'


def _normalize_intel_period(period):
    value = (period or 'today').strip().lower()
    if value in ('week', 'weekly', 'هفته'):
        return 'week'
    if value in ('month', 'monthly', 'ماه'):
        return 'month'
    return 'today'


def _top_performers_for_period(period):
    days = 7 if period in ('today', 'week') else 30
    date_to = date.today()
    date_from = date_to - timedelta(days=days)
    try:
        result = _build_leaderboard_result('همه', date_from, date_to)
    except Exception as e:
        safe_log(f"[intelligence] خطا در نفرات برتر: {e}", level="warning")
        return [], [], []
    scores = sorted(result.get('scores') or [], key=lambda s: s.get('total_score') or 0, reverse=True)
    compact = []
    import org_structure
    for s in scores[:8]:
        disc = s.get('discipline')
        if org_structure.is_excluded_discipline(disc) and s.get('role') == 'manager':
            disc = 'مدیر مهندسی'
        elif org_structure.is_excluded_discipline(disc):
            continue
        compact.append({
            'person': s.get('person'),
            'role': s.get('role'),
            'discipline': disc,
            'total_score': int(round(float(s.get('total_score') or 0))),
            'core_score': int(round(float(s.get('core_score') or 0))),
        })
        if len(compact) >= 5:
            break
    disc_compact = [{
        'discipline': d.get('discipline'),
        'total_score': d.get('total_score'),
        'member_count': d.get('member_count'),
    } for d in (result.get('disciplines') or [])[:8]]
    return compact, disc_compact, result.get('scores') or []


def _apply_standing_instructions():
    """
    یادداشت‌های دستهٔ 'instruction' رو به project_intelligence می‌ده تا هر
    تماس با Qwen (خلاصه‌ی روزانه، پاسخ فعالیت، پاسخ دانش دستی) خودکار
    رعایتشون کنه. همه‌ی یادداشت‌ها (including دستهٔ 'fact') رو هم برمی‌گردونه
    تا برای جست‌وجوی سؤال استفاده بشه.
    """
    import project_intelligence
    entries = _get_all_knowledge_entries()
    instructions = [e['text'] for e in entries if e.get('category') == 'instruction']
    project_intelligence.set_standing_instructions(instructions)
    return entries


def _get_intelligence_bundle(force=False, doc_type='اصلی', period='today'):
    import project_intelligence

    doc_type = _normalize_intel_type(doc_type)
    period = _normalize_intel_period(period)
    data_version = data_store.get_data_version()
    scoring_version = _get_scoring_config_version()
    today_key = f"intel_v7:{data_version}:{scoring_version}:{doc_type}:{period}"
    facts_key = f"intel_v7:{data_version}:{scoring_version}:{doc_type}:{period}:facts"
    if not force:
        cached = cache_get(today_key)
        facts = cache_get(facts_key)
        if cached is not None and facts is not None:
            return cached, facts

    _, integrated = get_processed_snapshot()
    top_performers, discipline_scores, period_scores = _top_performers_for_period(period)
    people_directory = []
    try:
        people_directory = _build_people_directory(period_scores)
    except Exception as e:
        safe_log(f"[intelligence] خطا در نمایه افراد: {e}", level="warning")
        people_directory = []
    facts = project_intelligence.build_facts(
        integrated,
        doc_type=doc_type,
        extras={
            'top_performers': top_performers,
            'discipline_scores': discipline_scores,
            'people_directory': people_directory,
            'period': period,
        },
    )
    briefing = project_intelligence.generate_daily_briefing(facts, period=period)
    cache_set(facts_key, facts)
    cache_set(today_key, briefing)
    return briefing, facts


@app.route('/api/intelligence/today')
@update_activity
def intelligence_today():
    try:
        doc_type = request.args.get('type') or request.args.get('doc_type') or 'اصلی'
        period = request.args.get('period', 'today')
        _apply_standing_instructions()
        briefing, _ = _get_intelligence_bundle(doc_type=doc_type, period=period)
        return jsonify(briefing)
    except Exception as e:
        safe_log(f"[intelligence] خطا در خلاصه روزانه: {e}", level="error")
        safe_log(traceback.format_exc(), level="error")
        return jsonify({'ready': False, 'error': str(e)}), 500


def _infer_period_from_question(question: str, fallback_period: str) -> str:
    """
    اگه خودِ متن سؤال یه بازه‌ی زمانی مشخص کرده باشه («ماه گذشته»، «این
    هفته»...)، همون رو به‌جای period پیش‌فرض UI استفاده می‌کنیم — وگرنه
    کسی که «امتیاز فلانی ماه گذشته» می‌پرسه، بی‌سروصدا جواب هفتگی می‌گیره.
    """
    import person_activity
    q = person_activity._fold(question or '')
    if any(k in q for k in ('ماهگذشته', 'ماهپیش', 'ماهاخیر', 'اینماه', 'یکماه')):
        return 'month'
    if any(k in q for k in ('هفتهگذشته', 'هفتهپیش', 'هفتهاخیر', 'اینهفته')):
        return 'week'
    if 'امروز' in q:
        return 'today'
    return fallback_period


@app.route('/api/intelligence/ask', methods=['POST'])
@update_activity
def intelligence_ask():
    try:
        import project_intelligence
        import document_knowledge

        payload = request.get_json(silent=True) or {}
        question = (payload.get('question') or '').strip()
        project = (payload.get('project') or 'همه').strip() or 'همه'
        doc_type = payload.get('type') or payload.get('doc_type') or 'اصلی'
        period = _infer_period_from_question(question, payload.get('period') or 'today')
        if not question:
            return jsonify({'error': 'سؤال خالی است'}), 400

        _, facts = _get_intelligence_bundle(doc_type=doc_type, period=period)
        chunks = []
        if len(question) >= 6 and not project_intelligence._is_greeting(question):
            chunks = document_knowledge.search(question)

        activity_df = None
        try:
            activity_df = get_activity_index()
        except Exception as e:
            safe_log(f"[intelligence] خطا در ساخت ایندکس فعالیت: {e}", level="warning")

        knowledge_entries = []
        try:
            knowledge_entries = _apply_standing_instructions()
        except Exception as e:
            safe_log(f"[intelligence] خطا در بارگذاری دانش دستی: {e}", level="warning")

        # اطلاعات مالی فقط اگه کاربر تو لیست دسترسی باشه به دستیار داده می‌شه —
        # همون محدودیتی که رو خودِ صفحه‌ی /finance هم هست، اینجا هم رعایت می‌شه.
        finance_context = None
        contract_search_fn = None
        current_user = session.get('username')
        if _can_view_finance(current_user):
            try:
                finance_context = get_all_finance_summaries()
                import contract_knowledge
                contract_search_fn = contract_knowledge.search
            except Exception as e:
                safe_log(f"[intelligence] خطا در بارگذاری اطلاعات مالی: {e}", level="warning")

        result = project_intelligence.answer_question(
            question, facts, project, document_chunks=chunks,
            activity_df=activity_df, knowledge_entries=knowledge_entries,
            finance_context=finance_context, contract_search_fn=contract_search_fn,
        )
        return jsonify(result)
    except Exception as e:
        safe_log(f"[intelligence] خطا در پرسش: {e}", level="error")
        safe_log(traceback.format_exc(), level="error")
        return jsonify({'error': str(e)}), 500


# ==================== گرم کردن کش (Cache Warm-up) ====================
# محاسبات سنگین یک‌بار در پس‌زمینه انجام می‌شود و نتیجه بدون TTL در Redis
# می‌ماند تا با ری‌استارت Flask یا مرور زمان از بین نرود. باطل شدن فقط
# وقتی است که Excel جدید به Parquet تبدیل شود (نسخه داده عوض شود).
_warm_lock = threading.Lock()
_last_warmed_version = None


def _warm_up_cache_worker():
    global _last_warmed_version

    if not _warm_lock.acquire(blocking=False):
        safe_log("[warm-up] قبلاً در حال اجراست؛ این درخواست نادیده گرفته شد.")
        return

    try:
        data_store.initialize(find_files())
        data, integrated = get_processed_snapshot()
        data_version = data_store.get_data_version()

        project_list = sorted(set(
            str(d.get('project', '')).strip()
            for d in (integrated.get('master_docs') or []) + (integrated.get('vendor_docs') or [])
            if d.get('project') and str(d.get('project')).strip()
        ))
        projects_to_warm = ['همه'] + project_list

        # ۱) اول ترکیب‌های پرکاربرد داشبورد تا کاربر اول منتظر نماند
        try:
            with app.test_client() as client:
                priority = [
                    ('همه', 30, 'اصلی'),
                    ('همه', 30, 'وندور'),
                    ('همه', 7, 'اصلی'),
                    ('همه', 7, 'وندور'),
                    ('همه', 14, 'اصلی'),
                    ('همه', 14, 'وندور'),
                ]
                for prj, period, doc_type in priority:
                    cache_key = f"data:{data_version}:{prj}-{period}-{doc_type}"
                    if cache_get(cache_key) is None:
                        client.get(f'/api/data?project={prj}&period={period}&type={doc_type}')

                for prj in project_list:
                    for doc_type in ['اصلی', 'وندور']:
                        cache_key = f"data:{data_version}:{prj}-30-{doc_type}"
                        if cache_get(cache_key) is None:
                            client.get(f'/api/data?project={prj}&period=30&type={doc_type}')
        except Exception as e:
            safe_log(f"[warm-up] خطا در کش داشبورد: {e}", level="warning")

        # ۲) خلاصه هوشمند روزانه/هفتگی/ماهانه برای مهندسی و وندور
        try:
            import document_knowledge
            document_knowledge.ensure_docs_dir()
            document_knowledge.build_index()
            for intel_type in ['اصلی', 'وندور']:
                _get_intelligence_bundle(force=True, doc_type=intel_type, period='today')
        except Exception as e:
            safe_log(f"[warm-up] خطا در تحلیل هوشمند: {e}", level="warning")

        # ۳) کش کردن اینباکسِ همه‌ی افراد (حالت پیش‌فرض فیلترها)
        try:
            person_df = data_store.get_dataframe('person')
            person_col = find_column(person_df, ['Person Name'])
            if person_col:
                names = [
                    str(n).strip()
                    for n in person_df[person_col].dropna().unique().tolist()
                    if str(n).strip()
                ]
                for name in names:
                    cache_key = f"inbox:{data_version}:{name}:None:None"
                    if cache_get(cache_key) is None:
                        result = _build_inbox_result(data, name, None, None)
                        cache_set(cache_key, result)
        except Exception as e:
            safe_log(f"[warm-up] خطا در کش اینباکس: {e}", level="warning")

        # ۴) کش کردن جدول «مدارک تغییریافته» برای حالت‌های پیش‌فرض
        try:
            for period in ['week', 'month']:
                for project in projects_to_warm:
                    cache_key = f"changed:{data_version}:{period}::{project}:همه:True"
                    if cache_get(cache_key) is None:
                        result = _build_changed_docs_result(
                            data, period, '', project, 'همه', True
                        )
                        cache_set(cache_key, result)
        except Exception as e:
            safe_log(f"[warm-up] خطا در کش جدول تغییرات: {e}", level="warning")

        # ۵) کش کردن رتبه‌بندی رقابتی
        try:
            history_df = data_store.get_dataframe('history')
            if history_df is not None and not history_df.empty:
                history_project_list = _get_project_list(history_df)
                leaderboard_projects = ['همه'] + history_project_list
                for project in leaderboard_projects:
                    leaderboard_key = f"leaderboard_v14:{data_version}:{_get_scoring_config_version()}:{project}:all:all"
                    if cache_get(leaderboard_key) is None:
                        result = _build_leaderboard_result(project, None, None)
                        cache_set(leaderboard_key, result)
        except Exception as e:
            safe_log(f"[warm-up] خطا در کش رتبه‌بندی رقابتی: {e}", level="warning")

        cache_clear_old(data_version)
        _last_warmed_version = data_version
        safe_log("[warm-up] کش داشبورد، تحلیل هوشمند، اینباکس، جدول تغییرات و رتبه‌بندی آماده شد.")
    except Exception as e:
        safe_log(f"[warm-up] خطای کلی: {e}", level="error")
        safe_log(traceback.format_exc(), level="error")
    finally:
        _warm_lock.release()


def start_cache_warm_up():
    threading.Thread(target=_warm_up_cache_worker, daemon=True).start()


# ==================== چک دوره‌ای تغییر فایل (بدون نیاز به دکمه‌ی آپلود) ====================
# چون فایل اکسل به‌صورت دستی جایگزین می‌شه (نه از طریق دکمه‌ی آپلود سایت)،
# این حلقه هر چند دقیقه یک‌بار در پس‌زمینه چک می‌کنه: اگه فایل‌های اکسل
# جدیدتر از parquet موجود باشن، خودش parquet رو می‌سازه و کش رو گرم می‌کنه.
PERIODIC_WARM_UP_INTERVAL_SECONDS = 180  # هر ۳ دقیقه


def _periodic_cache_warm_up_loop():
    while True:
        try:
            data_store.initialize(find_files())
            version = data_store.get_data_version()
            default_key = f"data:{version}:همه-30-اصلی"
            if version != _last_warmed_version or cache_get(default_key) is None:
                _warm_up_cache_worker()
            else:
                safe_log("[warm-up] داده تغییر نکرده؛ کش معتبر است.")
        except Exception as e:
            safe_log(f"[warm-up-loop] خطا: {e}", level="error")
        time.sleep(PERIODIC_WARM_UP_INTERVAL_SECONDS)


def start_periodic_cache_warm_up():
    threading.Thread(target=_periodic_cache_warm_up_loop, daemon=True).start()


# ========== تنظیمات ==========
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
METADATA_PATH = os.path.join(BASE_DIR, "metadata.parquet")
STATIC_DIR = os.path.join(BASE_DIR, "static")
HOST = "0.0.0.0"
PORT = 5000
TITLE = "سرور داشبورد مدیریت مدارک آفام"
VERSION = "1.0.0"
MAX_PARTIAL_RESULTS = 100
# ==================== زمان‌بندی خودکار (APScheduler) ====================
# نکته‌ی مهم: چون سرور با debug=True اجرا می‌شه، Flask یک reloader داره که
# کل فایل رو در یک پردازش «فرزند» دوباره اجرا می‌کنه. بدون این چک، این بخش
# (و در نتیجه scheduler.add_job) دوبار اجرا می‌شد — یک‌بار در پردازش اصلی و
# یک‌بار در پردازش فرزند reloader — و همین باعث می‌شد ایمیل گزارش Distribute
# دقیقاً دوبار ارسال بشه. این چک تضمین می‌کنه فقط یک نسخه از scheduler بالا بیاد.
_should_start_scheduler = (
    os.environ.get('WERKZEUG_RUN_MAIN') == 'true'  # پردازش واقعیِ سرو (زیر reloader)
    or not app.debug  # اجرا بدون debug/reloader (مثلاً روی سرور تولید)
)

if _should_start_scheduler:
    try:
        from apscheduler.schedulers.background import BackgroundScheduler
        from apscheduler.triggers.cron import CronTrigger
        
        scheduler = BackgroundScheduler()
        scheduler.add_job(
            func=send_distribute_report_auto,
            trigger=CronTrigger(day_of_week='sun,tue', hour=9, minute=5),
            id='distribute_report_job',
            replace_existing=True
        )
        scheduler.start()
        
        
    except Exception as e:
        print(f"❌ خطا در راه‌اندازی زمان‌بندی: {e}")
        safe_log(f"❌ خطا در راه‌اندازی زمان‌بندی: {e}")

if __name__ == '__main__':
    safe_log("="*60)
    safe_log("Server started")
    safe_log("http://127.0.0.1:5000")
    safe_log(f"http://10.0.0.171:{PORT}")
    safe_log(f"http://5.160.148.115:{PORT}")
    safe_log("="*60)

    # فقط در پردازش واقعی سرو (نه parentِ reloader) کش را گرم کن
    if _should_start_scheduler:
        start_periodic_cache_warm_up()
    app.run(host='0.0.0.0', port=5000, debug=True)
